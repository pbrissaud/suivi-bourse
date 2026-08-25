"""Tests for the /api blueprint (issue #659, design #655).

The routes themselves are thin — five lines each — so what is tested here is the
**contract**, which is not thin at all: #655 decision 8 split absence into three
states the scheduler's own reads collapse onto ``None``, and getting them back
onto one value is exactly how a UI ends up rendering "the database is dead" and
"you own nothing yet" identically. Since #700 the store is **real** here and the
three states are structural: a declared table nobody wrote answers ``200`` +
``[]``, an unwritten column reads ``NULL``, and only a fault raises.

Plus the one rule of the SPA catch-all that is load-bearing: it must not swallow
an ``/api`` 404.
"""
from datetime import date, datetime, timedelta, timezone

import csv
import io
import json

import openpyxl
import pytest
from apscheduler.schedulers.background import BackgroundScheduler

import accounts as accounts_module
import installation_facts
import entries
import ledger
import main
import perf_series
import portfolio_view
import quotes
import runtime_state
import settings_registry
import store
import web as web_module
from events import EventLoader
from events import export as events_export
from events.schemas import AccountMetricPoint, PortfolioTotalPoint
from web import create_app, problem


class FakeMetrics:
    """Stands in for SuiviBourseMetrics: the replay, and the dials.

    ``ingest`` is real rather than a spy (issue #697). It is the seam
    *"the replay follows the write"* lands on, so a route that forgets an
    import has to be observable through the snapshot it republished — which
    only a real reload gives. The scrape-job reconciliation the production
    method also does needs a scheduler and is not what these tests are about.

    The InfluxDB half is gone with #700: a route reads the **store**, which
    these tests give it for real, so there is nothing left here to stand in for.
    """

    #: The real one, borrowed rather than re-implemented: it is a loop over the
    #: registry and a ``setattr``, so a copy here would be the second list
    #: ADR-0014 forbids — and the one that stops matching first.
    apply_dials = main.SuiviBourseMetrics.apply_dials

    def __init__(self, config_manager=None):
        self._config_manager = config_manager
        self.ingest_calls = 0
        self.apply_dials(settings_registry.defaults())
        # What ``rearm_regular_scrapes`` answers here. Re-arming needs a live
        # jobstore, which these tests have no business standing up; the split
        # itself is pinned in ``test_scheduling.py`` and its wiring in
        # ``test_scheduling_wiring.py``.
        self.rearm_result = (3, 8)
        self.rearm_calls = 0
        # The reconstruction's progress, from the scheduler's own memory
        # (issue #709). A **pair**, always: the real method has no ``None`` to
        # answer, since a process holding a metrics object is a process that can
        # see the memory. ``(0, 0)`` is what these tests' empty ledgers hold —
        # nothing to reconstruct. *Unobservable* is a runtime with no metrics at
        # all (``runtime.metrics is None``), which is what ``with_scheduler``
        # below builds.
        self.reconstruction = (0, 0)
        # What answering the reporting currency triggers (issue #704): the
        # lateral pass is brought forward on the backfill job. Counted rather
        # than performed — the real one moves a live jobstore, which these tests
        # have no business standing up — and answering ``False`` is the honest
        # shape of a runtime with no scheduler.
        self.repair_calls = 0
        self.repair_result = False

    def reconstruction_state(self):
        return self.reconstruction

    def rearm_regular_scrapes(self):
        self.rearm_calls += 1
        return self.rearm_result

    def repair_conversions_now(self):
        self.repair_calls += 1
        return self.repair_result

    def ingest(self, force=False):
        self.ingest_calls += 1
        if self._config_manager is None:
            return
        self._config_manager.replay() if force \
            else self._config_manager.reload()

    def recompute_perf(self):
        """The other half of the replay that follows the write (issue #812).

        A no-op, deliberately. The real one is the whole
        of :meth:`main.SuiviBourseMetrics.update_account_metrics` — a reporting
        currency, a price series, a horizon per account — so a copy of it in
        this class would be a simulation of the product rather than the product.
        What #812 claims is claimed where it can be read off the store:
        ``tests/test_replay_perf.py`` runs these same routes against a **real**
        metrics object and asserts on ``account_metrics`` and
        ``portfolio_totals``, with no tick anywhere.

        Not even a counter, therefore: a call count nothing asserts is evidence
        nobody reads, and it would leave this file green on the day the recompute
        stopped following the write.
        """


def build_client(tmp_path, accounts=None, events=None, seed=None,
                 break_store=False, with_scheduler=True):
    """A Flask test client over a real manager **and a real store**."""
    return build_client_and_store(
        tmp_path, accounts, events, seed, break_store, with_scheduler)[0]


def build_client_and_store(tmp_path, accounts=None, events=None, seed=None,
                           break_store=False, with_scheduler=True):
    """As above, plus the open store so a test can read the rows back.

    ``with_scheduler=False`` leaves ``runtime.metrics`` **unset**, which is the
    shape of a worker whose ``start_runtime`` has not run: it is the one state in
    which the reconstruction is genuinely *unobservable* (issue #709), and it is
    a missing object rather than a metrics object answering ``None``.

    ``accounts`` and ``events`` are **files**, and they are read into the store
    before the first publication rather than by it: the manager scans no
    directory since ADR-0032, so a fixture that wants rows puts them there
    itself. What a route then reads is the store, which is the only thing these
    tests ever assert on.

    They go in **through the roads the product has** since #816 — the accounts
    are declared as the app declares them, the events written by
    :func:`entries.create_many`, which is the function the upload route calls.
    The file stays the fixture's shape because it is what these tests already
    spell, and because an event file is still what an owner hands over; what is
    gone is the folder that read it on its own.

    ``seed`` runs **after** the first publication, and it has to: the replay
    rewrites ``position`` wholesale, so a row laid down before it would be
    swept away by it.

    ``break_store`` drops a table the read layer names. That is a genuine
    storage fault rather than a simulated one — the query raises, the blueprint
    answers ``503``, and nothing in between has to recognise an error message
    (which is exactly what ``_ABSENT_SCHEMA`` used to do, and why it is gone).
    """
    events_dir = tmp_path / 'events'
    events_dir.mkdir(exist_ok=True)
    if accounts is not None:
        (events_dir / 'accounts.csv').write_text(accounts, encoding='utf-8')
    if events is not None:
        (events_dir / '2024.csv').write_text(events, encoding='utf-8')

    # A real store under tmp_path, as ``post_fork`` would have opened (#696).
    # ``/health`` reaches it, and every other route here has to keep answering
    # with one present. Since #697 the ledger lives in it too, so the manager
    # is handed the **same** connection ``start_runtime`` hands it — two files
    # here would mean the routes read a different ledger from the one the
    # snapshot was published from.
    opened = store.open_store(tmp_path / 'store.duckdb')
    if accounts is not None:
        declare_accounts(opened, events_dir / 'accounts.csv')
    if events is not None:
        write_event_file(opened, events_dir / '2024.csv')
    manager = main.ConfigurationManager(config_dir=str(tmp_path),
                                        opened_store=opened)
    runtime = main.Runtime(manager, None)
    runtime.store = opened
    # The first publication, as ``build_runtime`` performs it in the master. It
    # reads the store and nothing else (ADR-0032), so what the fixture wrote
    # above is what a route reading the ledger reads.
    manager.reload()
    if seed is not None:
        seed(opened)
    if break_store:
        opened.execute('DROP TABLE position')
        opened.execute('DROP TABLE account_metrics')
        opened.execute('DROP TABLE portfolio_totals')
        opened.execute('DROP TABLE price_point')
    if with_scheduler:
        runtime.metrics = FakeMetrics(manager)
    return create_app(runtime).test_client(), opened


def _declared_rows(path):
    """The fixture file's ``id,type,label`` rows, read here and nowhere else.

    Reading it is the fixture's own business since ADR-0034: no accounts file
    enters the app any more, so the parser that used to live in :mod:`accounts`
    is gone and what a test writes for its own convenience it also reads.
    """
    with open(path, newline='', encoding='utf-8') as handle:
        return [
            (row['id'].strip(), row['type'].strip(),
             (row.get('label') or '').strip() or row['id'].strip())
            for row in csv.DictReader(handle)
            if (row.get('id') or '').strip()
        ]


def declare_accounts(opened, path):
    """The accounts a fixture wants, declared the way the app declares them.

    Accounts are born in the app (ADR-0034) and no file imports them, so this
    reads the fixture's file and makes the same three calls a reader clicking
    *declare* would. The seeded ``default`` row exists already and is relabelled
    rather than inserted — which is also the one way an install with a page and
    no file declares its single account.
    """
    for account_id, account_type, label in _declared_rows(path):
        if account_id in accounts_module.account_ids(opened):
            opened.execute(
                'UPDATE account SET type = ?, label = ? WHERE id = ?',
                [account_type, label, account_id])
            continue
        accounts_module.create_account(opened, account_id, account_type, label)


def write_event_file(opened, path):
    """One event file into the store, **through the road the upload takes**.

    :func:`entries.create_many` is what ``POST /api/events/import`` calls, and
    the currency question is put by the same function the route puts it to —
    so a fixture cannot write a ledger the product would have refused.
    """
    loader = EventLoader(str(path))
    rows = loader.load()
    entries.create_many(
        opened, rows,
        base_currency=ledger.currency_to_adopt(opened, loader.declared_currency))


# --------------------------------------------------------------------- #
# Seeding — the rows a page reads, written the way the jobs write them
# --------------------------------------------------------------------- #

def seed_position(opened, symbol='AAPL', name='Apple Inc', account='pea',
                  quantity=10.0, cost_basis=1500.0, realized_gain=0.0,
                  received_dividend=0.0):
    """One row of ``position`` — what the replay lays down."""
    opened.execute("INSERT INTO account (id, type, label) VALUES (?, 'CTO', ?) "
                   "ON CONFLICT (id) DO NOTHING", [account, account])
    opened.execute('INSERT INTO symbol (symbol) VALUES (?) '
                   'ON CONFLICT (symbol) DO NOTHING', [symbol])
    opened.execute(
        'INSERT INTO position (account, symbol, name, quantity, cost_basis, '
        '                      realized_gain, received_dividend) '
        'VALUES (?, ?, ?, ?, ?, ?, ?) '
        'ON CONFLICT (account, symbol) DO UPDATE SET '
        '  name = excluded.name, quantity = excluded.quantity, '
        '  cost_basis = excluded.cost_basis, '
        '  realized_gain = excluded.realized_gain, '
        '  received_dividend = excluded.received_dividend',
        [account, symbol, name, quantity, cost_basis, realized_gain,
         received_dividend])


def seed_quote(opened, symbol='AAPL', price=200.0,
               at=datetime(2024, 6, 1, 12, 0, tzinfo=timezone.utc),
               currency='USD', converted=None, rate=1.0, **attributes):
    """One observation — what the scrape writes, through the real writer.

    Converted as well as native (#702). `converted` defaults to the native
    price at a rate of 1: the pages read the converted column, because every
    money figure they draw is in the reporting currency. A point with only a
    native price is one whose conversion has not landed, and the tests that are
    about *that* say so by passing `converted=None, rate=None`.
    """
    opened.execute('INSERT INTO symbol (symbol) VALUES (?) '
                   'ON CONFLICT (symbol) DO NOTHING', [symbol])
    values = {'currency': currency, 'exchange': 'NMS', 'quote_type': 'EQUITY',
              'dividend_yield': 0.5, 'pe_ratio': 30.0, 'market_cap': 3.0e12}
    values.update(attributes)
    quotes.record_quote(opened, symbol, at, price, values,
                        price if converted is None and rate is not None
                        else converted, rate)


def seed_totals(opened, day=date(2026, 8, 5), **overrides):
    """One ``portfolio_totals`` row — the dashboard head's whole source."""
    values = {'cash_balance': 500.0, 'holdings_value': 12000.0,
              'total_value': 12500.0, 'net_contributed': 10000.0,
              'xirr': 0.12, 'gain_absolu': 2500.0, 'twr_index': 124.0}
    values.update(overrides)
    perf_series.write_portfolio_totals(
        opened, [PortfolioTotalPoint(day=day, **values)])


def seed_account_metrics(opened, account='pea', day=date(2026, 8, 5),
                         **overrides):
    """One ``account_metrics`` row, for the accounts comparison table."""
    values = {'cash_balance': 500.0, 'holdings_value': 12000.0,
              'total_value': 12500.0, 'net_contributed': 10000.0,
              'xirr': 0.12, 'gain_absolu': 2500.0, 'twr_index': 118.4}
    values.update(overrides)
    opened.execute("INSERT INTO account (id, type, label) VALUES (?, 'CTO', ?) "
                   "ON CONFLICT (id) DO NOTHING", [account, account])
    perf_series.write_account_metrics(opened, [AccountMetricPoint(
        account=account, account_type='CTO',
        day=day, **values)])


#: A declared setup — the `accounts` mode's precondition. An accounts **file**
#: in the drop folder (issue #698), in the events' own format.
ACCOUNTS_FILE = (
    "id,type,label\n"
    "pea,PEA,PEA Bourso\n"
)

ACCOUNTS_EVENTS = (
    "date,event_type,symbol,name,quantity,unit_price,account\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,pea\n"
)


# --------------------------------------------------------------------- #
# Movers
# --------------------------------------------------------------------- #

def test_movers_anchor_on_the_last_session_close_not_on_midnight_today(tmp_path):
    """#652 déc. 8's trap. The newest stored point is a Friday close; the
    baseline is midnight of *that* day, so a weekend still shows Friday's
    session instead of a column of zeros."""
    def seed(opened):
        seed_position(opened, quantity=10.0, cost_basis=1000.0)
        seed_quote(opened, price=100.0,
                   at=datetime(2026, 7, 30, 20, 0, tzinfo=timezone.utc))
        seed_quote(opened, price=110.0,
                   at=datetime(2026, 7, 31, 21, 0, tzinfo=timezone.utc))

    client = build_client(tmp_path, seed=seed)
    payload = client.get('/api/portfolio/movers').get_json()

    assert payload['since'] == '2026-07-31T00:00:00+00:00'
    # The cut and the session are two different instants, and only the second is
    # safe to put in front of a reader: naming the cut announced a close that
    # had not happened yet.
    assert payload['reference'] == '2026-07-30T20:00:00+00:00'
    assert payload['movers'][0]['change_pct'] == pytest.approx(0.1)
    assert payload['movers'][0]['contribution'] == pytest.approx(100.0)


def test_movers_on_a_fresh_install_is_empty_and_asks_nothing_further(tmp_path):
    """No observation means no session to anchor on, and nothing to compare."""
    client = build_client(tmp_path)
    payload = client.get('/api/portfolio/movers').get_json()

    assert payload == {'since': None, 'reference': None, 'movers': []}


# --------------------------------------------------------------------- #
# The v5 pair the front reads (contract #745, issue #763)
#
# `/api/positions` and `/api/portfolio-totals` were declared in `lib/api.ts`
# before the server served either, so what is pinned here is the **shape** as
# much as the arithmetic: a field the client reads and the server spells
# otherwise is a `404`'s worth of difference to a page that mounts both.
# --------------------------------------------------------------------- #

#: A ledger carrying a transfer fee, which no other fixture here has. The `BUY`
#: fee is the discriminating half: it is absorbed into the cost basis (ADR-0003)
#: and must **not** appear in ADR-0018's fourth term.
FEE_EVENTS = (
    "date,event_type,symbol,name,quantity,unit_price,fee,amount,account\n"
    "2024-01-10,DEPOSIT,,,,,1.50,1000.00,pea\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,3.00,,pea\n"
    "2024-06-01,WITHDRAWAL,,,,,0.75,200.00,pea\n"
)


def test_positions_serves_the_frozen_shape_in_one_read(tmp_path):
    """The whole portfolio in one query, spelled as `lib/api.ts` reads it.

    The two market objects are the shape that carries a distinction a single
    nullable number cannot: `price` is the quote **in its own currency**, and
    `converted` is the same observation in the reporting one, with the rate that
    got it there — so the row can be read back as a journal rather than believed.
    """
    def seed(opened):
        seed_position(opened, account='pea', quantity=10.0, cost_basis=1500.0,
                      realized_gain=25.0, received_dividend=12.5)
        seed_quote(opened, price=200.0, currency='USD', converted=180.0,
                   rate=0.9)
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    payload = build_client(tmp_path, seed=seed).get('/api/positions').get_json()

    assert payload['base_currency'] == 'EUR'
    assert len(payload['positions']) == 1
    row = payload['positions'][0]
    assert set(row) == {'account', 'symbol', 'name', 'quantity', 'cost_basis',
                        'realised', 'dividends', 'price', 'converted',
                        'closed_at', 'fundamentals'}
    assert (row['account'], row['symbol'], row['name']) == (
        'pea', 'AAPL', 'Apple Inc')
    assert row['quantity'] == 10.0 and row['cost_basis'] == 1500.0
    assert row['realised'] == 25.0 and row['dividends'] == 12.5
    assert row['price'] == {'value': 200.0, 'currency': 'USD',
                            'at': '2024-06-01T12:00:00+00:00'}
    assert row['converted'] == {'value': 180.0, 'currency': 'EUR', 'rate': 0.9,
                                'rate_at': '2024-06-01T12:00:00+00:00'}


def test_the_instrument_rides_on_the_holding_and_is_absent_unfetched(tmp_path):
    """The share sheet's fundamentals block, and its two absences (issue #720).

    They ride on the position's row the way ``price`` does — P1 hands them back
    per ``(account, symbol)`` and the front folds a symbol's rows into one line —
    and they are **read, never summed**: holding the same ETF on two accounts
    does not double its market capitalisation.

    The two absences are not one. A ``null`` **member** is the ordinary case,
    yfinance publishing no ``pe_ratio`` for an ETF; the object being ``null`` is
    the symbol the fetch has never reached, and the sheet then draws no block at
    all rather than five em dashes.
    """
    def seed(opened):
        seed_position(opened, symbol='AAPL', account='pea')
        seed_position(opened, symbol='MSFT', name='Microsoft', account='pea')
        seed_quote(opened, symbol='AAPL', price=200.0, pe_ratio=None)

    payload = build_client(tmp_path, seed=seed).get('/api/positions').get_json()
    rows = {row['symbol']: row for row in payload['positions']}

    assert rows['AAPL']['fundamentals'] == {
        'currency': 'USD', 'exchange': 'NMS', 'quote_type': 'EQUITY',
        'dividend_yield': 0.5, 'pe_ratio': None, 'market_cap': 3.0e12,
    }
    assert rows['MSFT']['fundamentals'] is None


def test_a_fundamental_json_cannot_spell_never_reaches_the_body(tmp_path):
    """A ``200`` whose body a browser refuses whole, and how it was found.

    ``store.finite`` was written against the NaN and spelled ``value != value``,
    which tests for a NaN and not for *a number JSON can carry* — so an infinite
    ``trailingPE`` walked through the writer, was stored, and came back out in
    ``fundamentals`` (the member #720 added). ``jsonify`` emits a bare
    ``Infinity`` token, ``JSON.parse`` refuses the **whole** body, every read of
    ``/api/positions`` fails in the client, and the four blocks of the dashboard
    correctly render nothing — a blank app, no console error, no band, because a
    read that has not landed is not a fact.

    The assertion is deliberately on the **bytes** and with ``parse_constant``
    armed, because that is the exact shape of the trap: Python's own parser
    accepts ``Infinity`` and ``NaN`` happily, so a ``curl | python -m json.tool``
    check — and ``response.get_json()`` beside it — passes on a payload no
    browser can read. Only a strict reader sees it.
    """
    def refuse(token):
        raise AssertionError(f'JSON cannot spell {token!r}')

    def seed(opened):
        seed_position(opened, symbol='AAPL', account='pea')
        seed_quote(opened, symbol='AAPL', price=200.0)
        # Written **behind** the writer on purpose. `record_quote` filters what
        # it stores, so going through it would seed a clean row and assert
        # nothing: the row that broke the app was written when the rule did not
        # cover the infinity, and a fundamental is only ever refreshed by a
        # successful fetch — which a sold line never gets (#699). This is that
        # row, and it is what makes the read guard the subject.
        opened.execute("UPDATE symbol_quote SET pe_ratio = 'Infinity'::DOUBLE, "
                       "market_cap = 'NaN'::DOUBLE WHERE symbol = 'AAPL'")

    response = build_client(tmp_path, seed=seed).get('/api/positions')
    assert response.status_code == 200

    payload = json.loads(response.get_data(as_text=True), parse_constant=refuse)
    fundamentals = payload['positions'][0]['fundamentals']
    assert fundamentals['pe_ratio'] is None
    assert fundamentals['market_cap'] is None
    # The object still stands: the currency and the exchange were observed, and
    # an unspellable figure is an absent member, never an absent block.
    assert fundamentals['currency'] == 'USD'


def test_a_sold_line_and_a_never_quoted_one_are_rows_and_never_absences(tmp_path):
    """Both halves of the second criterion, on the same payload.

    A sold position stays in the table (ADR-0017) — its realized gain is the
    figure it has left to say — and a position whose symbol was never fetched is
    a row with `null` market objects, which is P1's LEFT join: an inner one
    answers *"you own nothing"* to somebody who has just declared everything they
    own.
    """
    def seed(opened):
        seed_position(opened, symbol='AAPL', account='pea', quantity=0.0,
                      cost_basis=0.0, realized_gain=180.0)
        seed_position(opened, symbol='MSFT', name='Microsoft', account='pea',
                      quantity=4.0, cost_basis=800.0)
        seed_quote(opened, symbol='AAPL', price=200.0)

    payload = build_client(tmp_path, seed=seed).get('/api/positions').get_json()
    rows = {row['symbol']: row for row in payload['positions']}

    assert set(rows) == {'AAPL', 'MSFT'}
    assert rows['AAPL']['quantity'] == 0.0
    assert rows['AAPL']['realised'] == 180.0
    assert rows['MSFT']['price'] is None
    assert rows['MSFT']['converted'] is None


def test_a_quote_with_no_rate_keeps_its_price_and_loses_its_conversion(tmp_path):
    """*Waiting for a rate* and *no price at all* are two rows, not one.

    The first keeps the quote the reader's broker shows them and has no figure
    in the reporting currency; the second is carried at its cost (ADR-0004).
    """
    def seed(opened):
        seed_position(opened, account='pea')
        seed_quote(opened, price=200.0, currency='USD', converted=None,
                   rate=None)

    payload = build_client(tmp_path, seed=seed).get('/api/positions').get_json()
    row = payload['positions'][0]

    assert row['price']['value'] == 200.0
    assert row['price']['currency'] == 'USD'
    assert row['converted'] is None


#: A ledger where the three shapes `closed_at` has to tell apart coexist: a
#: line emptied in two sales, a line still held, and a line sold and bought
#: back. Written as **events**, because the day the field reports is a fact of
#: the ledger and a seeded `position` row could not carry it.
CLOSING_EVENTS = (
    "date,event_type,symbol,name,quantity,unit_price,account\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,pea\n"
    "2024-02-05,SELL,AAPL,Apple Inc,4,160.00,pea\n"
    "2024-03-10,SELL,AAPL,Apple Inc,6,170.00,pea\n"
    "2024-01-20,BUY,MSFT,Microsoft,5,300.00,pea\n"
    "2024-02-01,BUY,SAN,Sanofi,4,80.00,pea\n"
    "2024-04-01,SELL,SAN,Sanofi,4,90.00,pea\n"
    "2024-05-02,BUY,SAN,Sanofi,2,85.00,pea\n"
)


def test_closed_at_is_the_sale_that_emptied_the_line_and_nothing_else(tmp_path):
    """The member #719's folded section sorts on (issue #763).

    Three rows and three answers in one payload: the **last** sale of a line
    emptied in two goes — not its first — a held line, and a line sold then
    bought back, which is `null` again. The field describes the position's
    *current* state and not its history: a client cannot derive any of it, a
    position carrying a quantity and never the event that emptied it.
    """
    payload = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                           events=CLOSING_EVENTS).get('/api/positions').get_json()
    rows = {row['symbol']: row for row in payload['positions']}

    assert set(rows) == {'AAPL', 'MSFT', 'SAN'}
    assert rows['AAPL']['quantity'] == 0.0
    assert rows['AAPL']['closed_at'] == '2024-03-10'
    assert rows['MSFT']['closed_at'] is None
    # Sold on 2024-04-01 and held again since 2024-05-02: the field says what is
    # true now, and nothing on the wire remembers the line was once flat.
    assert rows['SAN']['quantity'] == 2.0
    assert rows['SAN']['closed_at'] is None


def test_positions_on_a_fresh_install_is_200_and_an_empty_list(tmp_path):
    """The first of the three absences, envelope included."""
    response = build_client(tmp_path).get('/api/positions')

    assert response.status_code == 200
    assert response.get_json() == {'base_currency': None, 'positions': []}


def test_positions_storage_failure_is_503_and_never_an_empty_list(tmp_path):
    """The third absence, and the one that must not look like the first."""
    response = build_client(tmp_path, break_store=True).get('/api/positions')

    assert response.status_code == 503
    assert response.mimetype == 'application/problem+json'
    assert response.get_json()['type'] == '/problems/storage-unavailable'


def test_portfolio_totals_serves_the_eleven_members(tmp_path):
    """Eight columns and three derivations, under the names #745 froze."""
    payload = build_client(
        tmp_path, seed=seed_totals).get('/api/portfolio-totals').get_json()

    assert set(payload) == {'base_currency', 'totals'}
    assert set(payload['totals']) == {
        'day', 'total_value', 'holdings_value', 'cash_balance',
        'net_contributed', 'xirr', 'twr_index', 'twr_since', 'transfer_fees',
        'gain_absolu', 'ytd'}
    assert payload['totals']['day'] == '2026-08-05'
    assert payload['totals']['total_value'] == 12500.0
    # The index counts from the first day of the series, and there is one here.
    assert payload['totals']['twr_since'] == '2026-08-05'
    # An install whose broker moves money for free: a figure worth nothing, not
    # an absence — the page drops it for being zero.
    assert payload['totals']['transfer_fees'] == 0.0


def test_portfolio_totals_with_no_figures_at_all_is_200_and_null(tmp_path):
    """`totals: null`, never a `404` and never `[]`.

    Two causes and one shape: no ledger, or no reporting currency answered.
    """
    response = build_client(tmp_path).get('/api/portfolio-totals')

    assert response.status_code == 200
    assert response.get_json() == {'base_currency': None, 'totals': None}


def test_the_head_keeps_its_positions_while_the_currency_is_unanswered(tmp_path):
    """The direct benefit of ADR-0018, proved across the two resources.

    The perf job writes nothing at all until the reporting currency is answered,
    so `portfolio_totals` is empty — and three of the four terms of the gain are
    read off `/api/positions`, which is under no such constraint. A field absent
    for the global row can no longer blank the headline.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)

    positions = client.get('/api/positions').get_json()
    totals = client.get('/api/portfolio-totals').get_json()

    assert positions['base_currency'] is None
    assert [row['symbol'] for row in positions['positions']] == ['AAPL']
    assert totals['totals'] is None


def test_transfer_fees_are_negative_as_they_enter_the_sum(tmp_path):
    """ADR-0018's fourth term, signed where it is produced (issue #763).

    Naming it as a cost and subtracting it at the point of use is one inversion
    too many for a figure whose entire interest is that the four terms add up.
    And a `BUY`'s fee is **not** in it: that one is absorbed into the cost basis
    (ADR-0003) and would otherwise be counted twice.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE, events=FEE_EVENTS,
                          seed=seed_totals)

    payload = client.get('/api/portfolio-totals').get_json()

    assert payload['totals']['transfer_fees'] == pytest.approx(-2.25)


def test_the_year_to_date_pins_a_positive_gain_against_a_negative_twr(tmp_path):
    """The configuration the pair exists to make readable, to the cent.

    `+40,69 €` of gain against `−1,25 %` of time-weighted return over the same
    period: opposite signs, both correct, because the portfolio grew by 6 673 €
    of deposits while its holdings lost 1,25 %. Subtracting the movement of
    contributions is what makes the first figure a performance rather than a
    receipt for a deposit — without that second term the gain would read
    `+6 713,69 €`, which is the misreading the couple exists to prevent.
    """
    def seed(opened):
        # The close of the previous exercise — the base the delta counts from.
        # `gain_absolu` is the value minus everything contributed — here
        # `total_value − net_contributed`, these rows carrying no valued grant
        # — **on every row the perf job writes**, and since #782 that last
        # clause is true. The field used to land on the last point of the
        # series alone, so this seed described precisely the row the job could
        # not produce and these four tests attested a screen no install could
        # show. What holds the route to the job is the test named for that
        # crossing below, which seeds no `portfolio_totals` row at all; these
        # stay seeded because what they are about is the **bound** — which day
        # is the base — and a ledger long enough to place one is a slow way to
        # ask that question.
        seed_totals(opened, day=date(2025, 12, 31), total_value=10000.00,
                    net_contributed=5000.00, gain_absolu=5000.00,
                    twr_index=160.00)
        seed_totals(opened, day=date(2026, 3, 2), total_value=16713.69,
                    net_contributed=11673.00, gain_absolu=5040.69,
                    twr_index=158.00)

    payload = build_client(
        tmp_path, seed=seed).get('/api/portfolio-totals').get_json()

    assert payload['totals']['day'] == '2026-03-02'
    assert payload['totals']['twr_since'] == '2025-12-31'
    assert payload['totals']['ytd']['gain'] == pytest.approx(40.69, abs=1e-9)
    assert payload['totals']['ytd']['twr'] == pytest.approx(-0.0125, abs=1e-9)


def test_the_year_to_date_base_is_the_close_of_the_previous_year(tmp_path):
    """The bound this ticket had to choose, and the day of market it turns on.

    *The first day on or after 1 January* would take 2026-01-02 as the base and
    leave that day's own move out of both figures; *the last day at or before
    31 December* is a state the measured year has not touched. The series being
    dense over calendar days (#707), both rows exist — so the choice cannot be
    settled by which one is there.
    """
    def seed(opened):
        seed_totals(opened, day=date(2025, 12, 31), total_value=10000.00,
                    net_contributed=5000.00, gain_absolu=5000.00,
                    twr_index=160.00)
        seed_totals(opened, day=date(2026, 1, 2), total_value=12000.00,
                    net_contributed=5000.00, gain_absolu=7000.00,
                    twr_index=176.00)
        seed_totals(opened, day=date(2026, 3, 2), total_value=16713.69,
                    net_contributed=11673.00, gain_absolu=5040.69,
                    twr_index=158.00)

    ytd = build_client(
        tmp_path, seed=seed).get('/api/portfolio-totals').get_json()[
            'totals']['ytd']

    assert ytd['gain'] == pytest.approx(40.69, abs=1e-9)
    assert ytd['twr'] == pytest.approx(-0.0125, abs=1e-9)


def test_the_year_to_date_is_null_only_while_the_series_misses_its_base(tmp_path):
    """The **one** state the reconstruction degrades — never a failed sum."""
    payload = build_client(
        tmp_path,
        seed=lambda opened: seed_totals(opened, day=date(2026, 3, 2)),
    ).get('/api/portfolio-totals').get_json()

    assert payload['totals']['ytd'] is None
    # And everything above it is exact from the first cycle, which is what makes
    # the head *normal* during the twenty-five minutes of a reconstruction.
    assert payload['totals']['total_value'] == 12500.0


def test_a_base_before_the_thirty_first_still_counts(tmp_path):
    """*At or before* 31 December, not *on* it.

    An install whose series stops on the 29th has closed its year there. Asking
    for the 31st exactly would report *the series does not reach the base* about
    a series that plainly does.
    """
    def seed(opened):
        seed_totals(opened, day=date(2025, 12, 29), total_value=10000.00,
                    net_contributed=5000.00, gain_absolu=5000.00,
                    twr_index=160.00)
        seed_totals(opened, day=date(2026, 3, 2), total_value=16713.69,
                    net_contributed=11673.00, gain_absolu=5040.69,
                    twr_index=158.00)

    ytd = build_client(
        tmp_path, seed=seed).get('/api/portfolio-totals').get_json()[
            'totals']['ytd']

    assert ytd['gain'] == pytest.approx(40.69, abs=1e-9)


def test_the_year_to_date_gain_survives_an_install_with_no_cash_event(tmp_path):
    """The ordinary v4 arrival, and the sentence #708 nearly made permanent.

    v4 has no cash events at all, so a ledger imported from one carries no
    `DEPOSIT` and no `WITHDRAWAL` — and since #708 that is exactly the install
    whose `total_value`, `net_contributed` and `twr_index` are `NULL` by the
    per-field rule. Subtracting the movement of two `NULL` columns gave a
    **present** `ytd` object with two `null` members, which the head read as
    *the history is not rebuilt that far back* — under a portfolio whose history
    is complete, permanently, and only for the population the rule was written
    to serve.

    `gain_absolu` is written **always** (ADR-0018) and *is* value minus
    contributions, so the euro figure is not merely rescued here: it is the same
    quantity, computed from the one column that survives. The percentage is not,
    and must not be faked — `twr_index` follows `total_value`, so a `null` there
    is the truth and the head owes it an em dash rather than a sentence.
    """
    def seed(opened):
        seed_totals(opened, day=date(2025, 12, 31), total_value=None,
                    cash_balance=None, net_contributed=None, xirr=None,
                    twr_index=None, holdings_value=10000.00,
                    gain_absolu=5000.00)
        seed_totals(opened, day=date(2026, 3, 2), total_value=None,
                    cash_balance=None, net_contributed=None, xirr=None,
                    twr_index=None, holdings_value=10500.00,
                    gain_absolu=5040.69)

    ytd = build_client(
        tmp_path, seed=seed).get('/api/portfolio-totals').get_json()[
            'totals']['ytd']

    assert ytd is not None
    assert ytd['gain'] == pytest.approx(40.69, abs=1e-9)
    # And the percentage stays absent rather than being invented from the euro.
    assert ytd['twr'] is None


def test_portfolio_totals_storage_failure_is_503_problem_json(tmp_path):
    """A query error **propagates**; it is never rescued into an empty answer."""
    response = build_client(
        tmp_path, break_store=True).get('/api/portfolio-totals')

    assert response.status_code == 503
    assert response.mimetype == 'application/problem+json'
    assert 'portfolio_totals' in response.get_json()['detail']


# --------------------------------------------------------------------- #
# `/api/portfolio-totals/history` — the global series, and the five members
# it shares with one account's own (#721)
# --------------------------------------------------------------------- #

def test_portfolio_totals_history_carries_the_account_history_members(tmp_path):
    """The accounts page rebases this series exactly as it rebases an account's.

    One client shape reads both, which is what keeps the rebasing written once:
    the `Portefeuille` row's `perf` and a `perf` cell above it are the same
    arithmetic over the same window, or the row nobody can check contradicts the
    two rows anybody can.
    """
    def seed(opened):
        seed_totals(opened, day=date(2026, 8, 4), twr_index=120.0)
        seed_totals(opened, day=date(2026, 8, 5), twr_index=124.0)

    payload = build_client(
        tmp_path, seed=seed).get('/api/portfolio-totals/history').get_json()

    assert [point['t'] for point in payload['points']] == [
        '2026-08-04', '2026-08-05']
    assert [point['twr_index'] for point in payload['points']] == [120.0, 124.0]
    assert set(payload['points'][0]) == {
        't', 'cash_balance', 'holdings_value', 'total_value',
        'net_contributed', 'twr_index'}


def test_portfolio_totals_history_is_200_and_empty_before_the_first_cycle(
        tmp_path):
    """The empty-collection state, never an error: an install whose perf cache
    is empty has nothing to compare and is perfectly healthy."""
    response = build_client(tmp_path).get('/api/portfolio-totals/history')

    assert response.status_code == 200
    assert response.get_json()['points'] == []


def test_portfolio_totals_history_keeps_a_null_index_a_null(tmp_path):
    """`twr_index` follows `total_value` (#708), so an install with no cash
    event has none — and a zero here would rebase to a curve at the floor."""
    def seed(opened):
        seed_totals(opened, day=date(2026, 8, 5), total_value=None,
                    cash_balance=None, net_contributed=None, twr_index=None)

    points = build_client(
        tmp_path, seed=seed).get(
            '/api/portfolio-totals/history').get_json()['points']

    assert points[0]['twr_index'] is None
    assert points[0]['total_value'] is None


def test_portfolio_totals_history_rejects_an_inverted_window(tmp_path):
    response = build_client(tmp_path).get(
        '/api/portfolio-totals/history?from=2026-08-05&to=2026-08-01')

    assert response.status_code == 400
    assert response.mimetype == 'application/problem+json'


def test_portfolio_totals_history_storage_failure_is_503_problem_json(tmp_path):
    response = build_client(
        tmp_path, break_store=True).get('/api/portfolio-totals/history')

    assert response.status_code == 503
    assert response.mimetype == 'application/problem+json'


# --------------------------------------------------------------------- #
# `/api/positions/history` — the chart's fallback reading (#727)
# --------------------------------------------------------------------- #

def test_positions_history_is_valuation_versus_investment(tmp_path):
    """The reading the dashboard falls back to, and the two names are the two
    figures: `invested` is what the positions cost, never money the owner put
    in — so the area between the curves is the **latent** gain."""
    events = (
        "date,event_type,symbol,name,quantity,unit_price,fee\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,0\n"
    )

    def seed(opened):
        seed_quote(opened, price=200.0,
                   at=datetime(2024, 6, 1, 17, 0, tzinfo=timezone.utc))

    client = build_client(tmp_path, events=events, seed=seed)
    payload = client.get(
        '/api/positions/history?from=2024-01-01&to=2024-12-31').get_json()

    assert payload['points'] == [{
        't': '2024-06-01', 'value': 2000.0, 'invested': 1500.0}]
    # No `mode`: what this resource answers does not depend on a configuration
    # the caller cannot see.
    assert 'mode' not in payload


def test_positions_history_answers_the_same_body_on_a_declared_install(tmp_path):
    """The whole reason it did not stay `/api/portfolio/history`.

    That route's discriminant was *are accounts declared*, which is not the
    question: an install declaring two accounts and recording no cash event
    answered `accounts` there and got `value`/`contributed` back — two series of
    `null` since #708's per-field rule, i.e. an empty chart on a full portfolio.
    """
    events = (
        "date,event_type,account,symbol,name,quantity,unit_price,fee\n"
        "2024-01-15,BUY,pea,AAPL,Apple Inc,10,150.00,0\n"
    )

    def seed(opened):
        seed_quote(opened, price=200.0,
                   at=datetime(2024, 6, 1, 17, 0, tzinfo=timezone.utc))

    client = build_client(tmp_path, accounts=ACCOUNTS_FILE, events=events,
                          seed=seed)
    payload = client.get(
        '/api/positions/history?from=2024-01-01&to=2024-12-31').get_json()

    assert payload['points'] == [{
        't': '2024-06-01', 'value': 2000.0, 'invested': 1500.0}]


def test_positions_history_is_200_and_empty_on_an_install_with_no_price(tmp_path):
    """`200` + `[]`, never an error: a portfolio nobody has quoted yet has no
    curve and is perfectly healthy."""
    response = build_client(tmp_path).get('/api/positions/history')

    assert response.status_code == 200
    assert response.get_json()['points'] == []


def test_positions_history_rejects_an_inverted_window(tmp_path):
    response = build_client(tmp_path).get(
        '/api/positions/history?from=2026-08-05&to=2026-08-01')

    assert response.status_code == 400
    assert response.mimetype == 'application/problem+json'


def test_positions_history_storage_failure_is_503_problem_json(tmp_path):
    response = build_client(
        tmp_path, break_store=True).get('/api/positions/history')

    assert response.status_code == 503
    assert response.mimetype == 'application/problem+json'


def test_the_four_terms_sum_to_the_absolute_gain_on_a_ledger_with_transfer_fees(
        tmp_path):
    """ADR-0018's identity, **through the API**, on a ledger with a fee (#763).

    The residual risk of deriving `transfer_fees` from `event` while
    `net_contributed` is computed in `performance.py` from the `Timeline`: two
    modules state what a cash movement is, and two statements eventually
    disagree — the symptom being an identity that quietly stops holding, on the
    page that exists to show that it holds. The sign alone is pinned above; this
    is the sum, and it is what would catch the drift.

    Nothing here is seeded: the ledger is a file, the perf cache is written by
    the **real** job, and the four terms are read off the two resources exactly
    as `lib/gain.ts` reads them. On this ledger — deposit 1 000,00 (fee 1,50),
    buy 10 × 80,00 (fee 3,00), dividend 12,00, sale of 4 at 90,00 (fee 2,00),
    withdrawal 200,00 (fee 0,75), quote at 100,00:

        latente +118,20 · réalisée +36,80 · dividendes +12,00 · frais −2,25
                                                          = gain_absolu 164,75
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,fee,amount,account\n"
        "2024-01-10,DEPOSIT,,,,,1.50,1000.00,pea\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,80.00,3.00,,pea\n"
        "2024-03-01,DIVIDEND,AAPL,Apple Inc,,,,12.00,pea\n"
        "2024-04-01,SELL,AAPL,Apple Inc,4,90.00,2.00,,pea\n"
        "2024-06-01,WITHDRAWAL,,,,,0.75,200.00,pea\n"
    )

    def seed(opened):
        seed_quote(opened, price=100.0, currency='EUR', converted=100.0,
                   rate=1.0, at=datetime(2024, 6, 5, 17, 0, tzinfo=timezone.utc))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=events, seed=seed)

    # The perf cache written by the production path, not by `seed_totals`: what
    # is being checked is that the two *modules* agree, so a hand-written
    # `gain_absolu` would check nothing at all.
    metrics = main.SuiviBourseMetrics(web_module.current_runtime().config_manager)
    metrics.base_currency = 'EUR'
    metrics.update_account_metrics()
    assert opened.query('SELECT count(*) FROM portfolio_totals')[0][0] > 0

    positions = client.get('/api/positions').get_json()['positions']
    totals = client.get('/api/portfolio-totals').get_json()['totals']

    # The three position terms, computed as `lib/gain.ts` computes them.
    latent = sum(row['quantity'] * row['converted']['value'] - row['cost_basis']
                 for row in positions)
    realised = sum(row['realised'] for row in positions)
    dividends = sum(row['dividends'] for row in positions)

    assert latent == pytest.approx(118.20, abs=5e-3)
    assert realised == pytest.approx(36.80, abs=5e-3)
    assert dividends == pytest.approx(12.00, abs=5e-3)
    assert totals['transfer_fees'] == pytest.approx(-2.25, abs=5e-3)
    assert totals['gain_absolu'] == pytest.approx(164.75, abs=5e-3)

    # And the identity itself, to the cent — the assertion the two spellings of
    # *a cash movement* have to keep true between them.
    assert (latent + realised + dividends + totals['transfer_fees']
            == pytest.approx(totals['gain_absolu'], abs=5e-3))


def test_the_identity_holds_when_the_dividend_itself_carries_a_fee(tmp_path):
    """The same identity, on the line the fourth term cannot reach.

    ADR-0018's fourth term is named for what a broker takes from a **transfer**,
    and `store_reads.transfer_fees` sums it over `DEPOSIT`/`WITHDRAWAL` alone —
    so a fee on a `DIVIDEND` row belongs to no term, while `_apply_share_cash`
    still takes it out of cash. It therefore landed inside `gain_absolu` and
    inside none of the four, and the head — which *computes* the total from the
    four — disagreed with `portfolio_totals` by exactly the withholding. The
    ledger below is the one above with a 4,00 withholding on the dividend, the
    commonest way this arrives: a PFU deducted at source, typed into `fee`.

        latente +118,20 · réalisée +36,80 · dividendes +8,00 · frais −2,25
                                                          = gain_absolu 160,75
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,fee,amount,account\n"
        "2024-01-10,DEPOSIT,,,,,1.50,1000.00,pea\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,80.00,3.00,,pea\n"
        "2024-03-01,DIVIDEND,AAPL,Apple Inc,,,4.00,12.00,pea\n"
        "2024-04-01,SELL,AAPL,Apple Inc,4,90.00,2.00,,pea\n"
        "2024-06-01,WITHDRAWAL,,,,,0.75,200.00,pea\n"
    )

    def seed(opened):
        seed_quote(opened, price=100.0, currency='EUR', converted=100.0,
                   rate=1.0, at=datetime(2024, 6, 5, 17, 0, tzinfo=timezone.utc))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=events, seed=seed)

    metrics = main.SuiviBourseMetrics(web_module.current_runtime().config_manager)
    metrics.base_currency = 'EUR'
    metrics.update_account_metrics()

    positions = client.get('/api/positions').get_json()['positions']
    totals = client.get('/api/portfolio-totals').get_json()['totals']

    latent = sum(row['quantity'] * row['converted']['value'] - row['cost_basis']
                 for row in positions)
    realised = sum(row['realised'] for row in positions)
    dividends = sum(row['dividends'] for row in positions)

    # The dividend is **net**: 12,00 received, 4,00 withheld.
    assert dividends == pytest.approx(8.00, abs=5e-3)
    assert totals['transfer_fees'] == pytest.approx(-2.25, abs=5e-3)
    assert totals['gain_absolu'] == pytest.approx(160.75, abs=5e-3)
    assert (latent + realised + dividends + totals['transfer_fees']
            == pytest.approx(totals['gain_absolu'], abs=5e-3))


def _fixed_today(mocker, y, mo, d):
    """The perf job's clock, fixed — `main` reads UTC and so does this (#781).

    The job's series runs to *today*, so a test asserting a year-to-date over it
    would otherwise change meaning on 1 January: the base day it counts from is
    read off the calendar, not off the ledger.

    It is `test_cash.py`'s helper, copied rather than hoisted, and that is an
    arbitration rather than an oversight: hoisting it means a fixture, therefore
    a signature edit on twenty-four tests, two of which take `mocker` for
    something else — a change of its own, in a ticket about a column. There is
    no rule to drift here, only a stub with one line in it.
    """
    class _FixedDatetime(datetime):
        @classmethod
        def now(cls, tz=None):
            return datetime(y, mo, d, 12, 0, tzinfo=tz)
    mocker.patch("main.datetime", _FixedDatetime)


def test_the_year_to_date_gain_is_a_figure_when_the_real_job_crosses_a_new_year(
        tmp_path, mocker):
    """The criterion of #782, and the seam nothing in this suite stood on.

    Every other year-to-date test here **seeds** `portfolio_totals`, and a seed
    can lay down a row the job does not produce — which is exactly what happened:
    `gain_absolu` was written on the last point of the series alone, the base row
    is by construction never the last point, and `ytd.gain` was therefore `null`
    on every real install since #763. A seeded row hid it for four tests.

    So nothing is seeded here but the prices and the dial: the ledger is a file,
    the cache is written by the **real** job, and the year-to-date is read off the
    route. On this ledger — deposit 1 000,00 on 2 June 2025, ten shares at 80,00
    the next day, the line quoted 92,00 in November and 96,50 in February:

        base 2025-12-31   1 000 cash-in · 200 cash + 920 holdings → gain +120,00
        latest 2026-03-02 1 000 cash-in · 200 cash + 965 holdings → gain +165,00
                                                    year-to-date  →       +45,00

    The euro figure is the movement of the gain and the percentage the ratio of
    the two indices, over one period and with nothing contributed in between —
    so the two agree here, where #718's measured pair disagreed in sign.
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,fee,amount,account\n"
        "2025-06-02,DEPOSIT,,,,,,1000.00,pea\n"
        "2025-06-03,BUY,AAPL,Apple Inc,10,80.00,,,pea\n"
    )

    def seed(opened):
        for day, price in ((date(2025, 6, 3), 80.00),
                           (date(2025, 11, 3), 92.00),
                           (date(2026, 2, 2), 96.50)):
            seed_quote(opened, price=price, currency='EUR', converted=price,
                       rate=1.0,
                       at=datetime(day.year, day.month, day.day, 17, 0,
                                   tzinfo=timezone.utc))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client = build_client(tmp_path, accounts=ACCOUNTS_FILE, events=events,
                          seed=seed)

    _fixed_today(mocker, 2026, 3, 2)
    metrics = main.SuiviBourseMetrics(web_module.current_runtime().config_manager)
    metrics.base_currency = 'EUR'
    metrics.update_account_metrics()

    totals = client.get('/api/portfolio-totals').get_json()['totals']

    assert totals['day'] == '2026-03-02'
    assert totals['gain_absolu'] == pytest.approx(165.00, abs=5e-3)
    assert totals['ytd']['gain'] == pytest.approx(45.00, abs=5e-3)
    # And the percentage over the same period, from the two base-100 indices:
    # 1 165 / 1 120 − 1, no flow having landed between the two days.
    assert totals['ytd']['twr'] == pytest.approx(1165 / 1120 - 1, abs=1e-9)


def test_the_year_to_date_gain_crosses_the_year_without_a_cash_ledger_too(
        tmp_path, mocker):
    """The population the choice was made for, through the real job (#782, #708).

    A ledger of purchases alone — the ordinary v4 arrival, v4 having no cash
    events — so `total_value`, `net_contributed` and `twr_index` are `NULL` by
    the per-field rule and only `holdings_value` and `gain_absolu` are written.
    That is exactly what departs the two repairs this ticket had to choose
    between: recomposing the year-to-date from `(total_value −
    net_contributed)` touches no writer and gives this install **nothing**,
    permanently. Making the gain a per-day field gives it the euro figure and
    keeps the percentage honestly absent.

    Same three quotes as above and the same two days, the cash ledger removed:
    the gain is `holdings − invested`, `−800 + 920` at the base and `−800 + 965`
    at the latest, so the year-to-date is the same `+45,00 €`.
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,account\n"
        "2025-06-03,BUY,AAPL,Apple Inc,10,80.00,pea\n"
    )

    def seed(opened):
        for day, price in ((date(2025, 6, 3), 80.00),
                           (date(2025, 11, 3), 92.00),
                           (date(2026, 2, 2), 96.50)):
            seed_quote(opened, price=price, currency='EUR', converted=price,
                       rate=1.0,
                       at=datetime(day.year, day.month, day.day, 17, 0,
                                   tzinfo=timezone.utc))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client = build_client(tmp_path, accounts=ACCOUNTS_FILE, events=events,
                          seed=seed)

    _fixed_today(mocker, 2026, 3, 2)
    metrics = main.SuiviBourseMetrics(web_module.current_runtime().config_manager)
    metrics.base_currency = 'EUR'
    metrics.update_account_metrics()

    totals = client.get('/api/portfolio-totals').get_json()['totals']

    assert totals['total_value'] is None
    assert totals['net_contributed'] is None
    assert totals['gain_absolu'] == pytest.approx(165.00, abs=5e-3)
    assert totals['ytd']['gain'] == pytest.approx(45.00, abs=5e-3)
    # And the percentage stays absent: `twr_index` follows `total_value`, so
    # *there is nothing to compute* is the truth here (ADR-0016) and no fifth
    # kind of absence is invented to say it.
    assert totals['ytd']['twr'] is None


# --------------------------------------------------------------------- #
# The chart's own resource — a rung, and the resolution it served (#719, #763)
#
# New, and not a rename of the v4 series route it outlived: that one took
# `?from=`/`?to=` and announced a `bucket`, this one takes a rung of the
# retention ladder and announces a `resolution`. They served side by side until
# the page reading the older one was rewritten.
# --------------------------------------------------------------------- #

def _hour_anchor() -> datetime:
    """The top of the current hour, so a bucket boundary is not a coin toss."""
    return datetime.now(timezone.utc).replace(
        minute=0, second=0, microsecond=0)


def test_the_price_series_announces_the_resolution_it_actually_served(tmp_path):
    """`resolution` is a statement about the answer, never a guess (#719).

    Two windows over the same three points: `1M` is the one rung where *as
    written* is reachable, and `1Y` comes back bucketed by the hour — two
    points, since two of the three share an hour. A reader deducing the
    fineness from the spacing of what came back would read the second as an
    outage; announced once, it is the caption under the chart.
    """
    anchor = _hour_anchor()

    def seed(opened):
        for minutes, price in ((90, 100.0), (70, 101.0), (30, 102.0)):
            seed_quote(opened, price=price, currency='EUR', converted=price,
                       rate=1.0, at=anchor - timedelta(minutes=minutes))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client = build_client(tmp_path, seed=seed)

    written = client.get('/api/prices/AAPL?window=1M').get_json()
    assert written['symbol'] == 'AAPL'
    assert written['base_currency'] == 'EUR'
    assert written['resolution'] == 'raw'
    assert [point['price'] for point in written['points']] == [100.0, 101.0,
                                                               102.0]

    hourly = client.get('/api/prices/AAPL?window=1Y').get_json()
    assert hourly['resolution'] == 'hour'
    # The bucket's survivor is its **last** point, so the 101 wins its hour.
    assert [point['price'] for point in hourly['points']] == [101.0, 102.0]


def test_the_widest_window_is_daily_and_reaches_the_whole_series(tmp_path):
    """`MAX` has no lower bound, and it says `day` because that is what it served.

    A bound computed from a guessed depth would cut an install older than the
    guess; there is none, and the coarsest rung is what a window that may hold
    twenty years of one line can carry.
    """
    def seed(opened):
        seed_quote(opened, price=50.0, currency='EUR', converted=50.0, rate=1.0,
                   at=datetime(2019, 3, 4, 17, 0, tzinfo=timezone.utc))
        seed_quote(opened, price=90.0, currency='EUR', converted=90.0, rate=1.0,
                   at=_hour_anchor() - timedelta(minutes=30))

    client = build_client(tmp_path, seed=seed)

    widest = client.get('/api/prices/AAPL?window=MAX').get_json()
    assert widest['resolution'] == 'day'
    assert [point['price'] for point in widest['points']] == [50.0, 90.0]

    # And the two-year rung stops short of it, which is what makes changing the
    # range change something a reader can see.
    two_years = client.get('/api/prices/AAPL?window=2Y').get_json()
    assert two_years['resolution'] == 'hour'
    assert [point['price'] for point in two_years['points']] == [90.0]


def test_a_point_whose_conversion_never_landed_is_null_and_never_absent(tmp_path):
    """The difference between this resource and the one it stands beside.

    A weekend is a **missing point** (#606); a quote whose rate never resolved
    is a point with **no price**, and it repairs itself when #704's lateral pass
    runs. Collapsing the two would make the chart say the market was closed.
    """
    anchor = _hour_anchor()

    def seed(opened):
        seed_quote(opened, price=100.0, currency='USD', converted=None,
                   rate=None, at=anchor - timedelta(minutes=90))
        seed_quote(opened, price=101.0, currency='USD', converted=95.0,
                   rate=0.94, at=anchor - timedelta(minutes=30))

    payload = build_client(
        tmp_path, seed=seed).get('/api/prices/AAPL?window=1M').get_json()

    assert [point['price'] for point in payload['points']] == [None, 95.0]


def test_an_unknown_window_is_refused_rather_than_defaulted(tmp_path):
    """`422`, and it names the parameter it refused.

    Falling back on a default would serve a curve nobody asked for under a
    caption that describes it correctly — the one failure an announced
    resolution cannot protect anybody from.
    """
    client = build_client(tmp_path)

    for query in ('?window=3M', '?window=', ''):
        response = client.get(f'/api/prices/AAPL{query}')
        assert response.status_code == 422, query
        assert response.mimetype == 'application/problem+json'
        body = response.get_json()
        assert body['key'] == 'window'
        assert '1M' in body['detail']


def test_a_symbol_with_no_stored_price_is_200_and_an_empty_series(tmp_path):
    """An empty series is a state of a fresh install, never a `404`."""
    response = build_client(tmp_path).get('/api/prices/AAPL?window=1Y')

    assert response.status_code == 200
    assert response.get_json() == {'symbol': 'AAPL', 'base_currency': None,
                                   'resolution': 'hour', 'points': []}


def test_the_price_series_propagates_a_storage_failure(tmp_path):
    """A query error is a `503`, never an empty chart."""
    response = build_client(
        tmp_path, break_store=True).get('/api/prices/AAPL?window=1Y')

    assert response.status_code == 503
    assert response.get_json()['type'] == '/problems/storage-unavailable'


# --------------------------------------------------------------------- #
# Accounts — the discriminator, not an empty list
# --------------------------------------------------------------------- #

def test_accounts_says_undeclared_and_still_serves_the_seeded_row(tmp_path):
    """`declared: false` is the opt-out setup every default install runs, and it
    is the **member** that says so — never an empty list.

    Letting the front infer it from `[]` is what would eventually make "no
    declared accounts" and "the config failed to load" render the same screen.
    And the list holds the one account ADR-0013 gives every install: `[]` was a
    resource answering *none* to a question the product says cannot be answered
    that way, which left the declaration block with nothing to render on a fresh
    install — no row to rename, and no way to declare a first account (#729).
    """
    payload = build_client(tmp_path).get('/api/accounts').get_json()

    assert payload['declared'] is False
    assert [a['id'] for a in payload['accounts']] == ['default']
    # The row **as a reader must see it**: what nobody declared is `null`, so the
    # interface names it from its own catalogue rather than recognising the
    # seed's English on the far side of HTTP — see the test below, which is where
    # that rule is guarded.
    seeded = payload['accounts'][0]
    assert (seeded['label'], seeded['type']) == (None, None)
    # And nothing says where the row came from: an account is born in the app
    # (ADR-0034), so the rename is an ordinary `PATCH` with no rule to consult.
    assert not {'source_id', 'editable'} & set(seeded)


def test_renaming_the_seeded_account_is_visible_on_the_resource(tmp_path):
    """The measurement #729 rests on, server-side.

    `PATCH /api/accounts/default` answered `200`, the store held the new label,
    and `GET /api/accounts` went on serving `{declared: false, accounts: []}` —
    so every page rebuilt the row from nothing and re-drew the catalogue's name
    over it. A rename is only a gesture if the resource carries its result, and
    `declared` stays `false`: renaming the row every install owns declares
    nothing beyond it.
    """
    client = build_client(tmp_path)

    renamed = client.patch('/api/accounts/default',
                           json={'label': 'Mon PEA', 'type': 'PEA'})

    assert renamed.status_code == 200
    payload = client.get('/api/accounts').get_json()
    assert payload['declared'] is False
    assert [(a['id'], a['label'], a['type']) for a in payload['accounts']] == [
        ('default', 'Mon PEA', 'PEA')]


def test_the_seed_never_crosses_the_wire_and_the_owners_name_does(tmp_path):
    """What nobody declared goes out as ``null``, and the recognising is here.

    ``store.DEFAULT_ACCOUNT_ROW`` writes ``Default account`` / ``OTHER`` into a
    row every install owns and nobody asked for. The front must not render
    either — they are the *server's* English, and ADR-0024 puts every rendering
    in the reader's language — so one side has to recognise them, and it is the
    side that writes them.

    Recognising them in the client was written first and undone: it put a third
    copy of this string across HTTP, where nothing spans both ends. The front's
    only faked edge is MSW, so its fixtures would have gone on agreeing with
    themselves; reworded here for a typo, the seed would have started rendering
    as a name its owner had typed, with every gate green. This assertion is that
    guard, and it is the one place both halves of the sentence run in one
    process.

    The store is **not** what changes: ``read_accounts`` keeps serving the row as
    written, which is what the export and the replay want. It is the wire that
    carries the declaration alone.
    """
    client = build_client(tmp_path)

    served = client.get('/api/accounts').get_json()['accounts']
    assert [(a['id'], a['label'], a['type']) for a in served] == [
        ('default', None, None)]
    # Read off the constant rather than quoted: this assertion has to fail when
    # the seed is reworded, which is the whole reason it exists.
    _, seeded_type, seeded_label = store.DEFAULT_ACCOUNT_ROW
    assert (seeded_label, seeded_type) not in [
        (a['label'], a['type']) for a in served]

    # An account that is **not** the seeded one keeps whatever it says, the
    # seed's own words included if its owner chose them: what is recognised is
    # one row, never a string.
    client.post('/api/accounts',
                json={'id': 'pea', 'label': seeded_label, 'type': seeded_type})
    named = [(a['id'], a['label'], a['type'])
             for a in client.get('/api/accounts').get_json()['accounts']]
    assert ('pea', seeded_label, seeded_type) in named
    # The seeded row leaves the list here for its own reason and not this one:
    # nothing names it, so ADR-0013 keeps it out of a declaration it did not
    # join (#698). What is asserted above is that its *words* are not the guard.


def test_accounts_returns_the_declaration_with_its_labels(tmp_path):
    """Reading the declaration rather than a DISTINCT on the tag (#652 déc. 4)
    hands over label and type — fields the app writes and zero Grafana panels
    read. An account is declared in the app and nowhere else (ADR-0034), so
    there is nothing beside them saying where the row came from."""
    accounts = (
        "id,type,label\n"
        "pea,PEA,PEA Bourso\n"
    )
    events = (
        "date,event_type,symbol,name,quantity,unit_price,account\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,pea\n"
    )
    client = build_client(tmp_path, accounts=accounts, events=events)
    payload = client.get('/api/accounts').get_json()

    assert payload['declared'] is True
    row = payload['accounts'][0]
    assert (row['id'], row['label'], row['type']) == ('pea', 'PEA Bourso', 'PEA')
    assert not {'source_id', 'editable'} & set(row)
    # #661 enriched the resource with the newest perf figures. With nothing
    # written yet they are all null and `as_of` says so — a declared account
    # whose first perf cycle has not run is a row with em dashes, not a missing
    # line.
    assert row['as_of'] is None
    assert row['total_value'] is None
    assert row['xirr'] is None


def test_accounts_carries_the_newest_figures_of_each_account(tmp_path):
    """The enrichment #661 added: one resource, two consumers.

    `total_value` in particular is written since v4.1 and displayed nowhere —
    the ticket's smallest item and the table's first column.
    """
    client = build_client(tmp_path, seed=seed_account_metrics,
                          accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    row = client.get('/api/accounts').get_json()['accounts'][0]

    assert row['total_value'] == 12500.0
    assert row['gain_absolu'] == 2500.0
    assert row['twr_index'] == 118.4
    assert row['as_of'] == '2026-08-05'
    # The declaration drives the identity fields, and since #700 it is the only
    # thing that could: `account_type` and `account_currency` were InfluxDB
    # tags, and the store has no column for either.
    assert row['type'] == 'PEA'


def test_accounts_keeps_a_declared_account_that_has_no_series_yet(tmp_path):
    """A declared account whose perf cycle has not run is a row of em dashes.

    Not a missing line: the declaration drives the list (#652 déc. 4), so
    absence of data cannot remove an account the human declared.
    """
    accounts = ACCOUNTS_FILE + "cto,CTO,CTO Degiro\n"
    client = build_client(tmp_path, seed=seed_account_metrics,
                          accounts=accounts, events=ACCOUNTS_EVENTS)
    payload = client.get('/api/accounts').get_json()

    # Store order, which is `ORDER BY id` — stable across restarts and across a
    # re-drop, where the file's own order was neither.
    assert [a['id'] for a in payload['accounts']] == ['cto', 'pea']
    assert payload['accounts'][0]['as_of'] is None
    assert payload['accounts'][0]['total_value'] is None


def test_an_absent_field_reaches_the_wire_as_null_and_never_as_a_zero(
        tmp_path, mocker):
    """#708's per-field rule, said on an API response (#806, ADR-0033).

    The rule was written against a *zero*, not against an absence: a zero makes
    *"no cash ledger"* and *"a ledger at zero"* the same figure, so anything
    summing or thresholding the column answers about an account it has never
    been told anything about. It was proved twice — on the store, by the writer's
    own tests, and on `/metrics`, where an absent field meant an absent series.
    The second proof leaves with the exporter; this is the one that replaces it,
    on the surface the front actually reads.

    The population is the ordinary one: a ledger of purchases alone, which is
    every v4 arrival and every owner who only ever recorded buys. `cash_balance`
    would be `−invested` and `net_contributed` `0`, so `total_value` would
    publish the latent gain under the label *total value* — those four and
    `xirr` are `NULL` at the writer and `null` here.

    And the discriminator in the same payload: `transfer_fees` **is** `0.0`. A
    broker that moves money for free is a figure worth nothing, not an absence,
    and the wire tells the two apart.
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,account\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,80.00,pea\n"
    )

    def seed(opened):
        for day, price in ((date(2024, 1, 15), 80.00),
                           (date(2024, 6, 3), 92.00)):
            seed_quote(opened, price=price, currency='EUR', converted=price,
                       rate=1.0,
                       at=datetime(day.year, day.month, day.day, 17, 0,
                                   tzinfo=timezone.utc))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client = build_client(tmp_path, accounts=ACCOUNTS_FILE, events=events,
                          seed=seed)

    _fixed_today(mocker, 2024, 6, 3)
    metrics = main.SuiviBourseMetrics(web_module.current_runtime().config_manager)
    metrics.base_currency = 'EUR'
    metrics.update_account_metrics()

    row = client.get('/api/accounts').get_json()['accounts'][0]

    # The two the ledger can answer exactly: ten shares at 92,00 against 800,00
    # invested, so the gain is `holdings − invested` with no contribution in it.
    assert row['as_of'] == '2024-06-03'
    assert row['holdings_value'] == pytest.approx(920.00, abs=5e-3)
    assert row['gain_absolu'] == pytest.approx(120.00, abs=5e-3)
    # And the five the ledger cannot: `null`, and `is None` is the assertion
    # precisely because a `0.0` would satisfy anything weaker.
    for field in ('cash_balance', 'total_value', 'net_contributed',
                  'xirr', 'twr_index'):
        assert row[field] is None
    assert row['transfer_fees'] == 0.0


def test_accounts_carry_the_fourth_term_of_the_gain_per_account(tmp_path):
    """ADR-0018's identity **per account**, through the API (#722).

    The account's own panel shows `Gain total` dominating its four terms, and
    three of them come off `/api/positions`. The fourth belongs to no position
    at all — it is what a broker takes out of a *transfer* — so the account
    resource derives it the way `/api/portfolio-totals` derives the global one.
    What this pins is that it is **not** the global one repeated: two accounts,
    two deposits, two different fees, and each row's four terms sum to its own
    `gain_absolu`.

    Nothing is seeded but the quote: the ledger is a file and the perf cache is
    written by the real job, so the assertion is that the two *modules* agree —
    `transfer_fees` off `event`, `net_contributed` off the `Timeline`.

        pea  latente +197,00 · réalisée 0,00 · dividendes 0,00 · frais −1,50
        cto  latente  +19,00 · réalisée 0,00 · dividendes 0,00 · frais −0,75
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,fee,amount,account\n"
        "2024-01-10,DEPOSIT,,,,,1.50,1000.00,pea\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,80.00,3.00,,pea\n"
        "2024-01-10,DEPOSIT,,,,,0.75,500.00,cto\n"
        "2024-02-01,BUY,AAPL,Apple Inc,2,90.00,1.00,,cto\n"
    )
    accounts = ACCOUNTS_FILE + "cto,CTO,CTO Degiro\n"

    def seed(opened):
        seed_quote(opened, price=100.0, currency='EUR', converted=100.0,
                   rate=1.0, at=datetime(2024, 6, 5, 17, 0, tzinfo=timezone.utc))
        opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')")

    client, opened = build_client_and_store(
        tmp_path, accounts=accounts, events=events, seed=seed)

    metrics = main.SuiviBourseMetrics(web_module.current_runtime().config_manager)
    metrics.base_currency = 'EUR'
    metrics.update_account_metrics()
    assert opened.query('SELECT count(*) FROM account_metrics')[0][0] > 0

    positions = client.get('/api/positions').get_json()['positions']
    rows = {row['id']: row
            for row in client.get('/api/accounts').get_json()['accounts']}

    # Each account's own fee, and never the −2,25 the global term would carry.
    assert rows['pea']['transfer_fees'] == pytest.approx(-1.50, abs=5e-3)
    assert rows['cto']['transfer_fees'] == pytest.approx(-0.75, abs=5e-3)

    for account, latent in (('pea', 197.00), ('cto', 19.00)):
        held = [row for row in positions if row['account'] == account]
        terms = (
            sum(row['quantity'] * row['converted']['value'] - row['cost_basis']
                for row in held)
            + sum(row['realised'] for row in held)
            + sum(row['dividends'] for row in held)
            + rows[account]['transfer_fees'])
        assert sum(row['quantity'] * row['converted']['value'] - row['cost_basis']
                   for row in held) == pytest.approx(latent, abs=5e-3)
        assert terms == pytest.approx(rows[account]['gain_absolu'], abs=5e-3)


def test_accounts_leave_the_fourth_term_absent_where_no_cycle_wrote_a_day(
        tmp_path):
    """No day to bound the fees by is no coherent statement to make (#722).

    An account whose perf cycle has not run has no `as_of`, so a fee total
    covering *everything* would be a term measured over another period sitting
    in a sum with figures that are all `null` anyway. `null`, and the panel
    drops the term exactly as it drops a zero.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    row = client.get('/api/accounts').get_json()['accounts'][0]

    assert row['as_of'] is None
    assert row['transfer_fees'] is None


def test_accounts_drops_a_series_left_by_an_undeclared_account(tmp_path):
    """Historical residue is not a row. The declaration is the list."""
    def seed(opened):
        seed_account_metrics(opened, account='pea')
        seed_account_metrics(opened, account='old', total_value=999.0)

    client = build_client(tmp_path, seed=seed, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    payload = client.get('/api/accounts').get_json()

    assert [a['id'] for a in payload['accounts']] == ['pea']


# --------------------------------------------------------------------- #
# One account's history — the empty / absent / failed triad (#661)
# --------------------------------------------------------------------- #

def test_account_history_is_200_and_empty_before_the_first_point(tmp_path):
    """Declared, no series yet: the empty-collection state, not an error."""
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    response = client.get('/api/accounts/pea/history')

    assert response.status_code == 200
    payload = response.get_json()
    assert payload['account'] == 'pea'
    assert payload['points'] == []


def test_account_history_returns_the_series_with_its_nulls_intact(tmp_path):
    """Trap 3 on a series: `xirr`/`twr_index` are absent, never zero."""
    def seed(opened):
        seed_account_metrics(opened, day=date(2026, 8, 4), total_value=11500.0,
                             holdings_value=11000.0, twr_index=None, xirr=None,
                             gain_absolu=None)
        seed_account_metrics(opened, day=date(2026, 8, 5))

    client = build_client(tmp_path, seed=seed, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    points = client.get('/api/accounts/pea/history').get_json()['points']

    assert [p['twr_index'] for p in points] == [None, 118.4]
    assert points[1]['total_value'] == 12500.0


def test_an_undeclared_account_history_is_404_problem_json(tmp_path):
    """Decided against the declaration, not against the data: a typo must not
    answer with an empty chart."""
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    response = client.get('/api/accounts/nope/history')

    assert response.status_code == 404
    assert response.mimetype == 'application/problem+json'
    assert response.get_json()['type'] == '/problems/not-found'


def test_account_history_without_any_declaration_is_404(tmp_path):
    """No `accounts:` block at all — every id *but the seeded one* is unknown."""
    response = build_client(tmp_path).get('/api/accounts/pea/history')

    assert response.status_code == 404


def test_the_seeded_account_has_a_history_on_an_install_that_declared_nothing(
        tmp_path):
    """The two resources decide against the same declaration (ADR-0013).

    `list_accounts` falls back to the seeded `default` row when nothing was
    declared — that is what a fresh install shows — while this route decided
    against `_snapshot().accounts`, which is `None` in exactly that state. So
    every id was unknown here, `default` included, and the page asks for the
    history of every row the collection just gave it: a failure band on the
    accounts page of the commonest install there is.
    """
    client = build_client(tmp_path, events=(
        "date,event_type,symbol,name,quantity,unit_price\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,150.00\n"))

    listed = client.get('/api/accounts').get_json()
    assert listed['declared'] is False
    assert [row['id'] for row in listed['accounts']] == ['default']

    response = client.get('/api/accounts/default/history')

    assert response.status_code == 200
    assert response.get_json()['account'] == 'default'


def test_account_history_storage_failure_is_503_problem_json(tmp_path):
    client = build_client(tmp_path, break_store=True,
                          accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    response = client.get('/api/accounts/pea/history')

    assert response.status_code == 503
    assert response.mimetype == 'application/problem+json'
    assert 'account_metrics' in response.get_json()['detail']


def test_a_fault_of_ours_is_a_500_and_not_a_storage_failure(tmp_path, mocker):
    """`503` is a claim about the store, and it was made about every fault.

    The handler answered `storage_unavailable` unconditionally, so a `TypeError`
    in a view module told the client the database was unreachable. The front
    branches on `problem.type` and nothing else (#745), so it drew the *store
    unreachable* screen for a bug of ours — and `503` is deliberately not `500`
    precisely because the condition is meant to be transient, where a `500`
    invites the bug report this one deserved. `problem.internal_error` was
    already written and exported, and had no caller.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    mocker.patch.object(portfolio_view, 'build_positions',
                        side_effect=TypeError('a fault of ours'))

    response = client.get('/api/positions')

    assert response.status_code == 500
    assert response.mimetype == 'application/problem+json'
    assert response.get_json()['type'] == '/problems/internal-error'


def test_a_bare_date_bounds_the_window_in_utc_and_keeps_its_first_day(tmp_path):
    """`?from=2024-06-01` is midnight **UTC**, and the day itself is in.

    Two traps in one assertion, and the route reaches both. `_parse_instant`
    reads a bare date and must stamp it UTC rather than leave it naive for the
    server's own zone to interpret — the front sends bare dates. And the bound
    then lands on a `DATE` column, where `app/CLAUDE.md` names the second trap:
    *a bound on a `DATE` column is cast, or DuckDB widens it to midnight and the
    first day of every window is dropped*.

    It was covered on `/api/shares/<symbol>/prices`, which left with the v4
    routes; `_parse_instant` did not leave with them — it still bounds this
    route, `/api/positions/history` and `/api/portfolio-totals/history`.
    """
    def seed(opened):
        seed_account_metrics(opened, day=date(2024, 6, 1), total_value=100.0)
        seed_account_metrics(opened, day=date(2024, 6, 2), total_value=110.0)

    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS, seed=seed)
    payload = client.get(
        '/api/accounts/pea/history?from=2024-06-01&to=2024-06-03').get_json()

    # The first day of the window is **in**. An hour either way is a
    # server-local reading of the bare date.
    assert [point['t'] for point in payload['points']] == ['2024-06-01', '2024-06-02']

    # And the second trap, which only an instant *inside* the day can spring: a
    # window always arrives as an instant — the route's own default is
    # `now − 365 days`, which is never midnight — while the series is keyed by
    # day. Compared raw, DuckDB widens the `DATE` to midnight and
    # `day >= 2024-06-01T14:23Z` drops 2024-06-01, so the curve silently starts
    # a day after the window it prints in its own `from`.
    inside = client.get(
        '/api/accounts/pea/history?from=2024-06-01T14:23:00Z&to=2024-06-03').get_json()

    assert [point['t'] for point in inside['points']] == ['2024-06-01', '2024-06-02']


def test_account_history_rejects_an_inverted_window(tmp_path):
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    response = client.get(
        '/api/accounts/pea/history?from=2026-06-01&to=2026-01-01')

    assert response.status_code == 400
    assert response.get_json()['type'] == '/problems/bad-request'


# --------------------------------------------------------------------- #
# Events
# --------------------------------------------------------------------- #

def test_an_empty_ledger_is_an_empty_list_not_an_error(tmp_path):
    """A fresh install whose events/ folder is still empty owns nothing yet."""
    response = build_client(tmp_path).get('/api/events')

    assert response.status_code == 200
    assert response.get_json() == []


def test_events_are_returned_without_an_address(tmp_path):
    """#662's opaque id and etag left with the editor (#711).

    They existed because **the file was the address**; nothing here addresses a
    row any more, and no successor was put in their place. The rows are the ones
    the aggregator ran on, in the order it sorted them.
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,amount\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,\n"
        "2024-03-01,DIVIDEND,AAPL,Apple Inc,,,8.50\n"
        "2024-02-01,BUY,MSFT,Microsoft,5,380.00,\n"
    )
    payload = build_client(tmp_path, events=events).get('/api/events').get_json()

    assert [row['date'] for row in payload] == [
        '2024-01-15', '2024-02-01', '2024-03-01']
    assert all(not {'etag', 'file'} & set(row) for row in payload)
    assert payload[0]['event_type'] == 'BUY'
    assert payload[2]['amount'] == 8.50


def test_no_event_carries_a_provenance_on_the_wire(tmp_path):
    """*"row 14 of 2024.csv"* is gone, and so is everything behind it (#816).

    The triplet and the sentence composed from it described a row a **mounted**
    file had provisioned, and they existed because that file was re-read
    (ADR-0032). A file is a payload now, so what the API can say about where a
    row came from is nothing — asserted as an absence on the payload, which is
    the only place a client would have looked.
    """
    events = (
        "date,event_type,symbol,name,quantity,unit_price,amount\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,\n"
        "2024-02-01,BUY,MSFT,Microsoft,5,380.00,\n"
    )
    payload = build_client(tmp_path, events=events).get('/api/events').get_json()

    assert len(payload) == 2
    for row in payload:
        assert set(row).isdisjoint(
            {'source_id', 'source_sheet', 'source_row', 'source_filename',
             'provenance'})


def test_events_can_be_narrowed_to_one_symbol_for_the_chart_markers(tmp_path):
    events = (
        "date,event_type,symbol,name,quantity,unit_price,amount\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,\n"
        "2024-02-01,BUY,MSFT,Microsoft,5,380.00,\n"
    )
    client = build_client(tmp_path, events=events)
    payload = client.get('/api/events?symbol=AAPL').get_json()

    assert [row['symbol'] for row in payload] == ['AAPL']


# --------------------------------------------------------------------- #
# The reassignment (issue #725, ADR-0013, ADR-0006)
#
# **The state is fabricated here and cannot be reached on the real portfolio**,
# whose 285 events all name an account — so `default` is nowhere in it. The
# ticket makes the fixture an obligation rather than a convenience: what it
# guards is an install that ran a month before declaring anything, whose whole
# history sits under the seeded row and whose owner is then locked out of the one
# action that repairs it.
# --------------------------------------------------------------------- #

#: Three events with a **blank** `account` column — legal at the instant they
#: were imported, and stored under `default` by the rule then in force (#698).
UNASSIGNED_EVENTS = (
    "date,event_type,symbol,name,quantity,unit_price,fee,amount,notes\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,2.50,,Initial purchase\n"
    "2024-02-01,BUY,MSFT,Microsoft,5,380.00,2.50,,Initial purchase\n"
    "2024-03-01,DIVIDEND,AAPL,Apple Inc,,,,2.40,Q1 2024 dividend\n"
)


def _accounts_named_by_events(opened):
    return sorted(row[0] for row in opened.query('SELECT account FROM event'))


def test_declaring_the_first_account_reassigns_in_the_same_gesture(tmp_path):
    """One request, and the ledger names the account when it answers."""
    client, opened = build_client_and_store(tmp_path, events=UNASSIGNED_EVENTS)
    assert _accounts_named_by_events(opened) == ['default'] * 3

    created = client.post('/api/accounts',
                          json={'id': 'pea', 'type': 'PEA', 'reassign': True})

    assert created.status_code == 201
    assert _accounts_named_by_events(opened) == ['pea'] * 3
    # And the page reads it: the snapshot the routes serve from was republished
    # by the same request, not by a timer.
    assert [row['account'] for row in client.get('/api/events').get_json()] == [
        'pea'] * 3


def test_the_trap_is_reached_from_the_keyboard_and_repaired_there(tmp_path):
    """The reassignment never had the import for a subject (ADR-0034).

    Its trap — a run of months under the seeded ``default`` row, then a real
    account — is reached by **typing** events exactly as it was by handing over
    a file: what puts a row in the population is its ``account`` column, and a
    typed row leaves that column blank as readily as a file's does. Not one
    event here comes out of a file, and the gesture answers identically.
    """
    client, opened = build_client_and_store(tmp_path)

    for day in ('2024-01-15', '2024-02-01', '2024-03-01'):
        assert client.post('/api/events',
                           json=_draft(date=day)).status_code == 201
    assert _accounts_named_by_events(opened) == ['default'] * 3

    created = client.post('/api/accounts',
                          json={'id': 'pea', 'type': 'PEA', 'reassign': True})

    assert created.status_code == 201
    assert _accounts_named_by_events(opened) == ['pea'] * 3


def test_the_declaration_is_never_refused_because_events_are_unassigned(tmp_path):
    """Refusing is the trap: it locks the owner out of the only repair there is.

    The flag is *offered*, never required — a body that omits it declares the
    account all the same, and leaves the rows exactly where they were.
    """
    client, opened = build_client_and_store(tmp_path, events=UNASSIGNED_EVENTS)

    created = client.post('/api/accounts', json={'id': 'pea', 'type': 'PEA'})

    assert created.status_code == 201
    assert _accounts_named_by_events(opened) == ['default'] * 3
    # And the offer stands afterwards: the window is the state, not the request.
    moved = client.post('/api/accounts/pea/reassignment')
    assert moved.get_json() == {'account': 'pea', 'reassigned': 3}


def test_after_that_instant_an_imported_row_is_not_writable(tmp_path):
    """*Jamais ensuite*, and the ``WHERE`` is what says it.

    A second reassignment moves nothing: the population is the column's own
    value, so once the rows name ``pea`` there is nothing left for it to reach.
    What *is* still reachable is each row on its own — the row-level ``PATCH``
    takes it, which since #816 is true of every row (ADR-0032) — so moving one
    of them back is a correction and never a second reassignment.
    """
    client, opened = build_client_and_store(tmp_path, events=UNASSIGNED_EVENTS)
    client.post('/api/accounts',
                json={'id': 'pea', 'type': 'PEA', 'reassign': True})
    client.post('/api/accounts', json={'id': 'cto', 'type': 'CTO'})

    again = client.post('/api/accounts/cto/reassignment')
    assert again.get_json() == {'account': 'cto', 'reassigned': 0}
    assert _accounts_named_by_events(opened) == ['pea'] * 3

    rows = client.get('/api/events').get_json()
    moved = client.patch(f"/api/events/{rows[0]['id']}",
                         json={'date': '2024-01-15', 'event_type': 'BUY',
                               'symbol': 'AAPL', 'quantity': 10,
                               'unit_price': 150.0, 'account': 'cto'})
    assert moved.status_code == 200
    assert sorted(_accounts_named_by_events(opened)) == ['cto', 'pea', 'pea']


def test_the_seeded_row_is_not_a_target_of_the_reassignment(tmp_path):
    client, opened = build_client_and_store(tmp_path, events=UNASSIGNED_EVENTS)
    client.post('/api/accounts', json={'id': 'pea', 'type': 'PEA'})

    refused = client.post('/api/accounts/default/reassignment')

    assert refused.status_code == 409
    assert _accounts_named_by_events(opened) == ['default'] * 3


def test_an_undeclared_target_is_a_404(tmp_path):
    client = build_client(tmp_path, events=UNASSIGNED_EVENTS)

    assert client.post('/api/accounts/nope/reassignment').status_code == 404


# --------------------------------------------------------------------- #
# The ledger's write path (issue #764, ADR-0005, ADR-0020, ADR-0021)
#
# The population is the whole subject: a row a file provisioned is read-only and
# revoked with its import, a row somebody typed here is reachable by no
# revocation and must therefore be editable. Every test below is about which of
# the two it is looking at.
# --------------------------------------------------------------------- #

#: What the create form sends — its exact shape, with **no ``name``**: a
#: security's name is an attribute of the security and not of each of its
#: events, which is the reason ``Nom`` left the ledger table (ADR-0020).
def _draft(**overrides) -> dict:
    body = {
        'date': '2024-06-03',
        'event_type': 'BUY',
        'account': '',
        'symbol': 'AAPL',
        'notes': 'Typed in the app',
        'quantity': 2,
        'unit_price': 100.0,
        'fee': 1.0,
        'amount': None,
    }
    body.update(overrides)
    return body


def test_every_row_carries_the_key_it_is_addressed_by(tmp_path):
    """``event.id`` reaches the wire, as text, and it is the store's own key."""
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)

    (row,) = client.get('/api/events').get_json()
    stored = [key for (key,) in opened.query('SELECT id FROM event')]

    assert row['id'] == str(stored[0])
    # Text, not a number: a BIGINT above 2^53 is not the integer that was sent,
    # and a client has no arithmetic to do with a key.
    assert isinstance(row['id'], str)


def test_a_typed_event_lands_and_is_visible_at_once(tmp_path):
    """``POST`` writes a row like every other, and the replay follows it.

    Both halves in one test because one without the other is the bug: a row
    written and not replayed is a ledger the app is not computing on, and the
    caller would have to wait for a timer to see their own gesture.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)

    created = client.post('/api/events', json=_draft())
    assert created.status_code == 201
    assert created.get_json()['id'] is not None

    rows = opened.query(
        "SELECT name, account, notes FROM event WHERE date = '2024-06-03'")
    # The name the form never asked for, read off what the ledger already calls
    # this security; the account, blank, resolved to the seeded bucket.
    assert rows == [('Apple Inc', 'default', 'Typed in the app')]

    # Visible in the ledger the app publishes, with no timer in between.
    ledger_rows = client.get('/api/events').get_json()
    assert [row['date'] for row in ledger_rows] == ['2024-01-15', '2024-06-03']
    # And the position the replay wrote counts it.
    assert opened.query(
        "SELECT quantity FROM position WHERE symbol = 'AAPL'") == [(12.0,)]


def test_two_strictly_identical_posts_both_succeed(tmp_path):
    """An order filled twice stays recordable at the keyboard (#813, story 7).

    The criterion is a **negative about the schema** read from the outside: the
    content key the import deduplicates on is declared in no constraint, so the
    same body sent twice lands twice, with two keys and two rows. A `UNIQUE` over
    those eight columns would answer the second one with a `409` nobody asked
    for — which is why the comparison lives at the import and nowhere else.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)
    body = _draft()

    first = client.post('/api/events', json=body)
    second = client.post('/api/events', json=body)

    assert (first.status_code, second.status_code) == (201, 201)
    assert first.get_json()['id'] != second.get_json()['id']
    assert opened.query(
        "SELECT count(*) FROM event WHERE date = '2024-06-03'") == [(2,)]
    # And the replay counted both: ten held before, two typed twice.
    assert opened.query(
        "SELECT quantity FROM position WHERE symbol = 'AAPL'") == [(14.0,)]


def test_a_security_nothing_has_named_yet_is_called_by_its_ticker(tmp_path):
    """The form does not ask for a name, so a first purchase carries the ticker.

    Left ``NULL`` the row would fail ``EventValidator``'s *name is required* on
    the next build — in the gunicorn master, i.e. a boot the owner cannot repair
    from an app that is down.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)

    client.post('/api/events', json=_draft(symbol='MSFT'))

    assert opened.query(
        "SELECT name FROM event WHERE symbol = 'MSFT'") == [('MSFT',)]
    # And the symbol got its row before the event referenced it.
    assert ('MSFT',) in opened.query('SELECT symbol FROM symbol')


def test_a_day_that_does_not_exist_is_refused_by_the_server(tmp_path):
    """``2026-02-31`` has the shape of a day and is not one.

    The rule is only observable from here: ``<input type="date">`` empties its
    own value before any script sees it, so the front measured the trap and
    cannot prove it.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)

    response = client.post('/api/events', json=_draft(date='2026-02-31'))

    assert response.status_code == 422
    assert response.mimetype == 'application/problem+json'
    assert response.get_json()['key'] == 'date'
    assert opened.query('SELECT count(*) FROM event') == [(1,)]


def test_an_instant_is_not_a_calendar_day(tmp_path):
    """The store's two kinds of time never mix, and the boundary is where."""
    response = build_client(tmp_path).post(
        '/api/events', json=_draft(date='2024-06-03T10:00:00Z'))

    assert response.status_code == 422
    assert response.get_json()['key'] == 'date'


def test_a_refused_body_writes_nothing_at_all(tmp_path):
    """``PUT /api/settings``' rule, for the same reason.

    The body below is valid in every member but one, and the one is what the
    validator refuses — a BUY with no unit price. A half-applied body is a state
    nobody asked for.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)

    response = client.post('/api/events', json=_draft(unit_price=None))

    assert response.status_code == 422
    assert response.get_json()['key'] == 'unit_price'
    assert opened.query('SELECT count(*) FROM event') == [(1,)]


def test_the_refusal_is_the_one_validator_s(tmp_path):
    """An event typed here obeys the rules an imported one obeys, word for word.

    The message is ``EventValidator``'s own — the one an accounts-less install
    reads in its logs when a file names an account nobody declared — and the
    field it names travels with it, which is what lets a form mark the input
    instead of printing a paragraph.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)

    response = client.post('/api/events', json=_draft(account='nope'))

    assert response.status_code == 422
    assert response.get_json()['key'] == 'account'
    assert "'nope' is not declared" in response.get_json()['detail']


def test_a_blank_account_is_refused_once_something_is_declared(tmp_path):
    """#698's rule, measured on the road the form actually takes.

    *A blank ``account`` means ``default`` until something is declared, and is
    an error afterwards.* An install declaring ``pea`` must therefore refuse the
    body the form sends with its account left empty — the file road refuses the
    same row whole — or this road quietly grows the phantom ``default`` whose
    figures are all zero, which is exactly what ``declared_portfolio`` exists to
    keep off the page.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    for body in (_draft(), _draft(account='   '), _draft(account=None)):
        response = client.post('/api/events', json=body)
        assert response.status_code == 422
        assert response.mimetype == 'application/problem+json'
        assert response.get_json()['key'] == 'account'

    # Nothing written, and no second account conjured beside the declared one.
    assert opened.query('SELECT count(*) FROM event') == [(1,)]
    assert {row[0] for row in opened.query('SELECT id FROM account')} == {
        'default', 'pea'}


def test_a_declared_account_is_written_as_it_was_named(tmp_path):
    """The refusal above is about the blank and never about naming an account."""
    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    created = client.post('/api/events', json=_draft(account='pea'))

    assert created.status_code == 201
    assert created.get_json()['account'] == 'pea'
    assert opened.query(
        "SELECT account FROM event WHERE date = '2024-06-03'") == [('pea',)]


def test_a_quantity_of_true_is_not_a_quantity(tmp_path):
    """Python calls ``True`` an integer; a ledger does not."""
    response = build_client(tmp_path).post(
        '/api/events', json=_draft(quantity=True))

    assert response.status_code == 422
    assert response.get_json()['key'] == 'quantity'


def test_a_typed_row_is_rewritten_whole_and_then_removed(tmp_path):
    """The two gestures a typo needs, on the row no revocation can reach."""
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)
    key = client.post('/api/events', json=_draft()).get_json()['id']

    rewritten = client.patch(f'/api/events/{key}',
                             json=_draft(quantity=5, notes='Corrected'))
    assert rewritten.status_code == 200
    assert rewritten.get_json()['quantity'] == 5
    assert opened.query(
        "SELECT quantity, notes FROM event WHERE id = ?",
        [int(key)]) == [(5.0, 'Corrected')]

    removed = client.delete(f'/api/events/{key}')
    assert removed.status_code == 200
    assert opened.query('SELECT count(*) FROM event') == [(1,)]
    # The replay followed the removal too: the position is back to the import's.
    assert opened.query(
        "SELECT quantity FROM position WHERE symbol = 'AAPL'") == [(10.0,)]


def test_a_patch_is_a_rewrite_and_never_a_merge(tmp_path):
    """An event's fields are not independent of one another.

    Turning a purchase into a transfer must not leave the purchase's quantity
    standing — the validator refuses a cash event carrying one, so a merge would
    produce a row nobody typed and the boot would meet it.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)
    key = client.post('/api/events', json=_draft()).get_json()['id']

    response = client.patch(f'/api/events/{key}', json={
        'date': '2024-06-04', 'event_type': 'DEPOSIT', 'account': '',
        'symbol': None, 'notes': 'Virement', 'quantity': None,
        'unit_price': None, 'fee': None, 'amount': 500.0,
    })

    assert response.status_code == 200
    assert opened.query(
        "SELECT symbol, quantity, amount FROM event WHERE id = ?",
        [int(key)]) == [(None, None, 500.0)]


def test_an_uploaded_row_is_taken_by_both_row_gestures(tmp_path):
    """**The ticket's seam, on the API** (ADR-0032, #816, stories 13 and 14).

    Both gestures used to answer ``409`` on a row a file had laid down, naming
    the import to forget. A file is handed over once and never re-read now, so
    the argument for that refusal is gone: the row is corrected in place, and
    then removed on its own, and the two rows beside it stay exactly where they
    were.
    """
    client, opened = build_client_and_store(tmp_path, events=_THREE_LINES)
    keys = [key for (key,) in opened.query('SELECT id FROM event ORDER BY id')]
    assert len(keys) == 3

    corrected = client.patch(f'/api/events/{keys[0]}',
                             json=_draft(date='2024-01-15', quantity=12))
    assert corrected.status_code == 200
    assert opened.query(
        'SELECT quantity FROM event WHERE id = ?', [keys[0]]) == [(12.0,)]

    removed = client.delete(f'/api/events/{keys[0]}')
    assert removed.status_code == 200
    assert [key for (key,) in
            opened.query('SELECT id FROM event ORDER BY id')] == keys[1:]


def test_an_unaddressable_event_is_a_named_404(tmp_path):
    """Both shapes of *no such row*, and both answered by this blueprint.

    The route takes a string rather than Flask's ``<int:…>`` converter, so
    ``/api/events/nope`` is problem+json with the id in it instead of the
    router's own bare page.
    """
    client = build_client(tmp_path, events=_ONE_BUY)

    unknown = client.delete('/api/events/9999')
    assert unknown.status_code == 404
    assert unknown.mimetype == 'application/problem+json'

    unaddressable = client.delete('/api/events/nope')
    assert unaddressable.status_code == 404
    assert 'nope' in unaddressable.get_json()['detail']


def test_a_removal_that_would_leave_an_oversell_is_refused(tmp_path):
    """Overselling is a property of the **ledger**, never of a row.

    So a deletion is refused exactly as an insertion is, and the store is never
    left holding a ledger that would fail the next boot.
    """
    client, opened = build_client_and_store(tmp_path)
    bought = client.post('/api/events', json=_draft(quantity=10)).get_json()
    client.post('/api/events',
                json=_draft(date='2024-06-10', event_type='SELL', quantity=10))

    response = client.delete(f"/api/events/{bought['id']}")

    assert response.status_code == 409
    assert opened.query('SELECT count(*) FROM event') == [(2,)]


# --------------------------------------------------------------------- #
# The oversell says its own name (issue #824, ADR-0024)
#
# Four write paths meet a ledger that does not replay, and all four used to
# answer ``/problems/conflict`` — the type whose one sentence was written for
# #698's refusals (*what this names is already there, or something still rests
# on it*) and which describes nothing at all about a file selling shares that
# were never bought. The status is right and stays; the identifier is what the
# front branches on, so the identifier is what changes.
# --------------------------------------------------------------------- #

def test_an_oversell_answers_a_type_of_its_own_and_writes_nothing(tmp_path):
    """The refusal names its subject as **data**, never as prose.

    ``symbol``, ``wanted`` and ``owned`` are extension members — the three facts
    the useful sentence needs — so the front can compose it in the reader's
    language without rendering a word of the server's English. ``detail`` is
    that English, unchanged: it is what a log and a ``curl`` read.
    """
    client, opened = build_client_and_store(tmp_path)
    client.post('/api/events', json=_draft(quantity=10))

    response = client.post('/api/events', json=_draft(
        date='2024-06-10', event_type='SELL', quantity=12, unit_price=180.0))

    assert response.status_code == 409
    assert response.mimetype == 'application/problem+json'
    body = response.get_json()
    assert body['type'] == problem.TYPE_UNREPLAYABLE
    assert body['title'] == 'Ledger does not replay'
    assert (body['symbol'], body['wanted'], body['owned']) == ('AAPL', 12.0, 10.0)
    assert body['day'] == '2024-06-10'
    assert body['detail'] == (
        'Cannot sell 12.0 shares of AAPL (only 10.0 owned) on 2024-06-10')
    # Refused whole: the sale is not half-written and the purchase is untouched.
    assert opened.query('SELECT count(*) FROM event') == [(1,)]


def test_the_refused_gesture_is_named_rather_than_deduced(tmp_path):
    """*Your file oversells* and *what you are taking away is depended on*.

    Two pieces of news, and no payload tells them apart: the oversell reached by
    writing and the one reached by **withdrawing** carry the same three numbers.
    So the route says which it was — ``write`` or ``remove`` — and the front
    selects a sentence on it rather than guessing from the verb it used.
    """
    client, opened = build_client_and_store(tmp_path)
    bought = client.post('/api/events', json=_draft(quantity=10)).get_json()
    client.post('/api/events',
                json=_draft(date='2024-06-10', event_type='SELL', quantity=10))

    # Writing: the purchase edited down under the sale that rests on it.
    edited = client.patch(f"/api/events/{bought['id']}",
                          json=_draft(quantity=1))
    # Removing: the purchase taken away outright, and the same on the bulk
    # gesture, whose subject is a reduction rather than a row.
    removed = client.delete(f"/api/events/{bought['id']}")
    reduced = client.delete('/api/events?type=BUY')

    assert edited.get_json()['gesture'] == 'write'
    assert edited.get_json()['owned'] == 1.0
    assert removed.get_json()['gesture'] == 'remove'
    assert reduced.get_json()['gesture'] == 'remove'
    for response in (edited, removed, reduced):
        assert response.status_code == 409
        assert response.get_json()['type'] == problem.TYPE_UNREPLAYABLE
        assert response.get_json()['symbol'] == 'AAPL'
    assert opened.query('SELECT count(*) FROM event') == [(2,)]


def test_a_declaration_conflict_keeps_the_conflict_type(tmp_path):
    """What this ticket does **not** move.

    An id already taken and an account an event still names are exactly what
    ``/problems/conflict``'s sentence was written for, and it is true of them.
    Only the case it described badly leaves.
    """
    client = build_client(tmp_path, events=ACCOUNTS_EVENTS,
                          accounts=ACCOUNTS_FILE)

    twice = client.post('/api/accounts', json={'id': 'pea', 'type': 'PEA'})
    named = client.delete('/api/accounts/pea')

    for response in (twice, named):
        assert response.status_code == 409
        assert response.get_json()['type'] == problem.TYPE_CONFLICT


def test_a_typed_event_is_exported_like_any_other_row(tmp_path):
    """Provenance is deliberately not exported (issue #710).

    So a row typed here is rendered by exactly the columns an imported one is,
    and the file carries nothing a re-import could read a ``source_id`` out of —
    the export *replaces* the imports it came from, and describing them would
    describe a file the exported one is not.
    """
    client = build_client(tmp_path, events=_ONE_BUY)
    client.post('/api/events', json=_draft(symbol='MSFT'))

    body = client.get('/api/export/events.csv').get_data(as_text=True)
    header, *rows = body.strip().splitlines()

    assert not {'source_id', 'source_row', 'source_sheet', 'source_filename',
                'provenance', 'id'} & set(header.split(','))
    assert len(rows) == 2
    # Same column count on both, the typed one included.
    assert len({len(row.split(',')) for row in rows}) == 1
    assert any(row.startswith('2024-06-03,BUY,default,MSFT,MSFT,') for row in rows)


# --------------------------------------------------------------------- #
# Export (issue #710)
# --------------------------------------------------------------------- #

_EXPORTABLE = (
    "date,event_type,account,symbol,name,quantity,unit_price,fee,amount,notes\n"
    "2024-01-15,BUY,pea,AAPL,Apple Inc,10,150.00,2.50,,Initial purchase\n"
    "2024-03-01,DIVIDEND,pea,AAPL,Apple Inc,,,,8.50,\n"
)

_EXPORTABLE_ACCOUNTS = "id,type,label\npea,PEA,PEA Boursorama\n"


def test_the_export_is_reachable_over_http(tmp_path):
    """The gesture is on the API, so a headless install has it too.

    *Headless means without an interface, not without HTTP* — the same argument
    that keeps ``PUT /api/settings`` a route. One ``curl`` is a complete backup.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_EXPORTABLE,
        seed=lambda opened: opened.execute(
            "INSERT INTO setting (key, value) VALUES ('base_currency', 'EUR')"))

    response = client.get('/api/export/events.csv')

    assert response.status_code == 200
    assert response.headers['Content-Type'] == 'text/csv; charset=utf-8'
    assert 'attachment' in response.headers['Content-Disposition']
    body = response.get_data(as_text=True)
    assert body.splitlines()[0].split(',') == list(events_export.EVENT_COLUMNS)
    # The reporting currency rides on every row, or a round trip reinterprets
    # every amount in the file.
    assert body.count(',EUR\n') == 2
    opened.close()


def test_there_is_no_accounts_file_to_export(tmp_path):
    """The export drops its second file, and the route with it (ADR-0034).

    It is the worst of the residues while it stands: nothing reads an accounts
    file back in since the upload started refusing one, so a
    ``suivi-bourse-accounts.csv`` filed beside the events **looks** like half of
    a round trip without being one — its owner keeps it with their backup and
    believes they can restore from it. A ``404`` is the honest answer, and the
    accounts are redeclared by hand.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_EXPORTABLE)

    assert client.get('/api/export/accounts.csv').status_code == 404
    # And the events' own export is untouched by its going.
    assert client.get('/api/export/events.csv').status_code == 200
    opened.close()


def test_an_empty_ledger_exports_the_header_and_no_row(tmp_path):
    """Emptiness is a state, not an error: the file is valid and carries nothing."""
    client, opened = build_client_and_store(tmp_path)

    body = client.get('/api/export/events.csv').get_data(as_text=True)

    assert body == ','.join(events_export.EVENT_COLUMNS) + '\n'
    opened.close()


_SELECTABLE = (
    "date,event_type,account,symbol,name,quantity,unit_price,fee,amount,notes\n"
    "2024-01-15,BUY,pea,AAPL,Apple Inc,10,150.00,2.50,,Initial purchase\n"
    "2024-03-01,DIVIDEND,pea,AAPL,Apple Inc,,,,8.50,\n"
    "2025-02-02,DEPOSIT,default,,,,,,500.00,Versement de février\n"
)


def test_the_workbook_is_reachable_over_http(tmp_path):
    """The second file the menu offers, and it is a spreadsheet's own shape."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.get('/api/export/events.xlsx')

    assert response.status_code == 200
    assert response.headers['Content-Type'] == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    assert 'suivi-bourse-events.xlsx' in response.headers['Content-Disposition']

    book = openpyxl.load_workbook(io.BytesIO(response.data), data_only=True)
    assert book.sheetnames == ['2024', '2025']
    opened.close()


def test_the_export_takes_the_reduction_the_chips_hold(tmp_path):
    """The four parameters are the table's four, on the resource itself."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    def rows(query):
        body = client.get(f'/api/export/events.csv?{query}').get_data(as_text=True)
        return body.strip().splitlines()[1:]

    assert len(rows('type=BUY')) == 1
    assert len(rows('account=pea')) == 2
    assert len(rows('symbol=AAPL')) == 2
    # Accents folded, as the search field on the table folds them.
    assert len(rows('q=FEVRIER')) == 1
    # Composed: two reductions are an intersection.
    assert rows('type=BUY&account=pea') == rows('type=BUY')
    opened.close()


def test_a_reduction_does_not_take_the_backup_s_name(tmp_path):
    """A partial file is not a backup, and must not replace one on a disk.

    A re-import identifies a source by its **file name**, so a selection saved
    under ``suivi-bourse-events.csv`` would overwrite the whole ledger's export
    in the reader's downloads folder and, dropped back in, replace the import
    that carried every other row.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    whole = client.get('/api/export/events.csv')
    reduced = client.get('/api/export/events.csv?type=BUY')
    workbook = client.get('/api/export/events.xlsx?type=BUY')

    assert 'suivi-bourse-events.csv' in whole.headers['Content-Disposition']
    assert 'suivi-bourse-selection.csv' in reduced.headers['Content-Disposition']
    assert 'suivi-bourse-selection.xlsx' in workbook.headers['Content-Disposition']
    opened.close()


def test_a_parameter_naming_no_type_is_refused_rather_than_served_empty(tmp_path):
    """A backup that silently comes back empty is worse than one that fails.

    ``?type=ACHAT`` is not a ledger this install does not have — it is a word
    the product does not know, and answering it with a valid file holding no row
    would read as *you have recorded nothing of that kind*.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.get('/api/export/events.csv?type=ACHAT')

    assert response.status_code == 422
    assert response.mimetype == 'application/problem+json'
    assert response.get_json()['key'] == 'type'
    opened.close()


def test_a_blank_parameter_is_no_reduction_at_all(tmp_path):
    """``?type=&account=`` is a client with empty fields, not a reduction.

    Read as one, it would answer a file with no row in it under the *selection*
    name — a reader told that a reduction they never made retained nothing.
    ADR-0014's rule about the environment, one level out.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.get('/api/export/events.csv?type=&account=&q=&symbol=')

    assert len(response.get_data(as_text=True).strip().splitlines()) == 4
    assert 'suivi-bourse-events.csv' in response.headers['Content-Disposition']
    opened.close()


def test_a_type_is_read_however_it_is_spelled(tmp_path):
    """``?type=buy`` is the same question as ``?type=BUY``.

    The write path already reads a type case-insensitively
    (``web.api._event_from_body``), and two spellings of one rule eventually
    disagree — here the disagreement would be a ``422`` on a ``curl`` that works
    against every other route.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    lower = client.get('/api/export/events.csv?type=buy')
    upper = client.get('/api/export/events.csv?type=BUY')

    assert lower.get_data() == upper.get_data()
    assert len(lower.get_data(as_text=True).strip().splitlines()) == 2
    opened.close()


def test_a_reduction_that_retains_nothing_is_a_header_and_not_an_error(tmp_path):
    """The empty collection, in a file: valid, and carrying nothing."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.get('/api/export/events.csv?account=zzz')

    assert response.status_code == 200
    assert response.get_data(as_text=True) == \
        ','.join(events_export.EVENT_COLUMNS) + '\n'
    opened.close()


def test_the_export_serves_the_period_bounds_included(tmp_path):
    """The fifth parameter, on both event routes (issue #810).

    Bounds **inclusive**: the ledger is dated to the day (ADR-0008), so
    ``?since=2024-01-15&until=2024-03-01`` is the two days it names and
    everything between them — a half-open reading would drop the last day of
    every year a reader extracts, in a file that looks complete.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.get(
        '/api/export/events.csv?since=2024-01-15&until=2024-03-01')
    rows = response.get_data(as_text=True).strip().splitlines()[1:]

    assert [row.split(',')[0] for row in rows] == ['2024-01-15', '2024-03-01']

    # The workbook takes the same reduction, and its tabs are the years of what
    # is left: the 2025 row is outside the interval, so there is no 2025 sheet.
    workbook = client.get(
        '/api/export/events.xlsx?since=2024-01-15&until=2024-03-01')
    book = openpyxl.load_workbook(io.BytesIO(workbook.data), data_only=True)

    assert book.sheetnames == ['2024']
    opened.close()


def test_one_bound_alone_opens_the_interval_on_the_other_side(tmp_path):
    """*Everything since 2025* is a reduction, and each bound is optional."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    def days(query):
        body = client.get(f'/api/export/events.csv?{query}').get_data(as_text=True)
        return [row.split(',')[0] for row in body.strip().splitlines()[1:]]

    assert days('since=2025-01-01') == ['2025-02-02']
    assert days('until=2024-01-15') == ['2024-01-15']
    opened.close()


def test_a_blank_period_is_no_bound_at_all(tmp_path):
    """``?since=&until=`` is a client with two empty date fields.

    The rule ``?type=&account=`` already follows, and it matters more here: a
    reduction read out of two blanks would name the file a *selection* and hand
    the reader a partial-looking backup of the whole ledger.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.get('/api/export/events.csv?since=&until=')

    assert len(response.get_data(as_text=True).strip().splitlines()) == 4
    assert 'suivi-bourse-events.csv' in response.headers['Content-Disposition']
    opened.close()


def test_a_bound_that_is_not_a_day_is_refused_and_produces_no_file(tmp_path):
    """``?since=hier`` is refused the way ``?type=ACHAT`` is (issue #810).

    Two shapes of the same defect, and the second is the one no browser catches:
    ``2024-02-31`` has the shape of a day and is not one, and a bound silently
    dropped would answer a *file* holding a decade nobody asked for — or missing
    one.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    for query, key in (('since=hier', 'since'),
                       ('until=2024-02-31', 'until'),
                       ('since=20240115', 'since')):
        response = client.get(f'/api/export/events.csv?{query}')

        assert response.status_code == 422
        assert response.mimetype == 'application/problem+json'
        assert response.get_json()['key'] == key
        assert 'Content-Disposition' not in response.headers

    assert client.get(
        '/api/export/events.xlsx?since=hier').status_code == 422
    opened.close()


def test_a_reduction_on_the_period_alone_takes_the_selection_name(tmp_path):
    """An extract of one year is not a backup, and must not be named as one."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    csv_file = client.get('/api/export/events.csv?since=2024-01-01&until=2024-12-31')
    workbook = client.get('/api/export/events.xlsx?until=2024-12-31')

    assert 'suivi-bourse-selection.csv' in csv_file.headers['Content-Disposition']
    assert 'suivi-bourse-selection.xlsx' in workbook.headers['Content-Disposition']
    opened.close()


# --------------------------------------------------------------------- #
# The bulk removal — the reduction is the subject (issue #814, ADR-0032)
# --------------------------------------------------------------------- #

#: Each of the five reductions, and the days the ledger is left holding.
#: Written as *what survives* rather than as *what leaves*: the assertion is on
#: the store's own contents, and a gesture that removed one row too many is
#: only visible from the side that stayed.
_BULK_REDUCTIONS = (
    ('type=BUY', [date(2024, 3, 1), date(2025, 2, 2)]),
    ('account=pea', [date(2025, 2, 2)]),
    ('symbol=AAPL', [date(2025, 2, 2)]),
    # Accents folded, as the search field on the table folds them.
    ('q=FEVRIER', [date(2024, 1, 15), date(2024, 3, 1)]),
    ('since=2025-01-01', [date(2024, 1, 15), date(2024, 3, 1)]),
)


def _bulk_client(tmp_path, name):
    """A client and a store of its own, so five deletions do not share one."""
    room = tmp_path / name
    room.mkdir()
    return build_client_and_store(
        room, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)


def test_the_bulk_delete_takes_the_five_reduction_parameters(tmp_path):
    """``DELETE /api/events`` speaks the export routes' own vocabulary.

    One vocabulary over one contract: the reduction the table shows is the one
    the deletion consumes, so *undo this import* is the chips the reader is
    already looking at rather than a second spelling of them.

    **Every row here came from a file**, which is the other half of the case:
    the predicate *this line was imported* is not consulted by this gesture,
    and one ticket from now it will not exist at all.
    """
    for index, (query, survivors) in enumerate(_BULK_REDUCTIONS):
        client, opened = _bulk_client(tmp_path, f'case{index}')

        response = client.delete(f'/api/events?{query}')

        assert response.status_code == 200
        assert response.get_json() == {'events_removed': 3 - len(survivors)}
        assert [row[0] for row in
                opened.query('SELECT date FROM event ORDER BY date')] == \
            survivors
        opened.close()


def test_a_bulk_delete_with_no_reduction_is_refused_and_writes_nothing(tmp_path):
    """A truncated request must not be able to empty a history (issue #814).

    Emptying the whole ledger stays possible — by reducing on something that
    covers all of it, and therefore deliberately. A client that forgot its query
    string is not that.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.delete('/api/events')

    assert response.status_code == 422
    assert response.mimetype == 'application/problem+json'
    assert opened.query('SELECT count(*) FROM event') == [(3,)]
    opened.close()


def test_blank_parameters_are_no_reduction_and_are_refused_too(tmp_path):
    """``?type=&account=&since=`` is a client with empty fields.

    Blank counts as absent on this resource — ADR-0014's rule one level out, and
    the one the export already follows — so a form submitted with nothing in it
    reaches exactly the refusal an empty query string reaches, rather than
    deleting the ledger it retained by reducing on nothing.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.delete('/api/events?q=&type=&account=&symbol=&since=&until=')

    assert response.status_code == 422
    assert opened.query('SELECT count(*) FROM event') == [(3,)]
    opened.close()


def test_a_bulk_delete_that_retains_nothing_removes_nothing_and_is_no_error(
        tmp_path):
    """Zero is a state, exactly as an export of no row is a valid file."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    response = client.delete('/api/events?account=zzz')

    assert response.status_code == 200
    assert response.get_json() == {'events_removed': 0}
    assert opened.query('SELECT count(*) FROM event') == [(3,)]
    opened.close()


def test_a_bound_that_is_not_a_day_is_refused_by_the_bulk_delete_too(tmp_path):
    """The parameters are read by one function, so they refuse by one sentence.

    ``?since=hier`` names no interval, and a deletion answered under it would be
    a perimeter nobody can state — worse here than on the export by exactly the
    difference between a wrong file and a missing history.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_SELECTABLE)

    for query, key in (('type=ACHAT', 'type'), ('since=hier', 'since'),
                       ('until=2024-02-31', 'until')):
        response = client.delete(f'/api/events?{query}')

        assert response.status_code == 422
        assert response.get_json()['key'] == key
    assert opened.query('SELECT count(*) FROM event') == [(3,)]
    opened.close()


def test_a_bulk_delete_that_would_leave_an_oversell_is_refused_whole(tmp_path):
    """A reduction can take the purchases away and leave the sales.

    ``DELETE /api/events/<id>``'s ``409`` on a wider perimeter, and it rolls the
    whole reduction back: a ledger committed half-deleted raises on every
    reload, and that raise is fatal in the gunicorn master.
    """
    client, opened = build_client_and_store(tmp_path)
    client.post('/api/events', json=_draft(quantity=10))
    client.post('/api/events',
                json=_draft(date='2024-06-10', event_type='SELL', quantity=10))

    response = client.delete('/api/events?type=BUY')

    assert response.status_code == 409
    assert opened.query('SELECT count(*) FROM event') == [(2,)]
    opened.close()


def test_the_bulk_delete_reaches_an_uploaded_row(tmp_path):
    """*The removal is the gesture* (ADR-0032), on a row a file laid down.

    What replaces losing ``forget_import`` has to reach those rows, and it does
    so without asking any of them where they came from — which since #816 is not
    a restraint it shows but a question nothing can ask.
    """
    client, opened = build_client_and_store(tmp_path, events=_ONE_BUY)

    removed = client.delete('/api/events?symbol=AAPL')

    assert removed.status_code == 200
    assert removed.get_json() == {'events_removed': 1}
    assert opened.query('SELECT count(*) FROM event') == [(0,)]
    opened.close()


def test_an_unreadable_store_fails_the_workbook_too(tmp_path):
    """Same contract as the CSV: a query error is a ``503``, never a file."""
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_EXPORTABLE)
    opened.execute('DROP TABLE event')

    assert client.get('/api/export/events.xlsx').status_code == 503
    opened.close()


def test_an_unreadable_store_fails_the_export_rather_than_emptying_it(tmp_path):
    """A backup that silently comes back empty is worse than one that fails.

    Same contract as every other route in this blueprint: a query error
    propagates and becomes a ``503``, never an empty collection.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=_EXPORTABLE_ACCOUNTS, events=_EXPORTABLE)
    opened.execute('DROP TABLE event')

    response = client.get('/api/export/events.csv')

    assert response.status_code == 503
    opened.close()


# --------------------------------------------------------------------- #
# The catch-all
# --------------------------------------------------------------------- #
#
# The bundle's location is resolved from the package and reads no environment
# variable since #740 — the environment says four things and this was never one
# of them. The seam is therefore the resolver itself rather than
# ``SB_STATIC_DIR``, which is what a test seam should have been all along: the
# variable existed for an operator who does not exist.

def _serve_bundle_from(monkeypatch, directory):
    """Point the SPA resolver at ``directory`` for the duration of a test."""
    monkeypatch.setattr(web_module, '_static_dir', lambda: directory)


def test_the_spa_catch_all_does_not_swallow_an_api_404(tmp_path):
    """Without this guard a typo'd endpoint returns the HTML shell with a 200,
    and the front fails on a JSON parse error far from the cause — the most
    confusing failure this arrangement can produce."""
    response = build_client(tmp_path).get('/api/does-not-exist')

    assert response.status_code == 404
    assert response.mimetype == 'application/problem+json'


def test_the_catch_all_does_not_serve_html_at_metrics(tmp_path, monkeypatch):
    """Found by an existing #651 test the moment the SPA landed.

    The endpoint is gone (ADR-0033) and this is the seam where its absence
    could be hidden: an unknown path is answered with the shell, so a leftover
    scraper would read **200 with HTML** and keep reporting nothing wrong.
    The bundle is present on purpose — that is the install the owner runs, and
    the one where the catch-all is in the way.
    """
    bundle = tmp_path / 'bundle'
    bundle.mkdir()
    (bundle / 'index.html').write_text(
        '<!doctype html><div id=spa-shell>', encoding='utf-8')
    _serve_bundle_from(monkeypatch, bundle)

    client = build_client(tmp_path)

    assert client.get('/metrics').status_code == 404
    assert client.get('/metrics/').status_code == 404
    # And it is the shell specifically that must not be what came back: a 404
    # carrying the SPA would still leave the front routing on a path that has
    # no page, which is the same lie one status code later.
    assert b'spa-shell' not in client.get('/metrics').data


def test_health_still_wins_over_the_catch_all(tmp_path):
    """The container healthcheck's only target (#651) must not become the SPA."""
    response = build_client(tmp_path).get('/health')

    assert response.status_code == 200
    # It is the route and not the shell: the shell has no ``jobs`` (#818).
    assert set(response.get_json()['jobs']) == {
        'scrape', 'backfill', 'performance'}


def test_a_build_without_a_bundle_says_so_instead_of_404ing_blankly(tmp_path,
                                                                    monkeypatch):
    _serve_bundle_from(monkeypatch, tmp_path / 'no-bundle-here')
    response = build_client(tmp_path).get('/')

    assert response.status_code == 404
    assert 'API' in response.get_json()['detail']


def test_the_spa_is_served_for_an_unknown_client_route(tmp_path, monkeypatch):
    bundle = tmp_path / 'bundle'
    bundle.mkdir()
    (bundle / 'index.html').write_text('<!doctype html><div id=root>', encoding='utf-8')
    _serve_bundle_from(monkeypatch, bundle)

    response = build_client(tmp_path).get('/titres/AAPL')

    assert response.status_code == 200
    assert b'id=root' in response.data


# --------------------------------------------------------------------------- #
# The configuration resource
# --------------------------------------------------------------------------- #
#
# #662's write path — the row edit, the opaque token, the ETag, the 409, and
# `PUT /api/accounts` — is gone with `config_writer.py` and `events/editor.py`
# (#711), and so are the six problem types that only it raised. What is left of
# this resource is a read and the one toggle that was never a file.

LEDGER_CSV = (
    "date,event_type,symbol,name,quantity,unit_price,fee,amount,notes,account\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,2.50,,Initial purchase,\n"
)


def ledger_client(tmp_path):
    return build_client(tmp_path, events=LEDGER_CSV)


def test_a_malformed_body_is_400_not_503(tmp_path):
    """The blueprint's catch-all renders every exception as a storage 503.

    Flask's own `get_json` raises on a bad body, so without `silent=True` a
    client that sent broken JSON would be told the database is down.
    """
    response = ledger_client(tmp_path).put(
        '/api/config/log-level', data='{not json',
        content_type='application/json')

    assert response.status_code == 400


def test_the_config_route_carries_the_declared_portfolio(tmp_path):
    """What the ledger *declares*, as opposed to what has been observed of it.

    No `mode`, no `editable`, no `read_only_reason`: there is one loading path
    and no write path left for either to be about.
    """
    body = ledger_client(tmp_path).get('/api/config').get_json()

    assert body['shares'][0]['symbol'] == 'AAPL'
    assert 'mode' not in body
    assert 'editable' not in body
    assert 'read_only_reason' not in body


def test_the_log_level_toggle_answers_with_what_it_set(tmp_path):
    client = ledger_client(tmp_path)
    try:
        assert client.put('/api/config/log-level',
                          json={'level': 'debug'}).get_json() == {'log_level': 'DEBUG'}
        assert client.put('/api/config/log-level',
                          json={'level': 'nope'}).status_code == 400
    finally:
        main.set_log_level('INFO')


def test_the_file_era_routes_are_gone_rather_than_refusing(tmp_path):
    """Demolition, not a stub: #662's file gestures do not exist (issue #711).

    ``/api/events/files`` was the file import and conversion, and
    ``PUT /api/accounts`` was the settings block written back — both existed
    because **the file was the address**. Neither has a successor.

    The row-level writes on ``/api/events`` are **not** in that list any more
    (issue #764): they came back for every row there is, and they refuse nothing
    for its origin — see
    ``test_an_uploaded_row_is_taken_by_both_row_gestures``.

    **The two import routes join the list at #816** (criterion 3, ADR-0032).
    ``GET /api/imports`` listed the sources and ``DELETE /api/imports/<id>``
    revoked one; nothing persists that could be named any more, so they are
    demolished rather than answering an empty collection — which would be a
    resource claiming to exist.
    """
    client = ledger_client(tmp_path)

    assert client.get('/api/events/files').status_code == 404
    assert client.put('/api/accounts', json={'accounts': []}).status_code == 405
    assert client.get('/api/imports').status_code == 404
    # ``405`` and not ``404`` for the same reason ``PUT /api/accounts`` gets one,
    # one line up: the catch-all takes the path and not the verb. What both
    # answers say is *no such route*.
    assert client.delete('/api/imports/1').status_code == 405


# --------------------------------------------------------------------------- #
# The app's own runtime state (issue #668, design #656)
#
# One case here is specific to this resource and is the reason it exists: it
# must answer 200 while the store is unreadable. That is decision 6's whole
# point, and it is the one thing a test can prove that looking cannot — on a
# healthy stack the two designs are indistinguishable.
# --------------------------------------------------------------------------- #

def test_the_runtime_resource_answers_200_with_the_store_unreadable(tmp_path):
    """#656 decision 6, and the reason #659's `status` slot was retired.

    `/api/positions` is a query and this blueprint answers 503 when one fails,
    so a pill riding on that payload would **disappear exactly when it is the
    only thing able to explain the empty table** — #655's error contract turned
    against itself one storey up, and worse than the original, because the
    diagnostic dies with what it diagnoses.
    """
    client = build_client(
        tmp_path, break_store=True,
        accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    # The table is dead...
    assert client.get('/api/positions').status_code == 503
    # ...and the thing that explains why is not.
    response = client.get('/api/runtime')

    assert response.status_code == 200
    assert response.get_json()['symbols'][0]['symbol'] == 'AAPL'


def test_the_runtime_says_whether_the_reconstruction_still_has_windows(tmp_path):
    """`rebuilding` (contract #745, issue #763), read from process memory.

    It is on the app-state resource and not beside the figures because it is a
    fact about *this process*, and what it decides on screen is one thing: the
    time-weighted return carries its base date **only while that date is still
    moving**. It says nothing about the year-to-date, which becomes computable
    long before the reconstruction concludes and has its own carrier.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = web_module.current_runtime()

    # Nothing to reconstruct is an observation, and it is not a rebuild.
    runtime.metrics.reconstruction = (0, 0)
    assert client.get('/api/runtime').get_json()['rebuilding'] is False

    runtime.metrics.reconstruction = (1, 2)
    assert client.get('/api/runtime').get_json()['rebuilding'] is True

    runtime.metrics.reconstruction = (2, 2)
    assert client.get('/api/runtime').get_json()['rebuilding'] is False


def test_a_runtime_with_no_scheduler_does_not_claim_a_rebuild(tmp_path):
    """A process that cannot see the scheduler asserts nothing about it.

    A boolean has no room for #709's third answer, and `false` is the safe
    reading: what the member enables is the *claim* that a date is still moving.
    """
    client = build_client(tmp_path, with_scheduler=False,
                          accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    assert client.get('/api/runtime').get_json()['rebuilding'] is False


def test_the_runtime_publishes_the_mount_observation_to_the_front(tmp_path):
    """`store.persistence` (issue #741, ADR-0015) — *the fact is published for
    the front by the same path as the rest of the runtime state*.

    Here rather than on a resource of its own, for the reason that put
    `rebuilding` here: it is a fact about *this process* and its mount
    namespace, answered from memory with no query. Which is also why the
    assertion above matters more than it looks — a store that cannot be opened
    is exactly when *"where did my data go"* gets asked, and this route is the
    one still answering.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)
    runtime = web_module.current_runtime()
    runtime.store_path = tmp_path / 'store.duckdb'
    path = str(tmp_path / 'store.duckdb')

    # A test runtime observed nothing, and says so rather than claiming a kept
    # store.
    assert client.get('/api/runtime').get_json()['store'] == {
        'persistence': 'unknown', 'path': path}

    runtime.store_persistence = 'ephemeral'
    assert client.get('/api/runtime').get_json()['store'] == {
        'persistence': 'ephemeral', 'path': path}

    runtime.store_persistence = 'persistent'
    assert client.get('/api/runtime').get_json()['store'] == {
        'persistence': 'persistent', 'path': path}


def test_the_mount_observation_survives_an_unreadable_store(tmp_path):
    """The fact stays readable on the one failure that empties every page. It is
    observed once in the master and carried in memory, so nothing about it goes
    through the store it describes."""
    client = build_client(
        tmp_path, break_store=True,
        accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    web_module.current_runtime().store_persistence = 'ephemeral'

    assert client.get('/api/positions').status_code == 503
    assert client.get('/api/runtime').get_json()['store']['persistence'] \
        == 'ephemeral'


# --------------------------------------------------------------------- #
# The store block (issue #724, spec #695 § 10)
# --------------------------------------------------------------------- #

def test_the_store_resource_states_its_size_and_its_last_ledger_write(tmp_path):
    """The two figures the block leads with, and what each one is *not*.

    ``size_bytes`` is the file plus its write-ahead log, and it is published
    rather than hidden: hiding it removes only its explanation — ``du`` still
    finds the number — and the explanation is what stops the purge button beside
    it reading as a way to get bytes back.

    ``ledger_last_write`` is when the **ledger** last moved, and never the newest
    observed price. The second is liveness and belongs to the banner; shown here
    it would make a store whose last write was a year ago read as freshly
    written. It was ``max(import_source.imported_at)`` while a file was a row;
    the writer stamps the instant since #816, so a correction and a deletion
    move it too — which the old query, being about imports, never did.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    body = client.get('/api/store').get_json()

    assert body['size_bytes'] > 0
    first = body['ledger_last_write']
    assert first is not None

    ((key,),) = opened.query('SELECT id FROM event ORDER BY id LIMIT 1')
    assert client.delete(f'/api/events/{key}').status_code == 200

    after = client.get('/api/store').get_json()['ledger_last_write']
    assert after is not None and after >= first
    assert body['orphans'] == []


def test_a_store_with_no_import_has_no_last_write_rather_than_a_zero(tmp_path):
    """``null`` and never an epoch. A fresh install has never been written to,
    which is a state and not a date at the beginning of time — and the block
    renders the absence rather than *1 January 1970*."""
    client = build_client(tmp_path)

    assert client.get('/api/store').get_json()['ledger_last_write'] is None


def test_an_orphan_is_named_with_the_series_it_holds(tmp_path):
    """The symbols nothing declares any more, kept **deliberately** (#695 § 10).

    Forgetting an import is reversible — re-drop the file — while a
    reconstructed price series is not, so the app never throws one away by
    itself. What it owes in exchange is that they be named and purgeable on
    demand, and the count is what makes the gesture answerable *before* it is
    made.
    """
    def seed(opened):
        seed_quote(opened, symbol='ZZORPHAN', price=10.0,
                   at=datetime(2024, 1, 1, 12, 0, tzinfo=timezone.utc))
        seed_quote(opened, symbol='ZZORPHAN', price=11.0,
                   at=datetime(2024, 1, 2, 12, 0, tzinfo=timezone.utc))

    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS, seed=seed)

    assert client.get('/api/store').get_json()['orphans'] == [
        {'symbol': 'ZZORPHAN', 'points': 2}]


def test_a_sold_position_is_not_an_orphan(tmp_path):
    """The predicate is *no event names it*, never *its quantity is zero*.

    A line the owner closed years ago still has every one of its events in the
    ledger, so it is still declared — and offering to purge it would offer to
    throw away the history of a position whose realised gain the product still
    shows. ADR-0003 spent a table on that distinction; this is the one place it
    could quietly be lost.
    """
    sold = ('date,event_type,account,symbol,name,quantity,unit_price\n'
            '2024-01-02,BUY,pea,AAPL,Apple Inc,10,150.00\n'
            '2024-06-02,SELL,pea,AAPL,Apple Inc,10,180.00\n')
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE, events=sold)

    assert client.get('/api/store').get_json()['orphans'] == []


def test_the_purge_removes_rows_and_says_how_many(tmp_path):
    """It answers **rows**, never bytes.

    Measured on a real store: 79 % of its rows purged for zero bytes returned —
    126,0 Mo before, 126,0 Mo after, the same content rebuilt from scratch
    fitting in 26,0. DuckDB reuses its blocks. The figure the API can honestly
    report is the one it removed, and the sentence beside the button is the rest.
    """
    def seed(opened):
        seed_quote(opened, symbol='ZZORPHAN', price=10.0)

    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS, seed=seed)

    body = client.delete('/api/store/orphans').get_json()

    assert body == {'symbols': ['ZZORPHAN'], 'points_removed': 1}
    # The symbol row goes with its series: a symbol surviving its own prices
    # would read as a line the app still knows about and can no longer chart.
    assert opened.query(
        "SELECT count(*) FROM symbol WHERE symbol = 'ZZORPHAN'")[0][0] == 0
    assert opened.query(
        "SELECT count(*) FROM price_point WHERE symbol = 'ZZORPHAN'")[0][0] == 0
    # And the declared symbol is untouched.
    assert opened.query(
        "SELECT count(*) FROM symbol WHERE symbol = 'AAPL'")[0][0] == 1
    assert client.get('/api/store').get_json()['orphans'] == []


def test_purging_nothing_is_a_success_and_not_a_refusal(tmp_path):
    """An empty purge is a legitimate answer, the way an import carrying no
    event is still an import. The list is absent from the page at zero, so this
    is what a second click on a stale page gets."""
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE,
                          events=ACCOUNTS_EVENTS)

    response = client.delete('/api/store/orphans')

    assert response.status_code == 200
    assert response.get_json() == {'symbols': [], 'points_removed': 0}


def test_the_store_resource_fails_with_the_store_it_describes(tmp_path):
    """A ``503``, deliberately — and it is the split with ``/api/runtime``.

    Everything on this resource needs the file, so it cannot claim to describe
    an installation it can no longer read. The two facts that *do* survive that
    failure — the path and its persistence — are on the runtime, which touches
    nothing at all, which is why the block reads two resources rather than one.
    """
    client = build_client(tmp_path, break_store=True,
                          accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    assert client.get('/api/store').status_code == 503
    assert client.get('/api/runtime').status_code == 200


def test_the_runtime_resource_issues_no_query_at_all(tmp_path, mocker):
    """Not merely tolerant of a dead store — it never asks it anything.

    Decision 4 makes that free: the backfill job already *reads* the oldest
    stored point, so it *remembers* it, and the progress bar survives an
    unreadable store rather than merely degrading politely.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    queried = mocker.spy(opened, 'query')
    arrow = mocker.spy(opened, 'arrow')

    assert client.get('/api/runtime').status_code == 200

    assert queried.call_count == 0
    assert arrow.call_count == 0


def test_a_never_scraped_symbol_is_a_row_rather_than_a_missing_line(tmp_path):
    """The row set comes from the configuration snapshot (#656 déc. 3).

    Driving it from the recorder would race the scrape threads *and* drop this
    row; driving it from the declaration gives the honest answer for free — the
    position exists, nothing has been observed about it yet.
    """
    client = build_client(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)

    body = client.get('/api/runtime').get_json()

    assert [s['symbol'] for s in body['symbols']] == ['AAPL']
    assert body['symbols'][0]['pill'] == 'unknown'
    # The accounts are a plain list of names since #700: nothing per-account is
    # measured about a series that has no account dimension.
    assert body['symbols'][0]['accounts'] == ['pea']


def test_a_sold_line_publishes_its_backward_progress_and_says_it_is_not_polled(
        tmp_path):
    """The reconstruction the owner is waiting for is the sold one (issue #703).

    The backfill is driven by the replay now, so a closed position has a
    backward pass of its own. Its row was filtered out of this payload while the
    two sets were one; leaving it out today hides the only progress there is to
    show — and the banner's bar would count 1 series out of 1 with a second one
    still walking back through five years.
    """
    events = ACCOUNTS_EVENTS + "2025-02-03,SELL,AAPL,Apple Inc,10,180.00,pea\n"
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=events)
    runtime = web_module.current_runtime()
    runtime.recorder.record_backfill(runtime_state.BackfillRecord(
        symbol='AAPL', direction=runtime_state.BACKWARD,
        at=datetime(2026, 8, 5, 15, 0, tzinfo=timezone.utc),
        target=datetime(2024, 1, 15, tzinfo=timezone.utc),
        oldest=datetime(2024, 6, 1, tzinfo=timezone.utc), written=42))

    body = client.get('/api/runtime').get_json()
    row = body['symbols'][0]

    assert row['symbol'] == 'AAPL'
    assert row['held'] is False
    assert row['next_run_state'] == 'not_held'
    assert row['backward']['written'] == 42
    assert body['backfill']['in_scope'] == 1


def test_the_runtime_publishes_an_unconvertible_series_with_its_reason(tmp_path):
    """#704's terminal, on the route that answers while nothing else can.

    A conversion that will never resolve is the one backfill state that asks the
    owner to *do* something, so it travels with the pair it is about — a state
    word alone would leave a reader in front of an empty column with no
    explanation. It is a **reply** and not a failure, so it stays out of the
    errors list.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = web_module.current_runtime()
    runtime.recorder.record_backfill(runtime_state.BackfillRecord(
        symbol='AAPL', direction=runtime_state.LATERAL,
        at=datetime(2026, 8, 5, 15, 0, tzinfo=timezone.utc),
        terminal=runtime_state.TERMINAL_UNCONVERTIBLE,
        reason='no exchange rate exists between XYZ and EUR (XYZEUR=X)'))

    body = client.get('/api/runtime').get_json()
    lateral = body['symbols'][0]['lateral']

    assert lateral['state'] == 'unconvertible'
    assert 'XYZEUR=X' in lateral['reason']
    assert lateral['error'] is None
    assert body['errors'] == []
    # And it is not counted as a reconstruction that finished.
    assert body['backfill']['complete'] == 0


def test_the_runtime_publishes_the_perf_horizon_of_each_account(tmp_path):
    """Criterion 9 of #708, and it obeys this route's one rule.

    The horizon is computed inside the recompute and written down nowhere — the
    rows say where a series *starts*, which is a different question the moment an
    account's first activity is later than its horizon. So it rides on the perf
    record, and the route reads it from **process memory** like everything else
    here: it answers while the store is unreadable, which is exactly when *"the
    page is filling in towards the left"* is the thing a reader needs told.

    The shape is the one ``lib/api.ts`` announced before either half was written
    — ``accounts: [{ account, horizon }]``, beside ``symbols`` — the same way
    ``rebuilding`` and ``/api/portfolio-totals`` arrived.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = web_module.current_runtime()
    runtime.recorder.record_perf(runtime_state.PerfRecord(
        at=datetime(2026, 8, 5, 15, 0, tzinfo=timezone.utc),
        verdict=runtime_state.PERF_RAN,
        horizons={'pea': date(2021, 6, 3), 'cto': None}))

    body = client.get('/api/runtime').get_json()

    assert body['accounts'] == [{'account': 'cto', 'horizon': None},
                                {'account': 'pea', 'horizon': '2021-06-03'}]


def test_a_published_record_reaches_the_payload_with_its_pill(tmp_path):
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = web_module.current_runtime()
    runtime.recorder.record_scrape(runtime_state.ScrapeRecord(
        symbol='AAPL', at=datetime(2026, 8, 5, 15, 0, tzinfo=timezone.utc),
        market_state='CLOSED', closed=True, price_present=True,
        verdict=runtime_state.SCRAPE_CLOSED, failure_count=0,
        next_delay=3600.0))

    body = client.get('/api/runtime').get_json()

    assert body['symbols'][0]['pill'] == 'closed'
    assert body['symbols'][0]['market_state'] == 'CLOSED'
    assert body['symbols'][0]['last_pass'] == '2026-08-05T15:00:00+00:00'
    # No scheduler in a test process, and that is not trap 1's ambiguity.
    assert body['symbols'][0]['next_run_state'] == 'unavailable'


# --------------------------------------------------------------------- #
# Health — the two registers, on the app over a real store (issue #818)
# --------------------------------------------------------------------- #

HEALTH_PASS = datetime(2026, 8, 5, 15, 0, tzinfo=timezone.utc)


def _arm_the_scheduler(runtime):
    """Give the runtime a scheduler, so ``/health`` reads a worker, not a master.

    A **real** ``BackgroundScheduler``, never started: it holds no thread until
    ``start()``, and what the health body asks of it — *is there one, and what
    is armed* — an empty pending set answers truthfully. Standing in for it
    would have been standing in for the very thing the field is about.
    """
    runtime.scheduler = BackgroundScheduler()
    return runtime


def test_health_names_each_job_with_its_last_pass_and_its_verdict(tmp_path):
    """Criterion 1 of #818: the body is for a person, and it carries the jobs.

    Three jobs and one word for the whole. The material is process memory —
    the recorder's last-pass records — so this answers on a store that has just
    been proved to answer and asks it nothing more.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = _arm_the_scheduler(web_module.current_runtime())
    runtime.recorder.record_scrape(runtime_state.ScrapeRecord(
        symbol='AAPL', at=HEALTH_PASS, market_state='REGULAR', closed=False,
        price_present=True, verdict=runtime_state.SCRAPE_WROTE,
        failure_count=0, next_delay=120.0, wrote=True))
    runtime.recorder.record_backfill(runtime_state.BackfillRecord(
        symbol='AAPL', direction=runtime_state.BACKWARD, at=HEALTH_PASS,
        target=datetime(2024, 1, 15, tzinfo=timezone.utc),
        terminal=runtime_state.TERMINAL_COMPLETE))
    runtime.recorder.record_perf(runtime_state.PerfRecord(
        at=HEALTH_PASS, verdict=runtime_state.PERF_RAN))

    response = client.get('/health')
    body = response.get_json()

    assert response.status_code == 200
    assert body['status'] == 'ok'
    assert body['jobs']['scrape'] == {
        'status': 'ok', 'at': '2026-08-05T15:00:00+00:00', 'verdict': 'open',
        'held': 1, 'attention': []}
    assert body['jobs']['backfill']['verdict'] == 'complete'
    assert body['jobs']['backfill']['complete'] == 1
    assert body['jobs']['backfill']['in_scope'] == 1
    assert body['jobs']['performance'] == {
        'status': 'ok', 'at': '2026-08-05T15:00:00+00:00', 'verdict': 'ran',
        'error': None}


def test_a_frozen_scrape_leaves_the_code_at_200_and_is_read_in_the_body(
        tmp_path):
    """**The assertion that holds the whole decision of the two registers.**

    A writer that fetches happily and persists a value that never moves is what
    #628's sonde flags on the scrape record. Restarting the container repairs
    nothing yfinance or the market broke, so the orchestrator must be told
    nothing at all — and the owner must be told everything. Amber with a
    ``200``, and both halves are read here: the status code, and the payload.

    There is a status code to read, so there is **no spy to write**: the trap
    #804 names by hand is assuring this by checking that some function was not
    called.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = _arm_the_scheduler(web_module.current_runtime())
    runtime.recorder.record_scrape(runtime_state.ScrapeRecord(
        symbol='AAPL', at=HEALTH_PASS, market_state='REGULAR', closed=False,
        price_present=True, verdict=runtime_state.SCRAPE_WROTE,
        failure_count=0, next_delay=120.0, wrote=True, stale=True))

    response = client.get('/health')
    body = response.get_json()

    assert response.status_code == 200
    assert body['status'] == 'attention'
    assert body['jobs']['scrape']['status'] == 'attention'
    assert body['jobs']['scrape']['verdict'] == 'frozen'
    # And named, because *which line to go and read* is the one thing a `curl`
    # cannot follow up on from a count.
    assert body['jobs']['scrape']['attention'] == ['AAPL']


def test_a_wedged_backfill_is_read_in_the_body_and_never_in_the_code(tmp_path):
    """The same rule, on the second job: a stuck pass is not a restart.

    ``failures`` is the backfill's own consecutive counter, folded by the
    recorder — the piece #656 grew out of, since ``_backfill_backward`` logs a
    warning and returns ``0``, leaving nothing to tell *pacing normally* from
    *wedged on yfinance*.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = _arm_the_scheduler(web_module.current_runtime())
    for _ in range(3):
        runtime.recorder.record_backfill(runtime_state.BackfillRecord(
            symbol='AAPL', direction=runtime_state.BACKWARD, at=HEALTH_PASS,
            target=datetime(2024, 1, 15, tzinfo=timezone.utc),
            failed=True, error='yfinance answered nothing'))

    response = client.get('/health')
    body = response.get_json()

    assert response.status_code == 200
    assert body['jobs']['backfill']['verdict'] == 'failing'
    assert body['jobs']['backfill']['attention'] == ['AAPL']


def test_a_failed_perf_pass_is_read_in_the_body_with_what_it_raised(tmp_path):
    """The third job, and the one whose record is global.

    The error rides on the job rather than in a list because there is one of it:
    the record holds the last pass only, so it disappears the moment the next
    cycle succeeds.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = _arm_the_scheduler(web_module.current_runtime())
    runtime.recorder.record_perf(runtime_state.PerfRecord(
        at=HEALTH_PASS, verdict=runtime_state.PERF_FAILED,
        error='the replay raised'))

    response = client.get('/health')
    body = response.get_json()

    assert response.status_code == 200
    assert body['status'] == 'attention'
    assert body['jobs']['performance']['verdict'] == 'failed'
    assert body['jobs']['performance']['error'] == 'the replay raised'


def test_a_container_that_has_observed_nothing_is_unknown_and_not_well(tmp_path):
    """The boot window, said as itself — the whole included.

    A container that started ninety seconds ago has run no scrape, no backfill
    cycle and no perf cycle. *Nothing is wrong* and *nothing is known* are two
    different sentences, and answering ``ok`` to the second is how a page ends up
    announcing a reconstruction that has not started as one that finished. The
    code is the other register and it does not move: ``unknown`` is a ``200``,
    because there is nothing here a restart repairs.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    _arm_the_scheduler(web_module.current_runtime())

    response = client.get('/health')
    body = response.get_json()

    assert response.status_code == 200
    assert body['status'] == 'unknown'
    assert [job['status'] for job in body['jobs'].values()] == [
        'unknown', 'unknown', 'unknown']
    assert body['jobs']['scrape']['at'] is None
    assert body['jobs']['performance']['verdict'] == 'unknown'


def test_a_first_pass_takes_the_whole_out_of_unknown_into_well(tmp_path):
    """And the boot window closes on the first observation, not on the last.

    ``unknown`` is *this process has seen nothing*; one job that ran and nothing
    asking to be looked at is ``ok``, although the two other cycles are still
    ahead of their first pass.
    """
    client, _ = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = _arm_the_scheduler(web_module.current_runtime())
    runtime.recorder.record_scrape(runtime_state.ScrapeRecord(
        symbol='AAPL', at=HEALTH_PASS, market_state='REGULAR', closed=False,
        price_present=True, verdict=runtime_state.SCRAPE_WROTE,
        failure_count=0, next_delay=120.0, wrote=True, stale=False))

    response = client.get('/health')
    body = response.get_json()

    assert response.status_code == 200
    assert body['status'] == 'ok'
    assert body['jobs']['scrape']['status'] == 'ok'
    assert body['jobs']['backfill']['status'] == 'unknown'


def test_the_health_code_is_the_store_and_the_store_alone(tmp_path):
    """Criterion 2 of #818: the predicate of the code did not move.

    The body grew; what makes the answer a ``503`` is what #696 settled — the
    store does not answer — and it is still a ``problem+json``, which is the
    contract every other failing route on this socket obeys.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    runtime = _arm_the_scheduler(web_module.current_runtime())
    # A record a reader would have wanted — and it goes with the store, which
    # is the trade ADR-0036 makes by name: the colour that survives is red.
    runtime.recorder.record_perf(runtime_state.PerfRecord(
        at=HEALTH_PASS, verdict=runtime_state.PERF_RAN))
    opened.close()

    response = client.get('/health')

    assert response.status_code == 503
    assert response.mimetype == 'application/problem+json'


def test_the_health_body_issues_no_query_at_all(tmp_path, mocker):
    """Beyond the ``ping`` the code register runs, the body asks nothing.

    The same rule ``/api/runtime`` keeps, and here it is what keeps the two
    registers apart: a body built from the store would fail for reasons the
    status code has just finished saying the process does not have.
    """
    client, opened = build_client_and_store(
        tmp_path, accounts=ACCOUNTS_FILE, events=ACCOUNTS_EVENTS)
    _arm_the_scheduler(web_module.current_runtime())
    queried = mocker.spy(opened, 'query')
    arrow = mocker.spy(opened, 'arrow')

    assert client.get('/health').status_code == 200

    # One query, and it is the probe's own read of a declared table.
    assert queried.call_count == 1
    assert arrow.call_count == 0


def _dials(client):
    """The dials `/api/config` publishes, by key.

    By key and never by position: the registry's order is not a contract, and a
    test that indexes `[0]` starts asserting about a different dial the day one
    is inserted above it — silently, and still green.
    """
    return {row['key']: row for row in client.get('/api/config').get_json()['settings']}


def test_the_config_route_carries_what_the_process_had_to_know_first(
        tmp_path, monkeypatch):
    """#654's one survivor, on `/api/config` rather than `/api/runtime`.

    #661's argument: one noun, two consumers. "What is this container running?"
    is a question about the configuration, and the data page is already the
    screen that asks it — while putting it on the runtime resource would start
    that resource down the road to a junk drawer.

    What is left in the list is ADR-0014's line: what the process must know
    **before** it can open the store. The three ``INFLUXDB_*`` names left with
    the database (#700) and are reported as set-and-unread instead, which is
    what stops an operator concluding their token is wrong.
    """
    monkeypatch.setenv('INFLUXDB_TOKEN', 'apiv3_supersecret')
    client = build_client(tmp_path)

    body = client.get('/api/config').get_json()
    env = {s['name']: s for s in body['environment']}

    assert 'INFLUXDB_TOKEN' not in env
    assert 'INFLUXDB_TOKEN' in body['unread_environment']
    assert env['SB_STORE_DIR']['value'] is not None
    # Compose-only variables are not app settings (#654 trap 13).
    assert 'SB_VERSION' not in env


def test_the_config_route_lists_the_dials_through_the_registry(tmp_path):
    """#701: the content of ``setting``, and not a second enumeration of it.

    The bounds ride along, so the form validates on the rule the write path
    enforces rather than on a copy of it — and a dial added to the registry
    appears here with no route to edit.
    """
    body = build_client(tmp_path).get('/api/config').get_json()

    dials = {row['key']: row for row in body['settings']}

    assert list(dials) == [spec.key for spec in settings_registry.SETTINGS]
    assert dials['regular_interval']['value'] == 120
    assert dials['regular_interval']['minimum'] == 10
    assert dials['regular_interval']['effect'] == settings_registry.REARM_SCRAPE
    # A dial has no environment form at all, not even a reported one.
    assert 'SB_REGULAR_INTERVAL' not in {
        row['name'] for row in body['environment']}


def test_the_config_route_names_what_is_set_and_no_longer_read(
        tmp_path, monkeypatch):
    """ADR-0014's gesture, published where the page that explains it can show it."""
    monkeypatch.setenv('SB_REGULAR_INTERVAL', '600')

    body = build_client(tmp_path).get('/api/config').get_json()

    assert 'SB_REGULAR_INTERVAL' in body['unread_environment']


# --------------------------------------------------------------------- #
# GET/PUT /api/settings — the dials, and their only writer (issue #701)
# --------------------------------------------------------------------- #

def test_the_dials_are_readable_on_the_resource_that_writes_them(tmp_path):
    """*Headless means without an interface, not without HTTP.*

    A client that could `PUT` a dial had no way to `GET` one, and the answer it
    got was worse than a bare 404: the SPA catch-all accepts any verb and wins
    the routing before Werkzeug can raise, so `curl` was told
    `No such API endpoint: /api/settings` about an endpoint that exists. The
    list is the one `put_settings` answers with, out of `settings.describe`, so
    the reader and the writer cannot describe a dial two ways.
    """
    client = build_client(tmp_path)

    payload = client.get('/api/settings')
    assert payload.status_code == 200

    read = {row['key']: row for row in payload.get_json()['settings']}
    # The registry is the list, and there is no seventh key (ADR-0014).
    assert set(read) == {spec.key for spec in settings_registry.SETTINGS}
    assert read['regular_interval']['default'] == 120

    # And it is the same list the writer answers with, field for field.
    written = client.put('/api/settings', json={'regular_interval': 300})
    assert written.status_code == 200
    assert written.get_json()['settings'] == client.get('/api/settings').get_json()['settings']


def test_the_reporting_currency_is_a_dial_and_its_absence_is_a_state(tmp_path):
    """How the API says *"nothing here has a unit yet"* (#702, ADR-0021).

    Nothing new is published for it and no route changes: the reporting currency
    is what every figure on a page is labelled with, so an absent one is the
    condition itself, and the dial is already on `/api/config` for the banner to
    read. A fourth kind of absence would make every page depend on one preamble,
    and a landing route that varies with the data is the one thing a bookmark
    cannot survive.
    """
    client = build_client(tmp_path)

    dial = _dials(client)['base_currency']
    assert dial['value'] is None and dial['default'] is None
    assert dial['stored'] is False

    assert client.put('/api/settings',
                      json={'base_currency': 'eur'}).status_code == 200

    assert _dials(client)['base_currency']['stored'] is True


def test_a_dial_is_written_and_read_back_in_the_same_request(tmp_path):
    """Reachable with one ``curl``: headless means without an interface, not
    without HTTP."""
    client = build_client(tmp_path)

    response = client.put('/api/settings', json={'regular_interval': 600})

    assert response.status_code == 200
    body = response.get_json()
    assert body['changed'] == ['regular_interval']
    dials = {row['key']: row for row in body['settings']}
    assert dials['regular_interval']['value'] == 600
    assert _dials(client)['regular_interval']['value'] == 600


def test_saving_a_dial_takes_effect_with_no_restart(tmp_path):
    """The attribute every cycle re-reads is assigned by the write path itself."""
    client = build_client(tmp_path)

    client.put('/api/settings', json={'backfill_chunk_days': 90})

    assert web_module.current_runtime().metrics.backfill_chunk_days == 90


def test_the_answer_quantifies_what_the_change_reached(tmp_path):
    """"3 symbols now, 8 more when their market opens" — never a bare 200.

    A portfolio-wide dial that reaches three symbols out of eleven has to say
    so, or the reader concludes the other eight are misconfigured.
    """
    client = build_client(tmp_path)

    body = client.put(
        '/api/settings', json={'regular_interval': 600}).get_json()

    assert body['effect']['symbols_rescheduled'] == 3
    assert body['effect']['symbols_at_market_open'] == 8
    assert web_module.current_runtime().metrics.rearm_calls == 1


def test_a_value_out_of_bounds_is_a_422_and_writes_nothing(tmp_path):
    """Well formed, and the content is what cannot be processed.

    ``400`` would read as "fix your encoding", and a client that retried on that
    reading would retry forever.
    """
    client = build_client(tmp_path)

    response = client.put('/api/settings', json={'regular_interval': 0})

    assert response.status_code == 422
    assert response.mimetype == 'application/problem+json'
    # The field is named, so a form marks the input rather than the page.
    assert response.get_json()['key'] == 'regular_interval'
    assert _dials(client)['regular_interval']['value'] == 120


def test_an_unknown_dial_is_a_422_rather_than_a_new_dial(tmp_path):
    response = build_client(tmp_path).put('/api/settings', json={'colour': 'blue'})

    assert response.status_code == 422


def test_a_rejected_body_writes_none_of_its_valid_keys(tmp_path):
    """A half-applied ``PUT`` is a state nobody asked for."""
    client = build_client(tmp_path)

    client.put('/api/settings',
               json={'regular_interval': 600, 'backfill_delay': -1})

    assert _dials(client)['regular_interval']['value'] == 120


def test_a_body_that_is_not_an_object_is_a_400(tmp_path):
    """Malformed syntax, not a refused value — and never a 503 (#655)."""
    response = build_client(tmp_path).put('/api/settings', json=[1, 2])

    assert response.status_code == 400


def test_reposting_the_same_value_re_arms_nothing(tmp_path):
    """``reschedule_job`` recomputes ``next_run_time`` from now.

    A save button that rewrote every row would put every timer back to zero on
    every click — including the timers of the dials nobody touched.
    """
    client = build_client(tmp_path)

    body = client.put(
        '/api/settings', json={'regular_interval': 120}).get_json()

    assert body['changed'] == []
    assert body['effect']['symbols_rescheduled'] == 0
    assert web_module.current_runtime().metrics.rearm_calls == 0


# --------------------------------------------------------------------- #
# Imports: the unit of revocation (issue #697)
# --------------------------------------------------------------------- #

_ONE_BUY = (
    "date,event_type,symbol,name,quantity,unit_price,amount\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,\n"
)

#: Three lines of one file, so a gesture on one of them can be shown to leave
#: the other two exactly where they were — story 14's whole point.
_THREE_LINES = (
    "date,event_type,symbol,name,quantity,unit_price,amount\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,\n"
    "2024-02-15,BUY,MSFT,Microsoft,4,380.00,\n"
    "2024-03-15,BUY,MSFT,Microsoft,1,390.00,\n"
)


def test_an_install_with_nothing_to_say_answers_an_empty_collection(tmp_path):
    """``200`` + ``[]``. Silence is the ordinary state, not a missing resource."""
    response = build_client(tmp_path).get('/api/installation-facts')

    assert response.status_code == 200
    assert response.get_json() == []


def test_an_installation_fact_is_listed_with_what_it_names(
        tmp_path, monkeypatch):
    monkeypatch.setenv('SB_EXECUTOR_POOL', '10')
    client, opened = build_client_and_store(tmp_path)
    installation_facts.refresh(opened, main.installation_fact_context(
        main.ConfigurationManager(config_dir=str(tmp_path))))

    (fact,) = client.get('/api/installation-facts').get_json()

    assert fact['key'] == installation_facts.UNREAD_ENVIRONMENT
    assert fact['acknowledged'] is False
    assert fact['acknowledged_at'] is None
    assert fact['first_seen_at'] is not None
    # The detail is **re-derived** by the read: the table has three columns.
    assert fact['detail']['variables'] == ['SB_EXECUTOR_POOL']
    assert 'SB_EXECUTOR_POOL' in fact['message']


def test_a_get_never_arms_an_installation_fact(tmp_path, monkeypatch):
    """The observation belongs to the jobs, never to somebody opening a page.

    A ``GET`` that armed them would date every installation fact with the
    moment a browser arrived — and log it there too.
    """
    monkeypatch.setenv('SB_EXECUTOR_POOL', '10')
    client, opened = build_client_and_store(tmp_path)

    assert client.get('/api/installation-facts').get_json() == []
    assert opened.query('SELECT count(*) FROM installation_fact')[0][0] == 0


def test_a_request_never_drops_what_a_running_scheduler_armed(tmp_path):
    """A runtime with no scheduler cannot see the reconstruction, and says so.

    ``UNOBSERVED`` rather than *finished*: the row stands, its detail is ``null``,
    and nothing is written — otherwise opening a page would take away the notice
    the backfill armed a minute earlier.

    *No scheduler* is a runtime with **no metrics object**, and that is the only
    shape it has: a metrics object always answers a pair, ``(0, 0)`` included,
    and ``(0, 0)`` is an observation that disarms rather than a silence.
    """
    client, opened = build_client_and_store(tmp_path, with_scheduler=False)
    installation_facts.refresh(
        opened, installation_facts.Context(reconstruction=(1, 3)))

    (fact,) = client.get('/api/installation-facts').get_json()

    assert fact['key'] == installation_facts.RECONSTRUCTION_RUNNING
    assert fact['detail'] is None
    assert opened.query('SELECT count(*) FROM installation_fact')[0][0] == 1


def test_acknowledging_hides_it_and_the_acknowledgement_persists(tmp_path):
    client, opened = build_client_and_store(tmp_path)
    installation_facts.refresh(
        opened, installation_facts.Context(unread_variables=('SB_EXECUTOR_POOL',)))

    response = client.post(
        f'/api/installation-facts/'
        f'{installation_facts.UNREAD_ENVIRONMENT}/acknowledgement')

    assert response.status_code == 200
    assert response.get_json()['acknowledged'] is True
    assert client.get('/api/installation-facts').get_json() == []
    # The row stays — that is what survives a restart, and what a toast cannot.
    assert opened.query(
        'SELECT acknowledged_at FROM installation_fact')[0][0] is not None


def test_acknowledging_twice_over_the_route_is_not_an_error(tmp_path):
    """The second gesture asserts what the first already did (issue #820).

    Held on the HTTP seam and not only under it, because the renamed resource is
    what a reader's second click reaches: a browser that retries, a page that is
    reopened, a ``curl`` run twice. The date is the row's, and the row does not
    move.
    """
    client, opened = build_client_and_store(tmp_path)
    installation_facts.refresh(
        opened, installation_facts.Context(unread_variables=('SB_EXECUTOR_POOL',)))
    route = (f'/api/installation-facts/'
             f'{installation_facts.UNREAD_ENVIRONMENT}/acknowledgement')

    first = client.post(route)
    second = client.post(route)

    assert (first.status_code, second.status_code) == (200, 200)
    assert first.get_json()['acknowledged_at'] == \
        second.get_json()['acknowledged_at']


def test_acknowledging_an_unknown_fact_is_a_404_not_a_503(tmp_path):
    response = build_client(tmp_path).post(
        '/api/installation-facts/no_such_notice/acknowledgement')

    assert response.status_code == 404
    assert response.mimetype == 'application/problem+json'


def test_acknowledging_one_that_is_not_standing_is_a_404(tmp_path):
    """The key is real and nothing stands under it: same answer to the client."""
    response = build_client(tmp_path).post(
        f'/api/installation-facts/'
        f'{installation_facts.UNREAD_ENVIRONMENT}/acknowledgement')

    assert response.status_code == 404


# --------------------------------------------------------------------- #
# Declaring an account from the app (issue #698)
# --------------------------------------------------------------------- #

def test_an_account_is_declared_here_renamed_here_and_removed_here(tmp_path):
    """The one place an account is born, and three members on the wire.

    No ``source_id`` and no ``editable``: an account is declared in the app and
    nowhere else (ADR-0034), so there is no second population to tell this row
    from and no rule about it for the front to re-implement.
    """
    client = build_client(tmp_path)

    created = client.post('/api/accounts',
                          json={'id': 'pea', 'type': 'PEA', 'label': 'PEA Bourso'})

    assert created.status_code == 201
    assert created.get_json() == {'id': 'pea', 'type': 'PEA',
                                  'label': 'PEA Bourso'}
    # The replay followed the write: the declaration is already published.
    listed = client.get('/api/accounts').get_json()
    assert listed['declared'] is True
    assert [a['id'] for a in listed['accounts']] == ['pea']

    renamed = client.patch('/api/accounts/pea', json={'label': 'PEA Fortuneo'})
    assert renamed.get_json()['label'] == 'PEA Fortuneo'

    assert client.delete('/api/accounts/pea').status_code == 200
    # Back to the install that has declared nothing: the discriminator says so,
    # and the seeded row is what is left to show (#729).
    after = client.get('/api/accounts').get_json()
    assert after['declared'] is False
    assert [a['id'] for a in after['accounts']] == ['default']


def test_declaring_an_id_twice_is_a_409(tmp_path):
    client = build_client(tmp_path)
    client.post('/api/accounts', json={'id': 'pea', 'type': 'PEA'})

    response = client.post('/api/accounts', json={'id': 'pea', 'type': 'CTO'})

    assert response.status_code == 409


def test_a_declared_account_is_renamed_and_removed_from_the_app(tmp_path):
    """An account is born in the app, so it is editable there (ADR-0034).

    The ``409`` that used to answer here was about a row an accounts **file**
    had provisioned; no file declares an account any more, and the refusal that
    stands is the one ADR-0013 has always held — an account an event names,
    which is the test below.
    """
    client = build_client(tmp_path, accounts=ACCOUNTS_FILE)

    assert client.patch('/api/accounts/pea',
                        json={'label': 'Renamed'}).status_code == 200
    assert client.delete('/api/accounts/pea').status_code == 200


def test_deleting_an_account_an_event_names_is_a_409(tmp_path):
    """ADR-0013's construction, on the wire: no orphan historical residue."""
    client = build_client(tmp_path, events=ACCOUNTS_EVENTS,
                          accounts=ACCOUNTS_FILE)

    response = client.delete('/api/accounts/pea')

    assert response.status_code == 409
    assert 'event' in response.get_json()['detail']


def test_deleting_an_unknown_account_is_a_404(tmp_path):
    assert build_client(tmp_path).delete('/api/accounts/nope').status_code == 404


def test_the_row_gestures_and_the_bulk_one_are_the_whole_map(tmp_path):
    """**Criterion 3 of #816**: the imports are not a resource any more.

    The URL map is where the decision reads. ``GET /api/imports`` and
    ``DELETE /api/imports/<id>`` are gone with the population they existed for
    (ADR-0032): nothing persists that could be named, so there is nothing to
    list and nothing to revoke. What is left is the four gestures on the
    collection an import writes — one row in, one row rewritten, one row out,
    and the reduction — plus the upload itself.
    """
    client = build_client(tmp_path, events=_ONE_BUY)
    rules = {
        (rule.rule, method)
        for rule in client.application.url_map.iter_rules()
        for method in (rule.methods or set())
        if method in {'PUT', 'PATCH', 'DELETE', 'POST', 'GET'}
    }
    paths = {rule for rule, _ in rules}

    assert '/api/imports' not in paths
    assert not any(path.startswith('/api/imports/') for path in paths)

    assert ('/api/events', 'POST') in rules
    assert ('/api/events', 'DELETE') in rules
    assert ('/api/events/import', 'POST') in rules
    assert ('/api/events/<event_id>', 'PATCH') in rules
    assert ('/api/events/<event_id>', 'DELETE') in rules
    assert ('/api/events/<event_id>', 'DELETE') in rules
    # And nothing writes a *file* — #711's demolition, still standing.
    assert not [rule for rule, _ in rules if rule.startswith('/api/events/files')]
