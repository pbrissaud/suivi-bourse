"""Tests for the editor read path (issue #659, design #653).

Two things earn tests here, and both are contract rather than arithmetic:

* the **address** is ``(file, sheet, row)`` and never ``(file, row)`` — the xlsx
  trap, where ``row_num`` restarts at 2 in every worksheet;
* the **fingerprint** changes when the row's content changes and not otherwise,
  because it is what turns a stale address into a ``409`` instead of a silently
  mis-edited row.

Plus the invariant the whole file exists to protect: ``Event``, the aggregator
and the validator are untouched by any of this.
"""
import openpyxl
import pytest

from events.editor import (
    EventEditorReader,
    decode_id,
    encode_id,
    fingerprint,
)
from events.loader import EventLoader
from events.schemas import EventType

CSV = (
    "date,event_type,symbol,name,quantity,unit_price,fee,amount,notes\n"
    "2024-01-15,BUY,AAPL,Apple Inc,10,150.00,2.50,,Initial purchase\n"
    "2024-03-01,DIVIDEND,AAPL,Apple Inc,,,,8.50,Q1 2024\n"
)


@pytest.fixture
def source(tmp_path):
    (tmp_path / '2024.csv').write_text(CSV, encoding='utf-8')
    return tmp_path


# --------------------------------------------------------------------- #
# The opaque token
# --------------------------------------------------------------------- #

def test_token_round_trips():
    token = encode_id('2024.csv', 'Feuil1', 7)
    address = decode_id(token)

    assert (address.file, address.sheet, address.row) == ('2024.csv', 'Feuil1', 7)


def test_token_does_not_leak_the_address_in_readable_form():
    """#653: `file` and `row` must never appear in the contract, or the door to
    a different store closes. Opaque is a rule about who may interpret it."""
    token = encode_id('2024.csv', '', 7)

    assert '2024.csv' not in token
    assert '/' not in token


@pytest.mark.parametrize("bad", ['not-base64!!', '', encode_id('a', 'b', 0)[:3]])
def test_malformed_tokens_raise_rather_than_resolve(bad):
    with pytest.raises(ValueError):
        decode_id(bad)


# --------------------------------------------------------------------- #
# The content fingerprint
# --------------------------------------------------------------------- #

def test_fingerprint_ignores_the_difference_between_blank_and_missing():
    """A CSV writes an empty cell, openpyxl hands back None. Same event."""
    assert fingerprint({'date': '2024-01-15', 'notes': ''}) == \
        fingerprint({'date': '2024-01-15', 'notes': None})


def test_fingerprint_changes_when_a_value_changes():
    before = fingerprint({'date': '2024-01-15', 'quantity': '10'})
    after = fingerprint({'date': '2024-01-15', 'quantity': '11'})

    assert before != after


def test_fingerprint_ignores_columns_the_format_does_not_define():
    """Adding an unrelated column to a file must not invalidate every id in it."""
    assert fingerprint({'date': '2024-01-15'}) == \
        fingerprint({'date': '2024-01-15', 'my_own_column': 'whatever'})


def test_duplicate_events_are_addressable_separately(tmp_path):
    """Why the id is an address and not a content hash: two identical DIVIDEND
    rows on the same date are legitimate, and content-addressing collides."""
    (tmp_path / 'dup.csv').write_text(
        "date,event_type,symbol,name,amount\n"
        "2024-03-01,DIVIDEND,AAPL,Apple Inc,8.50\n"
        "2024-03-01,DIVIDEND,AAPL,Apple Inc,8.50\n",
        encoding='utf-8')

    records = EventEditorReader(str(tmp_path)).list_records()

    assert len(records) == 2
    assert records[0].id != records[1].id
    assert records[0].etag == records[1].etag  # same content, different address


# --------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------- #

def test_csv_rows_are_addressed_from_line_two(source):
    records = EventEditorReader(str(source)).list_records()

    assert [decode_id(r.id).row for r in records] == [2, 3]
    assert all(decode_id(r.id).sheet == '' for r in records)
    assert all(r.editable for r in records)


def test_xlsx_addressing_carries_the_sheet(tmp_path):
    """The trap: `row_num` restarts at 2 per worksheet (`loader.py:125,139`), so
    two rows of one workbook share a row number and only the sheet tells them
    apart."""
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = 'PEA'
    first.append(['date', 'event_type', 'symbol', 'name', 'quantity', 'unit_price'])
    first.append(['2024-01-15', 'BUY', 'AAPL', 'Apple Inc', 10, 150.0])

    second = workbook.create_sheet('CTO')
    second.append(['date', 'event_type', 'symbol', 'name', 'quantity', 'unit_price'])
    second.append(['2024-02-01', 'BUY', 'MSFT', 'Microsoft', 5, 380.0])
    workbook.save(tmp_path / 'broker.xlsx')

    records = EventEditorReader(str(tmp_path)).list_records()
    addresses = [decode_id(r.id) for r in records]

    assert len(records) == 2
    assert [a.row for a in addresses] == [2, 2]        # identical row numbers…
    assert [a.sheet for a in addresses] == ['PEA', 'CTO']   # …disambiguated here
    assert {r.id for r in records} == {records[0].id, records[1].id}
    assert records[0].id != records[1].id


def test_xlsx_rows_are_not_editable(tmp_path):
    """openpyxl reads with data_only=True to see cached values, and saving from
    that state writes values over the formulas. #652 déc. 14 offers an explicit
    convert-to-CSV instead of a half-working save."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['date', 'event_type', 'symbol', 'name', 'amount'])
    sheet.append(['2024-03-01', 'DIVIDEND', 'AAPL', 'Apple Inc', 8.5])
    workbook.save(tmp_path / 'broker.xlsx')

    records = EventEditorReader(str(tmp_path)).list_records()

    assert records[0].editable is False


def test_values_are_typed_by_the_loader_s_own_rules(source):
    records = EventEditorReader(str(source)).list_records()

    assert records[0].event.event_type is EventType.BUY
    assert records[0].event.quantity == 10.0
    assert records[0].event.date.isoformat() == '2024-01-15'
    assert records[1].event.amount == 8.5


def test_a_bad_row_is_reported_rather_than_raised(tmp_path):
    """A ledger that 500s because one row is malformed cannot be used to *fix*
    that row. The app itself stays stricter — since #658 a rejected config is
    fatal at boot; this path is what lets a human see why."""
    (tmp_path / 'broken.csv').write_text(
        "date,event_type,symbol,name,quantity\n"
        "2024-01-15,BUY,AAPL,Apple Inc,10\n"
        "not-a-date,BUY,AAPL,Apple Inc,5\n",
        encoding='utf-8')

    records = EventEditorReader(str(tmp_path)).list_records()

    assert records[0].error is None
    assert records[1].event is None
    assert 'Invalid date format' in records[1].error


def test_a_missing_source_is_empty_not_an_error(tmp_path):
    assert EventEditorReader(str(tmp_path / 'nope')).list_records() == []


def test_to_dict_exposes_the_id_and_etag_but_no_file_path(source):
    payload = EventEditorReader(str(source)).list_records()[0].to_dict()

    assert payload['id'] and payload['etag']
    assert payload['source'] == '2024.csv'   # display-only provenance
    assert 'file' not in payload and 'row' not in payload
    assert payload['event_type'] == 'BUY'


# --------------------------------------------------------------------- #
# The invariant
# --------------------------------------------------------------------- #

def test_the_loader_still_produces_identical_events(source):
    """`Event`, the aggregator and the validator stay byte-identical (#653).

    The editor shares the loader's *parser* rather than reimplementing it, so
    the ledger and the aggregated position can never disagree about what a row
    means — but it adds nothing to `Event` itself.
    """
    loaded = EventLoader(str(source)).load()
    edited = [r.event for r in EventEditorReader(str(source)).list_records()]

    assert loaded == edited
    assert not hasattr(loaded[0], 'id')
    assert not hasattr(loaded[0], 'etag')
