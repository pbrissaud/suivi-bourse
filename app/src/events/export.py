"""The export: the ledger back out, in the format it comes in (issue #710).

The store is the truth of the ledger (#697), which closes a door: a portfolio
that used to be a folder of files a person could copy is now a binary DuckDB
file. Two questions follow from that, and one gesture answers both — *can I go
back to v4?* and *is my only backup a binary?*

**The format of the export is the format of the import**, so the round trip is
round by construction rather than by a mapping kept in step by hand. What is
rendered here is what :class:`events.loader.EventLoader` reads, in the column
order the documentation shows, and re-dropping the result into the drop folder
is an ordinary import with no special path anywhere.

Three things about it are decisions rather than defaults.

**The file states its reporting currency, and the column is called
``base_currency``.** Event amounts are the debit *in the reporting currency*
(ADR-0002) and nothing in the file says which one, so a round trip through an
install that answered differently would silently re-read three years of euros
as dollars. The name is the guard: a broker export routinely carries a
``currency`` column meaning the **security's quote** currency, and reading that
one as the reporting currency is exactly the reinterpretation this exists to
prevent. Two currency levels, two names.

**It is a column repeated on every row, not a preamble.** A CSV has one header
row and nowhere above it to put a fact about the file; a sidecar document would
not be read on the way back in at all, and the export has to re-enter by the
**normal** import path or it proves nothing. The cost — one repeated cell per
row — buys a file that opens in a spreadsheet like every other one.

**There is no provenance left to leave out** (ADR-0032, #816). The three columns
that said which import carried a row went with the second population, and the
round trip is the plainer for it: what leaves is every column a row has, and
what comes back is a row exactly like the one that left.

Since #796 it renders **two more things**, and neither is a second format.
A **workbook with one sheet per year** is the same rows and the same columns,
laid out the way a ledger is actually read in a spreadsheet — and it is
importable, the loader reading every worksheet of a file. A **selection** is the
ledger's own chips applied here rather than in the front: what leaves is the
importable form, so a partial file assembled on the other side of the wire would
be a second spelling of a format that only this module owns. The front sends the
names on a query string; the reduction is arrived at twice over **one** format
instead of once over two. Since #810 there are **five** of them: the period
joined the four, because an import is an interval of dates before it is anything
else — and because *extract a year* is what a backup with no period cannot do.

Since #836 it renders one more thing, and that one is **not a backup**. The
*accounts and positions* file is a **report**: balances, weighted-average unit
costs and valuations, which are derived state and not a declaration anybody
could hand back. Nothing reads it in — the loader refuses it by name, for want
of ``date`` and ``event_type`` — and that refusal is what keeps it on the right
side of ADR-0034, which retired the old ``accounts.csv`` precisely because it
*looked* like a restorable backup and was not one. A file the import turns away
in one sentence cannot be mistaken for half a restore.

The two files are not interchangeable, and the difference is a precision.
``openpyxl`` writes a double as ``%.16g``, one significant digit short of the
shortest string that reads back as the same double, so the CSV is the one that
survives a round trip bit for bit — which is why it, and not the workbook, keeps
the backup's name.

This is the *rendering* half of #711's deleted ``events/editor.py``, which spec
#695 § 6 had reserved for exactly this use. It is rewritten rather than restored:
what the old module rendered was a file being edited in place — an addressable
``CsvFile``, an atomic rename, a workbook conversion — and none of those three
has a subject once the rows come from the store and the bytes go into an HTTP
response.
"""
import csv
import io
import re
import unicodedata
from datetime import date
from typing import (
    Any, Dict, Iterable, List, Mapping, NamedTuple, Optional, Sequence, Tuple,
)

import instants

from .loader import BASE_CURRENCY_COLUMN
from .schemas import DEFAULT_ACCOUNT, Event, unit_cost

#: The columns of an exported event file, in the order the documentation shows
#: them (``website/docs/import-your-events.mdx``). ``base_currency`` closes the
#: list because it is the one column that is a fact about the *file*: a reader
#: scanning left to right meets the event first and its unit last.
EVENT_COLUMNS = (
    'date', 'event_type', 'account', 'symbol', 'name',
    'quantity', 'unit_price', 'fee', 'amount', 'notes',
    BASE_CURRENCY_COLUMN,
)

#: What a worksheet cannot carry: the C0 controls, tab, newline and carriage
#: return excepted. It is exactly the range ``openpyxl`` raises on, restated
#: here because it is a fact about the **format** rather than about the library
#: — and because naming it costs no import at module scope, which the workbook
#: deliberately does not take.
_UNWRITABLE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

#: The sheet a workbook falls back to when there is no year to name — an empty
#: ledger. A workbook with no sheet at all is not a file a spreadsheet opens, so
#: the header stands alone under a tab that names nothing rather than under a
#: year nothing happened in.
UNDATED_SHEET = 'events'


def render_events(events: Iterable[Event],
                  base_currency: Optional[str] = None) -> str:
    """The ledger as an importable ``.csv``.

    ``base_currency`` is written on every row when the install has answered the
    question, and left blank on every row when it has not. Blank is the honest
    answer: an install with no reporting currency has interpreted nothing, so
    the file has nothing to state — and a file whose column is empty imports
    everywhere, which is what an unanswered install's backup should do.
    """
    return _render(EVENT_COLUMNS, [
        _event_row(event, base_currency) for event in events
    ])


def _event_row(event: Event, base_currency: Optional[str]) -> dict:
    """One event as the cells a file holds.

    ``account`` is written as the store holds it, ``default`` included: the
    seeded account exists on every install (ADR-0013), so the cell always names
    something the re-import can resolve — where a blank cell would be resolved
    by a rule that depends on whether anything else has been declared yet.
    """
    return {
        'date': event.date,
        'event_type': event.event_type.value,
        'account': event.account,
        'symbol': event.symbol,
        'name': event.name,
        'quantity': event.quantity,
        'unit_price': event.unit_price,
        'fee': event.fee,
        'amount': event.amount,
        'notes': event.notes,
        BASE_CURRENCY_COLUMN: base_currency,
    }


def _render(columns: Sequence[str], rows: Sequence[dict]) -> str:
    """Header plus rows, ``\\n``-terminated, UTF-8 without a byte-order mark.

    ``lineterminator='\\n'`` rather than the module default ``\\r\\n``: the file
    is meant to be diffed and version-controlled beside the ones the user wrote
    by hand, and the loader reads either.
    """
    buffer = io.StringIO(newline='')
    writer = csv.DictWriter(buffer, fieldnames=list(columns), restval='',
                            extrasaction='ignore', lineterminator='\n')
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _cell(row.get(column)) for column in columns})
    return buffer.getvalue()


def _cell(value: Any) -> str:
    """One cell as text.

    ``None`` is an **empty cell**, never the string ``'None'`` — an empty cell
    is what *"this event has no fee"* looks like in a file, and what the loader
    reads back as ``None``. A ``float`` goes out through ``repr``, which since
    Python 3.1 is the shortest string that reads back as the same double, so a
    broker's ``0.34898399999999996`` survives the round trip bit for bit.

    A ``date`` goes out through :func:`instants.iso`, which is where the tree
    writes ISO once (#843). An event's date is a **calendar day** and leaves as
    one; the delegation is what keeps the rule from being re-spelled here, and
    it is why the empty-cell test comes first — ``iso(None)`` answers ``None``,
    and a file wants a blank.
    """
    if value is None:
        return ''
    if isinstance(value, date):
        return instants.iso(value)
    return str(value)


# --------------------------------------------------------------------- #
# The workbook — one sheet per year (issue #796)
# --------------------------------------------------------------------- #

def render_events_workbook(events: Iterable[Event],
                           base_currency: Optional[str] = None) -> bytes:
    """The ledger as an importable ``.xlsx``, **one sheet per year**.

    The same rows as :func:`render_events` and the same columns, laid out the
    way somebody actually reads a ledger in a spreadsheet — a tab is a year, and
    the years are the *events'* own: a ledger that skips a year skips a tab,
    because a range would state a year in which nothing happened.

    It stays importable, and that is not a bonus: the loader reads **every**
    worksheet of a workbook and validates the header of each, so the header is
    repeated on every tab and ``base_currency`` rides on every row exactly as it
    does in the CSV. One file, one reporting currency — sheets included, which
    the loader would refuse otherwise.

    The values go out **typed**: a date is a date and an amount is a number, so
    the file sorts and sums where it is opened, and comes back through
    :meth:`events.loader.EventLoader._load_xlsx` — which reads a ``datetime``
    and a ``float`` natively — as the double that left. Text is forced to text,
    which is the one thing a spreadsheet does not do by default and the reason
    a note is not silently a formula.

    The import is **local to this function**, in the taste of the loader's:
    ``openpyxl`` is a declared dependency, and imported at module scope a broken
    install would take the whole ``/api`` blueprint down instead of the one
    route that needs a workbook.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for year, rows in _by_year(events).items():
        sheet = workbook.create_sheet(title=year)
        _write_row(sheet, list(EVENT_COLUMNS))
        for event in rows:
            cells = _event_row(event, base_currency)
            _write_row(sheet, [cells[column] for column in EVENT_COLUMNS])
    if not workbook.sheetnames:
        _write_row(workbook.create_sheet(title=UNDATED_SHEET),
                   list(EVENT_COLUMNS))

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _by_year(events: Iterable[Event]) -> Dict[str, List[Event]]:
    """The events grouped by the year they are dated, oldest tab first.

    Ascending, so reading the tabs left to right reads the ledger forward — the
    order the store hands the rows over in, and the one a replay follows.
    """
    years: Dict[str, List[Event]] = {}
    for event in sorted(events, key=lambda event: event.date):
        years.setdefault(f'{event.date.year:04d}', []).append(event)
    return years


def _write_row(sheet, values: Sequence[Any]) -> None:
    """One row of cells, with text kept text and kept **writable**.

    Two things happen to a string on the way in, and each is a defect the
    library would otherwise hand the reader.

    ``openpyxl`` binds a string beginning with ``=`` as a **formula**, so a note
    reading ``=SUM(A1:A2)`` would leave as one, be evaluated where the file is
    opened, and come back as whatever that evaluated to. A ledger's note is what
    somebody typed; the type is restated after the binding, which is the only
    place the library offers to say so.

    And OOXML cannot carry ``\x00``–``\x1f``, so ``openpyxl`` **raises** on a
    cell holding one. Nothing upstream removes them — the API's text members are
    stripped, which never touches an interior character, and both loaders pass a
    cell through as it came — so one vertical tab in one note pasted out of a
    PDF would make the workbook route a ``500`` for the *whole* ledger, silently,
    the CSV going on working until the reader picks the other entry. The
    character is dropped here rather than the file being lost: it is not
    renderable in a spreadsheet under any encoding, and the CSV — the backup —
    keeps the byte.
    """
    sheet.append([_writable(value) for value in values])
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str):
            cell.data_type = 's'


def _writable(value: Any) -> Any:
    """One value a worksheet will accept — text minus what OOXML cannot hold."""
    if not isinstance(value, str):
        return value
    return _UNWRITABLE.sub('', value)


# --------------------------------------------------------------------- #
# The selection — the ledger's own chips, on this side (issue #796)
# --------------------------------------------------------------------- #

class Selection(NamedTuple):
    """What the ledger's chips retain, as the export resource takes it.

    The five are the table's five, and they are read **here** rather than
    applied in the front for the reason the whole module exists: what leaves is
    the *importable* form, and a file rendered on the other side of the wire
    would be a second spelling of it. What the front owns is the vocabulary —
    the same five names, on the query string — and the reduction it draws is the
    same reduction, arrived at twice over one contract rather than over two
    formats.

    Every member is *no reduction* by default, and no member is ever *no match*:
    an empty query is every row, and ``None`` is every type, every account,
    every security and every day.

    **The period is named** ``since`` **and** ``until`` **, on both sides of the
    wire** (issue #810). ``from`` is the natural word on a query string and it is
    a Python keyword, so a ``NamedTuple`` cannot carry it: one vocabulary means
    choosing the pair of names both sides can spell, rather than translating
    ``?from=`` into something else here and letting the two drift.

    **Two day bounds, inclusive, each optional.** The ledger is dated to the day
    (ADR-0008), so both bounds retain the days they name — what the reader sees
    on the chip is the interval they typed, and not a half-open one to reason
    about. One bound alone is a legitimate reduction: *everything since 2023*
    opens the interval on the other side.
    """

    query: str = ''
    event_type: Optional[str] = None
    account: Optional[str] = None
    symbols: Optional[Tuple[str, ...]] = None
    since: Optional[date] = None
    until: Optional[date] = None

    @property
    def reduces(self) -> bool:
        """Whether anything at all is being held back.

        It is what names the file on the way out: a reduction is **not a
        backup**, so it does not take the backup's name — under it the partial
        file would replace the whole one on the reader's own disk, which is the
        one protection a fixed export name buys (see ``EXPORT_FILENAMES``).

        **The period counts** (issue #810): an export reduced on the dates alone
        is an extract of one year, and under the backup's name that partial file
        would replace the whole one on the reader's own disk.
        """
        return bool(self.query.strip()) or self.event_type is not None \
            or self.account is not None or self.symbols is not None \
            or self.since is not None or self.until is not None


#: No parameter at all: the whole ledger, which is what the plain route serves.
NO_SELECTION = Selection()


def select(events: Iterable[Event], selection: Selection) -> List[Event]:
    """The events a selection retains, in the order they arrived.

    The two bounds are **inclusive**, which is the whole of the period's
    contract: a ledger dated to the day has no instant to be before or after, so
    ``since=2024-01-01`` and ``until=2024-12-31`` are the year 2024 and both of
    the days they name are in it.
    """
    needle = fold(selection.query.strip())
    symbols = None if selection.symbols is None else set(selection.symbols)
    kept = []
    for event in events:
        if selection.since is not None and event.date < selection.since:
            continue
        if selection.until is not None and event.date > selection.until:
            continue
        if selection.event_type is not None \
                and event.event_type.value != selection.event_type:
            continue
        if selection.account is not None \
                and account_of(event) != selection.account:
            continue
        if symbols is not None \
                and (not event.symbol or event.symbol not in symbols):
            continue
        if needle and needle not in _haystack(event):
            continue
        kept.append(event)
    return kept


def account_of(event: Event) -> str:
    """The account a row names — **a blank one names** ``default``.

    The aggregator's own rule (:mod:`entries`, ADR-0013) and not a rendering
    nicety: an install that declared nothing writes its events under the seeded
    row, and a blank cell resolved to nothing here would make the seeded
    account unreachable by the one chip that names it.
    """
    return (event.account or '').strip() or DEFAULT_ACCOUNT


def fold(value: str) -> str:
    """Accents dropped and case folded: a French label is searched as it is heard.

    ``Février`` and ``fevrier`` are the same word to somebody typing quickly,
    and a search that disagreed would be a table silently shorter than the
    reader expects — the defect the whole reduction bar is written against.
    """
    decomposed = unicodedata.normalize('NFD', value)
    return ''.join(character for character in decomposed
                   if not unicodedata.combining(character)).lower()


def _haystack(event: Event) -> str:
    """What the search reads: everything the identity and account columns show.

    The ticker, the free-text label and the account — and it is not a
    convenience: on nineteen purchases of one ETF the label is the only
    discriminant a row owns, and on a cash movement it is the only name at all.
    """
    return fold(' '.join(part for part in
                         (event.symbol, event.notes, account_of(event)) if part))


# --------------------------------------------------------------------- #
# Accounts and positions — the one file that is a report (issue #836)
# --------------------------------------------------------------------- #

#: The columns of the accounts-and-positions file, in reading order: the account
#: first, its cash next, then the holding and what it is worth. ``base_currency``
#: closes the list because it closes :data:`EVENT_COLUMNS` — it is the one column
#: that is a fact about the *file* rather than about the row.
#:
#: The store's own names are used throughout (``realized_gain``,
#: ``received_dividend``), the way ``event_type`` is one: this module renders
#: what the tables hold, and a second vocabulary invented on the way out is a
#: translation to keep in step by hand. ``unit_cost`` is the one exception and it
#: is not a rename — it is the *PMP*, derived and never stored (ADR-0003).
PORTFOLIO_COLUMNS = (
    'account', 'account_label', 'account_type',
    'cash_balance', 'net_contributed',
    'symbol', 'name', 'quantity', 'unit_cost', 'cost_basis',
    'price', 'market_value', 'realized_gain', 'received_dividend',
    BASE_CURRENCY_COLUMN,
)


def render_portfolio(accounts: Iterable[Any],
                     states: Mapping[str, Any],
                     positions: Iterable[Mapping[str, Any]],
                     base_currency: Optional[str] = None) -> str:
    """The accounts and their positions — balances, PMP and valuations (#836).

    The fourth entry of the export menu, and the only one of the four that is
    **not a backup**. The three others render the ledger, which is the truth
    (ADR-0032) and re-enters by the ordinary import path; this renders what the
    replay *derived* from it — a position, a unit cost, a cash balance — none of
    which anybody hands back.

    **The loader refuses this file by name**, for want of ``date`` and
    ``event_type``, and that refusal is the whole reason it may exist at all.
    ADR-0034 retired the old ``accounts.csv`` because a file nothing reads back
    *looks* like a restorable backup; one the import turns away in one sentence
    does not, and its name says report where the other said declaration. The
    accounts are still born in the app and nowhere else — nothing here declares
    anything.

    **A figure appears once, a name repeats.** An account's own two figures — the
    cash standing in it and what was put into it — are on the account's own row,
    where the position columns are empty; a position's figures are on the
    position's row, where the cash columns are. So every money column of this
    file sums to something true, which is the one thing a spreadsheet reader will
    actually do with it. The account's *label* and *type* repeat on the rows
    under it, because a name summed is nothing and a name missing is a filter
    that cannot be drawn.

    **An account with no cash ledger has empty cells, never zeros.** ``0.00`` and
    *nobody has ever moved money in this account* are two states, and the store
    keeps them apart by having no ``account_state`` row at all for the second —
    ADR-0006's rule read at the level of a cell.

    **The price is the observed one, and blank when there is none.** The carrying
    convention (ADR-0004) is deliberately not applied: it is a *named helper*
    whose domain is exactly *the symbol's backfill is terminal*, and establishing
    that takes the published snapshot's holding windows — which is the one read
    an export must not make, a backup being of **what is stored**. A file that
    substituted a cost for a price would publish a valuation nobody observed,
    and it would do it in the one place the reader cannot see the em dash that
    says so on the page.
    """
    declared = {account.id: account for account in accounts}
    held: Dict[str, List[Mapping[str, Any]]] = {}
    for position in positions:
        account = (position.get('account') or '').strip() or DEFAULT_ACCOUNT
        held.setdefault(account, []).append(position)

    rows: List[dict] = []
    # Every account the store names, declared or merely held in. A declared
    # account holding nothing is ordinary and keeps its row — its cash figures
    # are the point. The other side, held and undeclared, is unreachable by
    # construction: the validator refuses an event naming an account nothing
    # declared (#698), and ``accounts.delete_account`` refuses any account an
    # event names (ADR-0013). The union is defensive, and it costs one ``|``:
    # were that ever to stop holding, a position the reader owns would fall out
    # of the file in silence.
    for account in sorted(set(declared) | set(held)):
        rows.append(_account_row(account, declared.get(account),
                                 states.get(account), base_currency))
        for position in sorted(held.get(account, []),
                               key=lambda row: str(row.get('symbol') or '')):
            rows.append(_position_row(account, declared.get(account),
                                      position, base_currency))
    return _render(PORTFOLIO_COLUMNS, rows)


def _account_row(account: str, declaration: Optional[Any],
                 state: Optional[Any], base_currency: Optional[str]) -> dict:
    """The account itself: what it is called, and what cash stands in it.

    The declaration is written **as the store holds it**, the seeded ``default``
    row included: :func:`accounts.as_declared` is what the *wire* owes a front
    that must not render an English seed (ADR-0024), and a file is not a
    rendering — it states the row, and a reader who has never named that account
    reads the name the app gave it.
    """
    return {
        'account': account,
        'account_label': None if declaration is None else declaration.label,
        'account_type': None if declaration is None else declaration.type,
        'cash_balance': None if state is None else state.cash_balance,
        'net_contributed': None if state is None else state.net_contributed,
        BASE_CURRENCY_COLUMN: base_currency,
    }


def _position_row(account: str, declaration: Optional[Any],
                  position: Mapping[str, Any],
                  base_currency: Optional[str]) -> dict:
    """One holding: what is held, what it cost, and what it is worth.

    ``market_value`` is ``quantity × price``, and absent the moment either is —
    the arithmetic :mod:`portfolio_view` does for the wire, done here for a cell
    and returning nothing rather than zero, since *a share nobody has priced* and
    *a share worth nothing* are not the same line.

    A **sold** position travels like any other (ADR-0017): its quantity is zero,
    it has no unit cost — a position nobody holds has a realized gain instead —
    and it is that gain the row has left to state.
    """
    quantity = position.get('quantity')
    price = position.get('price')
    return {
        'account': account,
        'account_label': None if declaration is None else declaration.label,
        'account_type': None if declaration is None else declaration.type,
        'symbol': position.get('symbol'),
        'name': position.get('name'),
        'quantity': quantity,
        'unit_cost': unit_cost(quantity or 0.0,
                               position.get('cost_basis') or 0.0),
        'cost_basis': position.get('cost_basis'),
        'price': price,
        'market_value': (None if quantity is None or price is None
                         else quantity * price),
        'realized_gain': position.get('realized_gain'),
        'received_dividend': position.get('received_dividend'),
        BASE_CURRENCY_COLUMN: base_currency,
    }


__all__ = [
    'EVENT_COLUMNS', 'BASE_CURRENCY_COLUMN', 'PORTFOLIO_COLUMNS',
    'UNDATED_SHEET', 'NO_SELECTION', 'Selection',
    'render_events', 'render_events_workbook', 'render_portfolio',
    'select', 'account_of', 'fold',
]
