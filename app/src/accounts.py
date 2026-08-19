"""Accounts are declared by a file, and a bad file does not get in (issue #698).

An account is **user data with provenance, not a setting** (ADR-0013). So it is
declared the way an event is — by a file, in the events' own format, or from the
UI — and it is revoked **file by file**.

Four decisions carry the module.

**The file is what multi-account needs in the UI and headless alike.** An install
with no interface has nothing to click, so reserving the declaration to the UI
would forbid multi-account to every headless install — and with it the import of
a broker export that names accounts. The headless principle of the roadmap is
what bites here, not what decorates.

**The format is the events' format** (``.csv``/``.xlsx``), not a third one:
nobody should have to learn YAML to declare two accounts. The columns are the
columns of the ``account`` table — ``id``, ``type``, ``label`` — and **no
filename has a special meaning**: a file is an accounts source because of what
its header says, never because of what it is called.

**There is always at least one account.** The seeded ``default`` row is the one
row every install owns, so nothing in the app branches on *"are accounts
declared"* (ADR-0013): an empty ``account`` column means ``default`` until an
accounts source exists, and an error afterwards. That is v4's rule **minus the
opt-in**, and it is what makes a single-account v4's event files import without
a single edit.

**An account is undeletable while an event names it.** Whether the gesture is
"delete this account" or "forget the file that declared it", the answer is the
same refusal, and it retires by construction the historical residue the earlier
design merely tolerated. Cascade forgetting is therefore refused, not performed.

**Not in this module**: ``import_source`` itself and the event rows, which are
:mod:`ledger`'s (issue #697). This module is handed a ``source_id`` and owns the
``account`` table alone — one writer per row, as the schema's generating rule
requires.
"""
import csv
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Set

from logfmt_logger import getLogger

import perf_series
import store as store_module
from events.schemas import (
    ACCOUNT_FILE_COLUMNS, Account, DEFAULT_ACCOUNT, Portfolio)

logger = getLogger("accounts")

#: The columns of an account file — the columns of the ``account`` table.
#: ``label`` is optional and falls back to the id, the way v4's ``accounts:``
#: block did; the other two are what the DDL declares ``NOT NULL``.
#:
#: The list itself lives in :mod:`events.schemas`, and this is a second **name**
#: for it rather than a second value: the validator quotes it in the refusal it
#: raises when an event names an account nobody declared, and it cannot import
#: this module — but this one can import that one, so one tuple serves both
#: messages and neither can drift from the DDL without the other.
ACCOUNT_COLUMNS = ACCOUNT_FILE_COLUMNS
REQUIRED_ACCOUNT_COLUMNS = frozenset({'id', 'type'})

#: What tells an accounts file from an event file: its **header**, never its
#: name (spec #695 § 6 — no filename has a special meaning in v5). ``event_type``
#: is what an event file always carries and an accounts file never does, so a
#: file that somehow held both columns is read as the events file it looks most
#: like and refused by the event loader on its own terms.
_EVENT_MARKER = 'event_type'

#: ``utf-8-sig`` and not ``utf-8``: Excel's own *"CSV UTF-8"* export writes a
#: byte-order mark, which under plain ``utf-8`` turns the first column name into
#: ``﻿id`` — so the file would not even be *recognised* as a declaration,
#: and would be refused as an event file missing its ``date`` column. The codec
#: reads a mark-less file identically, so this costs nothing to a file that
#: never had one.
CSV_ENCODING = 'utf-8-sig'


class AccountSourceError(Exception):
    """The accounts file cannot be read, or what it declares cannot stand.

    Raised **before** anything is written, so a refused accounts file leaves the
    store exactly as it was — the same all-or-nothing contract an event file
    gets.
    """


class AccountInUse(Exception):
    """The account cannot go: an event names it.

    The one refusal ADR-0013 turns into a construction rather than a
    convention. It answers both gestures that could orphan an event — deleting
    the account, and forgetting the import that declared it — so the cascade is
    *refused*, never performed.
    """


class UnknownAccount(Exception):
    """No account has that id. A 404 at the API, never a silent create."""


class DuplicateAccount(Exception):
    """That id is taken. Two accounts with one id is what the PK forbids."""


class ReadOnlyAccount(Exception):
    """The account came from a file, so the app does not edit it.

    What came from a file is read-only and revoked by forgetting its import;
    what was created in the app is editable (issue #698). Editing a
    file-provisioned row in place would put the store and the file in
    disagreement with no way to tell which one is meant.
    """


@dataclass(frozen=True)
class AccountRow:
    """One row of an accounts file, with the position that names it in an error.

    ``sheet``/``row`` are the same displayable provenance an event carries
    (issue #697) — here they only ever reach an error message, because the
    ``account`` table has no room for them and an account is one row, not a
    line in a ledger.
    """
    id: str
    type: str
    label: str
    sheet: Optional[str] = None
    row: Optional[int] = None

    def where(self) -> str:
        """*"row 3"* / *"sheet Comptes, row 3"* — for the refusal message."""
        parts = []
        if self.sheet:
            parts.append(f"sheet {self.sheet}")
        if self.row is not None:
            parts.append(f"row {self.row}")
        return ", ".join(parts) or "?"


# --------------------------------------------------------------------------- #
# Reading the file
# --------------------------------------------------------------------------- #

def is_accounts_file(path: Path) -> bool:
    """Is this file an accounts source? Decided on its header alone.

    Unreadable, empty or headerless files answer ``False``: they are then taken
    for event files and refused by the event loader, which is the refusal that
    names the columns a user is most likely to have meant.
    """
    try:
        header = header_of(Path(path))
    except Exception:
        return False
    if _EVENT_MARKER in header:
        return False
    return REQUIRED_ACCOUNT_COLUMNS.issubset(header)


def header_of(path: Path) -> Set[str]:
    """The column names of the first row, lowercased and stripped.

    **One row, never the file.** This runs for every droppable file on every
    filesystem event, next to the fingerprint hash the always-on watch is
    budgeted around (issue #697), so a workbook is opened read-only and its
    first row pulled lazily — materialising every sheet here would put a full
    parse of the drop folder behind a watcher that exists to cost one hash.
    """
    suffix = path.suffix.lower()
    if suffix == '.csv':
        with open(path, 'r', encoding=CSV_ENCODING) as handle:
            first = handle.readline()
        return _normalised(next(csv.reader([first]), []))
    if suffix == '.xlsx':
        return _normalised(_xlsx_first_row(path))
    return set()


def load_account_rows(path: Path) -> List[AccountRow]:
    """Every account the file declares, in file order.

    Raises:
        AccountSourceError: the file cannot be read, lacks ``id``/``type``, or
            carries a row that is not a declarable account.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == '.csv':
        return _load_csv(path)
    if suffix == '.xlsx':
        return _load_xlsx(path)
    raise AccountSourceError(f"Unsupported accounts file format: {suffix}")


def _load_csv(path: Path) -> List[AccountRow]:
    try:
        with open(path, 'r', encoding=CSV_ENCODING) as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise AccountSourceError(f"Empty accounts file: {path.name}")
            _require_columns(_normalised(reader.fieldnames), path.name)
            # start=2: row 1 is the header, so the number in a refusal is the
            # one the user's editor shows them.
            return [
                _row_from(dict(_lower_keys(raw)), path.name, number)
                for number, raw in enumerate(reader, start=2)
                if any((value or '').strip() for value in raw.values())
            ]
    except OSError as exc:
        raise AccountSourceError(f"Cannot read {path.name}: {exc}") from exc


def _load_xlsx(path: Path) -> List[AccountRow]:
    rows: List[AccountRow] = []
    for title, values in _xlsx_sheets(path):
        if not values:
            continue
        header = [str(cell).lower().strip() if cell is not None else ''
                  for cell in values[0]]
        _require_columns(set(header), f"{path.name}/{title}")
        for number, line in enumerate(values[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in line):
                continue
            raw = {key: value for key, value in zip(header, line)}
            rows.append(_row_from(raw, path.name, number, sheet=title))
    return rows


def _open_workbook(path: Path):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - openpyxl is a hard dep
        raise AccountSourceError(
            "openpyxl is required to read .xlsx files") from exc
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _xlsx_first_row(path: Path) -> Sequence:
    """The first row of the first worksheet, and not a cell more."""
    workbook = _open_workbook(path)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(max_row=1, values_only=True):
                return row
            return ()
        return ()
    finally:
        workbook.close()


def _xlsx_sheets(path: Path) -> List[tuple]:
    """``[(sheet title, [row, …]), …]`` — the workbook, values only."""
    workbook = _open_workbook(path)
    try:
        return [(sheet.title, list(sheet.iter_rows(values_only=True)))
                for sheet in workbook.worksheets]
    finally:
        workbook.close()


def _require_columns(header: Set[str], where: str) -> None:
    missing = sorted(REQUIRED_ACCOUNT_COLUMNS - header)
    if missing:
        raise AccountSourceError(
            f"Missing required column(s) {missing} in {where}; an accounts "
            f"file has the columns {list(ACCOUNT_COLUMNS)}")


def _lower_keys(raw: dict):
    return ((str(key).lower().strip() if key else '', value)
            for key, value in raw.items())


def _row_from(raw: dict, filename: str, number: int,
              sheet: Optional[str] = None) -> AccountRow:
    account_id = _text(raw.get('id'))
    account_type = _text(raw.get('type'))
    label = _text(raw.get('label'))
    where = f"sheet {sheet}, row {number}" if sheet else f"row {number}"

    if not account_id:
        raise AccountSourceError(f"{filename} ({where}): id is required")
    if not account_type:
        raise AccountSourceError(
            f"{filename} ({where}): type is required for account "
            f"{account_id!r}")

    # The label falls back to the id rather than to the empty string: the
    # column is `NOT NULL`, and a row that cannot decline to name itself is
    # what makes the interface's account list readable without a second lookup.
    return AccountRow(id=account_id, type=account_type,
                      label=label or account_id, sheet=sheet, row=number)


def _text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _normalised(names: Iterable) -> Set[str]:
    return {str(name).lower().strip() for name in names if name is not None}


# --------------------------------------------------------------------------- #
# Reading the table
# --------------------------------------------------------------------------- #

def read_accounts(store) -> List[Account]:
    """Every row of ``account``, id-sorted. ``default`` is always among them."""
    rows = store.query(
        'SELECT id, type, label, source_id FROM account ORDER BY id')
    return [Account(id=r[0], type=r[1], label=r[2], source_id=r[3])
            for r in rows]


def account_ids(store) -> Set[str]:
    """The ids an event may name. Never empty — ``default`` is always in it."""
    return {row[0] for row in store.query('SELECT id FROM account')}


def accounts_are_declared(store) -> bool:
    """Has anything been declared beyond the account every install is given?

    The one predicate the empty-``account``-column rule turns on: blank means
    ``default`` until this is true, and is an error afterwards (spec #695 § 6).

    It is deliberately *not* "does the account table have rows" — it always has
    one — and deliberately not "was a **file** imported" either. An account
    created in the app declares just as much as a file does, so a blank column
    after it is the same omission, and answering otherwise would quietly pile a
    second account's events onto the first, which is the exact mistake the rule
    exists to refuse.
    """
    rows = store.query(
        "SELECT count(*) FROM account WHERE id <> ?", [DEFAULT_ACCOUNT])
    return bool(rows and rows[0][0])


def default_is_declared(store) -> bool:
    """Has anybody declared the row every install is given? (issue #725)

    The companion of :func:`accounts_are_declared`, and it is a different
    question: that one counts rows *beside* ``default``, this one asks whether
    ``default`` itself has stopped being what the schema seeded. Three roads
    reach that, and each of them is a declaration —

    * its owner **renamed** it, which is how an install with a page and no file
      declares its one account (the gesture #729 built the block for at N = 1);
    * its owner **retyped** it, the same clause on the other seeded column;
    * a **file** took it over (#698), which makes it a file's row like any other.

    It exists for the reassignment alone. *Events naming ``default``* and
    *events nobody assigned* are the same set only while nobody has declared
    that row, and reading them as one afterwards would take a whole ledger off
    the one line its owner had themselves put a name on.

    The comparison is :func:`as_declared`'s, not a second one: there is no
    *renamed* column to add to a row every install already has (ADR-0001 leaves
    no migration machinery), so what says *nobody declared this* is that it
    still says what the seed said. The other edge follows from the same
    sentence: an owner who names their account exactly what the seed named it
    is naming it, and this answers ``False`` — which costs them the offer and
    never a row.
    """
    row = next((a for a in read_accounts(store) if a.id == DEFAULT_ACCOUNT), None)
    if row is None:
        return False
    if row.source_id is not None:
        return True
    declared = as_declared(row)
    return declared.label is not None or declared.type is not None


def declared_portfolio(store) -> Optional[Portfolio]:
    """The declaration the app runs on, or ``None`` when nothing was declared.

    ``None`` is **ergonomics, not a discriminant** (ADR-0013): the store always
    holds at least one account, and every write path resolves an account without
    asking whether any were declared. What this answers is the narrower question
    the *pages* ask — is there a declaration to show — and it is why the seeded
    ``default`` row is left out of the list unless something actually names it:
    an install that declared two accounts should not grow a phantom third whose
    figures are all zero.
    """
    rows = read_accounts(store)
    declared = [a for a in rows if a.id != DEFAULT_ACCOUNT]
    if not declared:
        return None

    fallback = next((a for a in rows if a.id == DEFAULT_ACCOUNT), None)
    if fallback is not None and is_named_by_events(store, DEFAULT_ACCOUNT):
        declared.append(fallback)
    return Portfolio(accounts=declared)


def as_declared(account: Account) -> Account:
    """The row as a **reader** must see it: what nobody declared reads ``None``.

    The seed writes ``Default account`` / ``OTHER`` into a row every install
    owns and nobody asked for, and those two values are English the front must
    not render (ADR-0024) — so the interface names that row from its own
    catalogue until its owner gives it a name, and shows the name they gave the
    moment they do.

    **Which side recognises the seed is the whole decision.** The value is
    written here, in ``store.DEFAULT_ACCOUNT_ROW``, so it is recognised here:
    the wire carries ``null`` and the front folds *nothing declared* into its
    catalogue with no literal of its own. Recognising it in the client instead
    was written and undone — it put a **third copy** of a server-owned string on
    the far side of HTTP, where no compiler and no test spans both ends (the
    front's only faked edge is MSW, so its fixtures would have gone on agreeing
    with themselves). Reworded here for a typo, the seed would then have started
    rendering as a name its owner had typed, silently, every gate green — which
    is the exact failure the rule exists to prevent, arrived from the other side.

    Only the ``default`` row is concerned, and only while it still says what the
    seed said: an owner who names their own account ``Default account`` is
    naming it, and a comparison is what the store's lack of migration machinery
    leaves (ADR-0001) — there is no *renamed* column to add to a row every
    install already has.
    """
    if account.id != DEFAULT_ACCOUNT:
        return account
    _, seeded_type, seeded_label, _ = store_module.DEFAULT_ACCOUNT_ROW
    return replace(
        account,
        label=None if account.label == seeded_label else account.label,
        type=None if account.type == seeded_type else account.type,
    )


def is_named_by_events(store, account_id: str) -> bool:
    """Does any event name this account? The refusal's whole predicate."""
    rows = store.query(
        'SELECT count(*) FROM event WHERE account = ?', [account_id])
    return bool(rows and rows[0][0])


def source_account_ids(store, source_id: int) -> List[str]:
    """The ids an import declared, id-sorted."""
    return [row[0] for row in store.query(
        'SELECT id FROM account WHERE source_id = ? ORDER BY id', [source_id])]


# --------------------------------------------------------------------------- #
# Writing the table — the file's half
# --------------------------------------------------------------------------- #

def apply_source(store, source_id: int, rows: Sequence[AccountRow]) -> int:
    """Make ``source_id``'s accounts exactly ``rows``. Returns how many.

    Called **inside** :mod:`ledger`'s import transaction, which is what makes a
    refusal here a refusal of the whole file: the rollback takes the
    ``import_source`` row with it, so a bad accounts file is not imported at all
    (issue #698) and the previous declaration stands untouched.

    Three refusals, and each one is a mistake that would otherwise become a
    silent divergence:

    * the same id twice **in one file** — the PK would refuse it anyway, but not
      with a message naming both rows;
    * an id another source already declares, or one created in the app — two
      declarations of one account cannot both be the truth, and taking it over
      silently would make forgetting the other import delete a row it did not
      write;
    * an id this source declared before and no longer does, while an event names
      it (:class:`AccountInUse`) — the cascade ADR-0013 refuses.

    **A row that is already there is updated, never rewritten.** The upsert this
    replaced wrote the three columns unconditionally, ``source_id`` included, and
    that last one is why correcting a label and dropping the file again — the
    repair :class:`ReadOnlyAccount` tells the user to make — was impossible for
    every account an event named: DuckDB turns a write to a foreign-key column
    into a delete plus an insert, and the delete trips ``event.account``. The
    key is gone from the DDL for that reason (see :mod:`store`), and the write
    is split here for the other half of it: a re-drop of the same file finds its
    own ``import_source`` row, so ``source_id`` does not move, and the statement
    that does not move it is the one that says what actually changed.
    """
    _refuse_duplicates(rows)
    _refuse_taken(store, source_id, rows)

    incoming = {row.id for row in rows}
    declared = set(source_account_ids(store, source_id))
    retired = [i for i in declared if i not in incoming]
    _retire(store, retired)

    # Read after `_retire`: the ids it removed are gone, and an id this file now
    # declares that it did not declare before is an insert even when the row
    # existed a moment ago.
    existing = account_ids(store)
    for row in rows:
        if row.id in existing:
            store.execute(
                'UPDATE account SET type = ?, label = ? WHERE id = ?',
                [row.type, row.label, row.id])
            if row.id not in declared:
                # Ownership moves. `_refuse_taken` has already refused every id
                # another source or the app declares, so what reaches this line
                # is the seeded `default` row a file is allowed to take over.
                store.execute('UPDATE account SET source_id = ? WHERE id = ?',
                              [source_id, row.id])
            continue
        store.execute(
            'INSERT INTO account (id, type, label, source_id) '
            'VALUES (?, ?, ?, ?)', [row.id, row.type, row.label, source_id])
    return len(rows)


def forget_source(store, source_id: int) -> List[str]:
    """Drop the accounts an import declared. Returns the ids it removed.

    Raises:
        AccountInUse: an event names one of them. **Cascade forgetting is
            refused**, never performed: the alternative is an event pointing at
            an account that no longer exists, which is the orphan the whole
            design exists to make unrepresentable.
    """
    retired = source_account_ids(store, source_id)
    _retire(store, retired)
    return retired


def _refuse_duplicates(rows: Sequence[AccountRow]) -> None:
    seen = {}
    for row in rows:
        if row.id in seen:
            raise AccountSourceError(
                f"Account {row.id!r} is declared twice ({seen[row.id].where()} "
                f"and {row.where()}); an id names one account")
        seen[row.id] = row


def _refuse_taken(store, source_id: int, rows: Sequence[AccountRow]) -> None:
    """Refuse an id another source declares, or one created in the app."""
    owners = {
        row[0]: row[1] for row in store.query(
            'SELECT a.id, s.filename FROM account a '
            'LEFT JOIN import_source s ON s.id = a.source_id '
            'WHERE a.source_id IS DISTINCT FROM ?', [source_id])
    }
    for row in rows:
        if row.id not in owners:
            continue
        # `default` is the seeded row every install owns, not a declaration
        # somebody else made: a file is allowed to name it and take it over.
        if row.id == DEFAULT_ACCOUNT:
            continue
        by = f"by {owners[row.id]}" if owners[row.id] else "in the app"
        raise AccountSourceError(
            f"Account {row.id!r} ({row.where()}) is already declared {by}; "
            f"forget that declaration first, or rename this one")


def _retire(store, ids: Iterable[str]) -> None:
    """Remove accounts a source no longer declares — unless an event names one.

    **Every refusal is decided before the first deletion**, and that ordering is
    the whole correctness of the function: :func:`forget_source` is called
    outside a transaction by :func:`ledger.forget_import`, so raising halfway
    through the loop would commit the deletions already made and refuse the
    gesture — an import that is both refused and partly forgotten, with no way
    back (the file's fingerprint is unchanged, so a re-scan reports it
    unchanged and never restores what went). Checking first makes the refusal
    all-or-nothing without a transaction to lean on.

    ``default`` is never removed and never refused. It is the row every install
    has (ADR-0013: *there is always at least one account*), so a file that
    declared it hands it back rather than taking it away — and it hands back the
    **whole** row: ``type`` and ``label`` return to what the seed wrote, not to
    the file's own pair. Dropping the ownership alone would leave the install
    with an account nobody declares still wearing the name a forgotten file gave
    it, editable in the app and impossible to explain from anything the store
    still holds.
    """
    retiring = list(ids)

    removable = [i for i in retiring if i != DEFAULT_ACCOUNT]
    named = [i for i in removable if is_named_by_events(store, i)]
    if named:
        raise AccountInUse(
            f"Account(s) {', '.join(repr(i) for i in sorted(named))} cannot be "
            f"removed while an event names it; forget those events first")

    for account_id in retiring:
        if account_id == DEFAULT_ACCOUNT:
            _, seeded_type, seeded_label, _ = store_module.DEFAULT_ACCOUNT_ROW
            store.execute(
                'UPDATE account SET type = ?, label = ?, source_id = NULL '
                'WHERE id = ?', [seeded_type, seeded_label, DEFAULT_ACCOUNT])
            continue
        perf_series.forget_account(store, account_id)
        store.execute('DELETE FROM account WHERE id = ?', [account_id])


# --------------------------------------------------------------------------- #
# Writing the table — the app's half
# --------------------------------------------------------------------------- #

def create_account(store, account_id: str, account_type: str,
                   label: Optional[str] = None) -> Account:
    """Declare an account from the app. ``source_id`` stays ``NULL``.

    ``NULL`` is what "created in the UI" *is* (spec #695 § 6), and it is the
    same column that makes the row editable: what came from a file is read-only
    and revoked with its import, what was created here is neither.
    """
    account_id = _text(account_id)
    account_type = _text(account_type)
    if not account_id:
        raise AccountSourceError("id is required")
    if not account_type:
        raise AccountSourceError("type is required")
    if account_id in account_ids(store):
        raise DuplicateAccount(f"Account {account_id!r} already exists")

    store.execute(
        'INSERT INTO account (id, type, label, source_id) '
        'VALUES (?, ?, ?, NULL)',
        [account_id, account_type, _text(label) or account_id])
    logger.info(f"Declared account {account_id}")
    return Account(id=account_id, type=account_type,
                   label=_text(label) or account_id)


def update_account(store, account_id: str, *, account_type: Optional[str] = None,
                   label: Optional[str] = None) -> Account:
    """Relabel or retype an account created in the app.

    The id is not among what can change: it is the value events name, so
    changing it would be renaming every event that names it — which is an edit
    of imported rows, and imported rows are read-only.
    """
    current = _require(store, account_id)
    if current.source_id is not None:
        raise ReadOnlyAccount(
            f"Account {account_id!r} came from a file and is read-only; "
            f"correct the file and drop it again, or forget its import")

    new_type = _text(account_type) or current.type
    new_label = _text(label) or current.label
    store.execute('UPDATE account SET type = ?, label = ? WHERE id = ?',
                  [new_type, new_label, account_id])
    return Account(id=account_id, type=new_type, label=new_label)


def delete_account(store, account_id: str) -> None:
    """Remove an account created in the app.

    Three refusals, **in this order**: the ``default`` row (there is always at
    least one account), any account an event names (ADR-0013), and a row a file
    declared (forget the import instead).

    The middle one comes before the last on purpose. Both apply to a
    file-provisioned account an event names, and only one of them is actionable:
    *"forget its import"* is advice that would fail too, because forgetting is
    refused in cascade for exactly the same reason. Naming the event is what
    tells the user the gesture they actually have to make first.
    """
    current = _require(store, account_id)
    if account_id == DEFAULT_ACCOUNT:
        raise AccountInUse(
            "The default account is the one every install has, and it cannot "
            "be removed")
    if is_named_by_events(store, account_id):
        raise AccountInUse(
            f"Account {account_id!r} cannot be removed while an event names "
            f"it; forget those events first")
    if current.source_id is not None:
        raise ReadOnlyAccount(
            f"Account {account_id!r} came from a file; forget that import to "
            f"remove it")
    # The cached figures go with it (issue #700). ``account_metrics.account``
    # references this row, so the perf job's first cycle would otherwise make
    # every declared account undeletable — a constraint error the API renders as
    # a ``503``, in place of the ``200`` this gesture is designed to answer. The
    # series is a cache and rebuilds itself; the refusal that matters is the one
    # above, on an event, which cannot.
    perf_series.forget_account(store, account_id)
    store.execute('DELETE FROM account WHERE id = ?', [account_id])
    logger.info(f"Removed account {account_id}")


def _require(store, account_id: str) -> Account:
    for account in read_accounts(store):
        if account.id == account_id:
            return account
    raise UnknownAccount(f"No account with id {account_id!r}")


__all__ = [
    'ACCOUNT_COLUMNS', 'REQUIRED_ACCOUNT_COLUMNS',
    'AccountRow', 'AccountSourceError', 'AccountInUse', 'UnknownAccount',
    'DuplicateAccount', 'ReadOnlyAccount',
    'is_accounts_file', 'header_of', 'load_account_rows',
    'read_accounts', 'account_ids', 'accounts_are_declared',
    'default_is_declared', 'declared_portfolio',
    'is_named_by_events', 'source_account_ids',
    'apply_source', 'forget_source',
    'create_account', 'update_account', 'delete_account',
]
